# ── 最優先：process 啟動時就壓制 Windows 磁碟機未連線彈窗 ────────────
import sys as _sys
if _sys.platform == "win32":
    try:
        import ctypes as _ctypes
        # SEM_FAILCRITICALERRORS(0x0001) + SEM_NOOPENFILEERRORBOX(0x8000)
        # 壓制 H:/B: 磁碟機未就緒時的系統彈窗，讓 Python 直接收到 OSError
        _ctypes.windll.kernel32.SetErrorMode(0x0001 | 0x8000)
    except Exception:
        pass
    # stdout/stderr 改用 UTF-8，避免 emoji 在 cp950 下 UnicodeEncodeError 導致 thread 崩潰
    try:
        _sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        _sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    # 強制 stdout/stderr 使用 UTF-8，避免 emoji 在 cp950 終端機 crash
    try:
        _sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        _sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    # ── 防止 uvicorn reload 後 stdout/stderr 關閉導致 print crash ──
    import builtins as _builtins
    _original_print = _builtins.print
    def _safe_print(*args, **kwargs):
        try:
            _original_print(*args, **kwargs)
        except (ValueError, OSError):
            # I/O operation on closed file — 靜默忽略
            pass
    _builtins.print = _safe_print

import asyncio
import base64
import re
import secrets
import sqlite3
import uvicorn
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from fastapi import FastAPI, Request, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
from linebot.v3 import WebhookHandler
from linebot.v3.messaging import (
    Configuration, ApiClient, MessagingApi,
    ReplyMessageRequest, PushMessageRequest, TextMessage,
)
from linebot.v3.webhooks import MessageEvent, TextMessageContent, ImageMessageContent, VideoMessageContent
from linebot.v3.exceptions import InvalidSignatureError

from config import settings
from handlers.intent import detect_intent, Intent, CHECKOUT_KEYWORDS, AFFIRMATIVE_KEYWORDS
from handlers.inventory import handle_inventory
from handlers.orders import handle_order_tracking
from handlers.delivery import handle_delivery
from handlers.hours import handle_business_hours
from handlers.ordering import handle_order_quantity, handle_checkout, extract_quantity
from handlers.restock import handle_hq_reply
from handlers.internal import (
    handle_internal_arrival, handle_internal_order,
    handle_internal_image, handle_internal_order_from_state,
    handle_internal_notify_register, handle_internal_inventory,
    handle_internal_product_info, handle_internal_spec_query,
    handle_internal_tag_push,
    handle_internal_product_upload, handle_internal_save_images,
    handle_internal_add_images, handle_internal_save_text,
    handle_internal_mark_sold_out, handle_internal_unmark_sold_out,
    handle_internal_upload_start, handle_internal_upload_add_media,
    handle_internal_upload_text, handle_internal_upload_finish,
    handle_internal_upload_cancel,
    handle_ambiguous_resolve, handle_name_order_confirm, handle_new_customer_confirm,
    handle_internal_new_product,
    _split_new_product_entries,
    handle_internal_spec_inquiry, handle_spec_inquiry_reply, handle_spec_inquiry_qty,
    handle_internal_price_query,
    handle_internal_add_customer,
    handle_internal_product_info_by_name,
    handle_internal_consumable,
    handle_internal_rebate,
    handle_internal_rebate_push,
    handle_internal_set_rebate_target,
    handle_internal_unfulfilled,
    handle_internal_unclaimed,
    handle_internal_ready_for_pickup,
    handle_internal_ad_query,
    handle_internal_customer_orders,
    handle_internal_showcase_push,
    handle_internal_contact_group_push,
    handle_internal_recommend_push,
    handle_internal_product_photo,
    handle_internal_product_po_photo,
    handle_internal_competitor_price,
    handle_internal_label_queue,
    handle_internal_cart,
    _NEW_PROD_TRIGGER_RE,
    _PROD_CODE_RE,
    _SAVE_IMG_RE as _INTERNAL_SAVE_IMG_RE,
    _ADD_IMG_RE  as _INTERNAL_ADD_IMG_RE,
    _SOLD_OUT_RE as _INTERNAL_SOLD_OUT_RE,
    _RESTOCK_RE  as _INTERNAL_RESTOCK_RE,
    _UPLOAD_TRIGGERS as _INTERNAL_UPLOAD_TRIGGERS,
    _UPLOAD_FINISH_RE as _INTERNAL_UPLOAD_FINISH_RE,
    _UPLOAD_CANCEL_RE as _INTERNAL_UPLOAD_CANCEL_RE,
)
from handlers.ad_maker import handle_ad_update_trigger
from handlers.payment import is_payment_message, handle_payment
from handlers.price import handle_price
from handlers.summary import send_pending_summary, send_full_report, build_full_report, build_pending_text
from services.refresh import check_and_refresh
from handlers.escalate import handle_unknown
from handlers.service import (
    handle_bargaining, handle_spec, handle_return,
    handle_address_change, handle_complaint, handle_multi_product,
    handle_urgent_order, handle_image_product, handle_notify_request,
    detect_machine_query, handle_machine_query,
)
from handlers import tone
from handlers.visit import handle_visit, handle_visit_query, is_visit_query
from storage.state import state_manager
from storage.customers import customer_store
from storage.payments import payment_store
from storage.restock import restock_store
from storage.delivery import delivery_store
from storage.issues import issue_store
from storage.pending import pending_store
from storage.notify import notify_store
import storage.visits as visit_store
import storage.queue as queue_store

# 標記完成指令：支援單個、連續多個、範圍、全部
# 格式：✅ I2 I3 P1 / I1-I6已處理 / 全部已處理
_RESOLVE_TRIGGER_RE = re.compile(r"[✅☑️√v]|已處理|已完成|全部已處理")
_RESOLVE_ITEM_RE    = re.compile(r"([PRDIQprdiq])\s*(\d+)")
_RESOLVE_RANGE_RE   = re.compile(r"([PRDIQprdiq])\s*(\d+)\s*[-~–]\s*(?:[PRDIQprdiq]\s*)?(\d+)", re.IGNORECASE)
_RESOLVE_ALL_RE     = re.compile(r"全部已處理|已全部處理|全部\s*(?:已處理|完成|標記)")

# 內部群組呼叫 bot 名字的指令（到貨通知代客登記）
_BOT_NAME_RE = re.compile(r"^(新北小蠻牛|小蠻牛)\s*")
_PROD_CODE_RE_STAFF = re.compile(r"[A-Za-z][A-Za-z0-9\-]{2,}")
_STAFF_NOTIFY_KW = ["需要到貨通知", "要到貨通知", "到貨通知", "需要通知", "要通知"]

# ── 路徑常數 ──────────────────────────────────────────
_BASE_DIR   = Path(__file__).parent          # 專案根目錄（絕對路徑）
_ADMIN_HTML = _BASE_DIR / "static" / "admin.html"

# ── 機器人開關（admin 介面控制）──────────────────────
_bot_active: bool = True

# ── Server 啟動時間 ────────────────────────────────────
import time as _time_module
_server_start_time: float = _time_module.time()

# ── 最近看到的群組 ID（in-memory，重啟後清空）──────────
_seen_group_ids: set[str] = set()

# ── 未知群組追蹤（group_id -> last_seen timestamp）──────
_unknown_groups: dict[str, str] = {}

# ── Webhook 訊息去重（防止重複處理同一則訊息）──────────
_processed_msg_ids: dict[str, float] = {}  # message_id -> timestamp

# ── LINE profile 快取 ─────────────────────────────────
_profile_cache: dict[str, tuple[object, float]] = {}  # user_id -> (profile, timestamp)
_PROFILE_CACHE_TTL = 3600  # 1 hour

def _get_profile_cached(line_api, user_id: str):
    now = _time_module.time()
    cached = _profile_cache.get(user_id)
    if cached and now - cached[1] < _PROFILE_CACHE_TTL:
        return cached[0]
    try:
        profile = line_api.get_profile(user_id)
        _profile_cache[user_id] = (profile, now)
        return profile
    except Exception:
        if cached:
            return cached[0]
        return None

# ── 月額度 429 全域開關 ─────────────────────────────────────
_push_quota_exhausted = False        # True = 本月 push 額度已用完
_push_quota_exhausted_at: str = ""   # 記錄首次偵測到的時間
_pickup_notify_results: dict = {"notified_history": [], "no_line_id": []}  # 排程結果（供 admin 查看）

def _mark_push_exhausted():
    global _push_quota_exhausted, _push_quota_exhausted_at
    if not _push_quota_exhausted:
        _push_quota_exhausted = True
        _push_quota_exhausted_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        print(f"[quota] ⚠ 月額度用完，停止所有 push_message（{_push_quota_exhausted_at}）", flush=True)

def _is_quota_429(exc: Exception) -> bool:
    """檢查例外是否為月額度 429"""
    s = str(exc)
    return "429" in s and "monthly limit" in s.lower()

def _check_quota_reset():
    """每月 1 號自動重置額度開關"""
    global _push_quota_exhausted, _push_quota_exhausted_at
    if _push_quota_exhausted and _push_quota_exhausted_at:
        exhausted_month = _push_quota_exhausted_at[:7]  # "YYYY-MM"
        current_month = datetime.now().strftime("%Y-%m")
        if current_month != exhausted_month:
            _push_quota_exhausted = False
            _push_quota_exhausted_at = ""
            print(f"[quota] ✓ 新月份 {current_month}，額度開關已重置", flush=True)

_sync_failures: list[tuple[str, str, str]] = []  # [(時間, 任務名, 錯誤)]

def _notify_sync_failure(task_name: str, error: str):
    """記錄排程失敗，統一在 17:00 通知內部群"""
    _sync_failures.append((
        datetime.now().strftime("%H:%M"),
        task_name,
        error[:100],
    ))
    print(f"[sync-fail] 已記錄失敗：{task_name}", flush=True)

# ── 統一訊息合併緩衝（文字 + 圖片 + 影片，同一 timer）──────────
import threading as _threading

_MSG_COALESCE_SECS = 5.0           # 預設等待秒數（內部群）
_MSG_COALESCE_USER_SECS = 15.0    # 1對1客戶等待秒數
_MSG_UPLOAD_COALESCE_SECS = 15.0   # 上架指令等待秒數
_MSG_IMAGE_COALESCE_SECS = 8.0     # 有圖片 / 庫存問句等待秒數

# key: user_id
# value: {
#     "lines":        [str],           # 文字訊息
#     "media":        [{"msg_id": str, "type": str, "after_text": int}],
#     "context":      "user" | "group",
#     "group_id":     str | None,
#     "timer":        threading.Timer,
#     "reply_token":  str | None,
#     "quoted_msg_id": str | None,
# }
_msg_buffer: dict[str, dict] = {}
_msg_buffer_lock = _threading.Lock()
_user_flush_locks: dict[str, _threading.Lock] = {}
_user_flush_locks_lock = _threading.Lock()


# ── 已發送圖片 → 產品代碼對應表（客戶 tag 圖片時查詢，持久化到檔案）──
_SENT_IMAGE_MAP_FILE = Path(__file__).parent / "data" / "sent_image_map.json"
_sent_image_map: dict[str, dict] = {}  # message_id → {"code": "Z3031", "ts": 1234567890}
_sent_image_map_lock = _threading.Lock()
_SENT_IMAGE_MAP_TTL = 7 * 86400  # 保留 7 天


def _load_sent_image_map() -> None:
    """啟動時從檔案載入"""
    global _sent_image_map
    try:
        if _SENT_IMAGE_MAP_FILE.exists():
            import json as _json_sim
            data = _json_sim.loads(_SENT_IMAGE_MAP_FILE.read_text(encoding="utf-8"))
            # 清除超過 7 天的
            import time as _time_sim
            now = _time_sim.time()
            _sent_image_map = {
                k: v for k, v in data.items()
                if now - v.get("ts", 0) < _SENT_IMAGE_MAP_TTL
            }
            print(f"[sent_image_map] 載入 {len(_sent_image_map)} 筆（已清除過期）", flush=True)
    except Exception as e:
        print(f"[sent_image_map] 載入失敗: {e}", flush=True)

_load_sent_image_map()


def _save_sent_image_map() -> None:
    """存到檔案"""
    try:
        import json as _json_sim2
        _SENT_IMAGE_MAP_FILE.write_text(
            _json_sim2.dumps(_sent_image_map, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:
        print(f"[sent_image_map] 儲存失敗: {e}", flush=True)


def _store_sent_image_ids(sent_messages, image_codes: list[str]) -> None:
    """記錄 reply 回傳的 message_id → product_code"""
    import time as _time_store
    with _sent_image_map_lock:
        now = _time_store.time()
        for i, code in enumerate(image_codes):
            idx = i + 1  # 跳過第一則文字
            if idx < len(sent_messages):
                msg_id = sent_messages[idx].id
                _sent_image_map[msg_id] = {"code": code, "ts": now}
                print(f"[send_reply] 記錄圖片 msg_id={msg_id} → {code}", flush=True)
        # 清理超過 7 天的
        expired = [k for k, v in _sent_image_map.items() if now - v.get("ts", 0) > _SENT_IMAGE_MAP_TTL]
        for k in expired:
            _sent_image_map.pop(k, None)
        _save_sent_image_map()


def lookup_sent_image(msg_id: str) -> str | None:
    """查詢已發送圖片的產品代碼"""
    with _sent_image_map_lock:
        entry = _sent_image_map.get(msg_id)
        return entry["code"] if entry else None


# ── 客戶傳入文字 msg_id → text（客戶引用自己的文字時倒查用）──
_incoming_text_map: dict[str, dict] = {}  # msg_id → {"text": str, "ts": float}
_incoming_text_lock = _threading.Lock()
_INCOMING_TEXT_TTL = 86400  # 24h，in-memory 不持久化（引用超過一天很少見）


def _store_incoming_text(msg_id: str, text: str) -> None:
    import time as _t_it
    if not msg_id or not text:
        return
    with _incoming_text_lock:
        now = _t_it.time()
        _incoming_text_map[msg_id] = {"text": text, "ts": now}
        if len(_incoming_text_map) > 2000:
            cutoff = now - _INCOMING_TEXT_TTL
            expired = [k for k, v in _incoming_text_map.items() if v["ts"] < cutoff]
            for k in expired:
                _incoming_text_map.pop(k, None)


def lookup_incoming_text(msg_id: str) -> str | None:
    with _incoming_text_lock:
        entry = _incoming_text_map.get(msg_id)
        return entry["text"] if entry else None


def _send_reply(reply_token: str | None, to: str, text, line_api) -> bool:
    """
    優先用 reply_message（免費，不佔月額度），
    token 不存在或已過期才 fallback 到 push_message。
    支援 tuple: (text, image_urls) → 文字 + 圖片一起 reply
    回傳 True = 成功送出（reply 或 push 任一成功），False = 都失敗/額度用完。
    成功送出且 to 為客戶（U 開頭）時，自動寫 chat_history("bot", text)，
    避免「DB 有紀錄但 LINE 沒收到」的假紀錄。
    """
    from linebot.v3.messaging import ImageMessage

    image_urls = []
    image_codes = []
    if isinstance(text, tuple):
        text, image_urls = text[0], text[1]
        # 從 URL 提取產品代碼
        import re as _re_sr
        for url in image_urls:
            m = _re_sr.search(r'/([A-Za-z]{1,3}-?\d{3,6})[A-Z]?\.\w+$', url)
            if m:
                image_codes.append(m.group(1).upper())

    # LINE 訊息上限 5000 字元
    if len(text) > 4990:
        text = text[:4950] + "\n\n...（內容過長，已截斷）"

    # 組合訊息（文字 + 圖片，reply 最多 5 則）
    messages = [TextMessage(text=text)]
    for url in image_urls[:4]:  # 最多 4 張圖（文字佔 1 則，共 5 則上限）
        messages.append(ImageMessage(
            original_content_url=url, preview_image_url=url))

    def _log_bot_chat():
        if to and to.startswith("U") and text:
            try:
                from services.claude_ai import add_chat_history
                add_chat_history(to, "bot", text)
            except Exception as _ae:
                print(f"[send_reply] add_chat_history 失敗: {_ae}", flush=True)

    if reply_token:
        try:
            with ApiClient(_configuration) as _fresh_client:
                _fresh_api = MessagingApi(_fresh_client)
                resp = _fresh_api.reply_message(ReplyMessageRequest(
                    reply_token=reply_token,
                    messages=messages,
                ))
            # 記錄送出的圖片 message_id → product_code
            if image_codes and hasattr(resp, 'sent_messages') and resp.sent_messages:
                _store_sent_image_ids(resp.sent_messages, image_codes)
            # 記錄文字訊息中的貨號（客戶 tag 文字也能追蹤）
            elif not image_codes and hasattr(resp, 'sent_messages') and resp.sent_messages:
                import re as _re_txt_code
                _txt_codes = _PROD_CODE_RE.findall(text)
                if _txt_codes:
                    _first_code = _txt_codes[0].upper()
                    _txt_msg_id = resp.sent_messages[0].id
                    with _sent_image_map_lock:
                        _sent_image_map[_txt_msg_id] = {"code": _first_code, "ts": __import__('time').time()}
                        _save_sent_image_map()
                    print(f"[send_reply] 記錄文字 msg_id={_txt_msg_id} → {_first_code}", flush=True)
            _log_bot_chat()
            return True
        except Exception as _re:
            print(f"[send_reply] reply_message 失敗: {_re}", flush=True)
    else:
        print(f"[send_reply] 無 reply_token，直接 push to={to[:10]}...", flush=True)
    if _push_quota_exhausted:
        print(f"[send_reply] 月額度已用完，跳過 push", flush=True)
        return False
    try:
        resp = line_api.push_message(PushMessageRequest(
            to=to, messages=messages))
        # 跟 reply 路徑一致：記錄圖片 msg_id → product_code，文字裡的貨號也記
        if image_codes and hasattr(resp, 'sent_messages') and resp.sent_messages:
            _store_sent_image_ids(resp.sent_messages, image_codes)
        elif not image_codes and hasattr(resp, 'sent_messages') and resp.sent_messages:
            _txt_codes = _PROD_CODE_RE.findall(text)
            if _txt_codes:
                _first_code = _txt_codes[0].upper()
                _txt_msg_id = resp.sent_messages[0].id
                with _sent_image_map_lock:
                    _sent_image_map[_txt_msg_id] = {"code": _first_code, "ts": __import__('time').time()}
                    _save_sent_image_map()
                print(f"[send_reply] 記錄文字 msg_id={_txt_msg_id} → {_first_code} (push)", flush=True)
        _log_bot_chat()
        return True
    except Exception as _pe:
        if _is_quota_429(_pe):
            _mark_push_exhausted()
        else:
            print(f"[send_reply] push_message 也失敗: {_pe}", flush=True)
        return False


# (img_buffer / media_buf removed — merged into _msg_buffer above)

# ── 上架 Session 自動完成 timer（圖片傳完 30 秒無後續自動 finish）────
_UPLOAD_AUTO_FINISH_SECS = 15.0
_upload_finish_timers: dict[str, _threading.Timer] = {}
_upload_finish_timers_lock = _threading.Lock()

# ── 新增品項 Session timer（30 秒無後續自動處理）───────────────────────
_NP_SESSION_SECS = 30.0
_new_prod_timers: dict[str, _threading.Timer] = {}
_new_prod_timers_lock = _threading.Lock()


def _new_prod_auto_finish(user_id: str, group_id: str, reply_token: str | None) -> None:
    """30 秒無後續 → 自動處理累積的新增品項訊息"""
    with _new_prod_timers_lock:
        _new_prod_timers.pop(user_id, None)
    _state = state_manager.get(user_id)
    if not _state or _state.get("action") != "new_product_session":
        return
    lines = _state.get("lines", [])
    state_manager.clear(user_id)
    if not lines:
        return
    ack = handle_internal_new_product("\n".join(lines))
    if ack:
        _send_reply(reply_token, group_id, ack, _line_api)


def _new_prod_timer_reset(user_id: str, group_id: str, reply_token: str | None) -> None:
    """每次新訊息加入 session 就重置 30 秒 timer"""
    with _new_prod_timers_lock:
        old = _new_prod_timers.pop(user_id, None)
        if old:
            old.cancel()
        t = _threading.Timer(
            _NP_SESSION_SECS, _new_prod_auto_finish,
            args=(user_id, group_id, reply_token),
        )
        t.daemon = True
        _new_prod_timers[user_id] = t
    t.start()


def _new_prod_timer_cancel(user_id: str) -> None:
    """手動「完成」時取消 auto-finish timer"""
    with _new_prod_timers_lock:
        t = _new_prod_timers.pop(user_id, None)
        if t:
            t.cancel()


def _upload_auto_finish(session_id: str, group_id: str, reply_token: str | None) -> None:
    """15 秒無後續 → 自動執行 upload finish 並通知群組（需有至少 1 張圖片才執行）"""
    with _upload_finish_timers_lock:
        _upload_finish_timers.pop(session_id, None)
    # 確認有圖片，避免只有 PO 文就觸發
    _state = state_manager.get(session_id)
    if not _state:
        return
    _has_media = (
        len(_state.get("current_media", [])) > 0
        or any(len(g.get("media", [])) > 0 for g in _state.get("groups", []))
    )
    if not _has_media:
        print(f"[upload-auto] {session_id[:10]}... 無圖片，跳過 auto-finish", flush=True)
        return
    ack = handle_internal_upload_finish(session_id)
    if ack:
        _send_reply(reply_token, group_id, ack, _line_api)


def _upload_timer_reset(session_id: str, group_id: str, reply_token: str | None) -> None:
    """每次收到新媒體就重置 auto-finish timer"""
    with _upload_finish_timers_lock:
        old = _upload_finish_timers.pop(session_id, None)
        if old:
            old.cancel()
        t = _threading.Timer(
            _UPLOAD_AUTO_FINISH_SECS,
            _upload_auto_finish,
            args=(session_id, group_id, reply_token),
        )
        t.daemon = True
        _upload_finish_timers[session_id] = t
    t.start()


def _upload_timer_cancel(session_id: str) -> None:
    """手動說「完成」時取消 auto-finish timer"""
    with _upload_finish_timers_lock:
        t = _upload_finish_timers.pop(session_id, None)
        if t:
            t.cancel()


# (_media_buf_add / _media_buf_pop / _media_buf_flush / _txt_buffer removed — merged into _msg_buffer)


def _msg_buf_add(
    user_id: str,
    *,
    text: str | None = None,
    media_msg_id: str | None = None,
    media_type: str | None = None,   # "image" | "video"
    context: str,
    group_id: str | None = None,
    reply_token: str | None = None,
    quoted_msg_id: str | None = None,
) -> None:
    """
    統一訊息緩衝：文字 / 圖片 / 影片都走同一個 buffer + 單一 timer。
    任何新訊息重置 timer；timer 到期呼叫 _msg_buf_flush。
    """
    with _msg_buffer_lock:
        existing = _msg_buffer.get(user_id)
        if existing:
            existing["timer"].cancel()
            if text:
                existing["lines"].append(text)
                existing["line_quotes"].append(quoted_msg_id)
            if media_msg_id:
                _txt_count = len(existing.get("lines", []))
                existing["media"].append({
                    "msg_id": media_msg_id,
                    "type": media_type or "image",
                    "after_text": _txt_count,
                })
            if reply_token:
                existing["reply_token"] = reply_token
            if quoted_msg_id:
                existing["quoted_msg_id"] = quoted_msg_id
        else:
            _msg_buffer[user_id] = {
                "lines":        [text] if text else [],
                "line_quotes":  [quoted_msg_id] if text else [],
                "media":        [{"msg_id": media_msg_id, "type": media_type or "image", "after_text": 0}] if media_msg_id else [],
                "context":      context,
                "group_id":     group_id,
                "reply_token":  reply_token,
                "quoted_msg_id": quoted_msg_id,
            }
            existing = _msg_buffer[user_id]

        _type = "text" if text else f"media({media_type})"
        print(f"[msg-buf-add] user={user_id[:10]}... +{_type} lines={len(existing['lines'])} media={len(existing['media'])}", flush=True)

        # ── 決定 timer 秒數 ──
        all_text = "\n".join(existing["lines"])
        _is_upload_cmd = any(kw in all_text for kw in ("上架", "存圖", "加圖", "存文"))
        _may_have_image = any(k in all_text for k in ("有嗎", "有沒有", "還有貨", "有貨嗎", "還有嗎"))
        has_media = bool(existing["media"])
        has_text = bool(existing["lines"])

        _is_finish_cmd = _INTERNAL_UPLOAD_FINISH_RE.match(all_text.strip()) if all_text.strip() else False
        _is_cancel_cmd = _INTERNAL_UPLOAD_CANCEL_RE.match(all_text.strip()) if all_text.strip() else False
        if _is_finish_cmd or _is_cancel_cmd:
            wait_secs = 0.3  # 「完成/取消」立即處理
        elif _is_upload_cmd:
            wait_secs = _MSG_UPLOAD_COALESCE_SECS
        elif has_media and not has_text:
            _is_user_img = existing.get("context") == "user"
            wait_secs = _MSG_COALESCE_USER_SECS if _is_user_img else _MSG_IMAGE_COALESCE_SECS
        elif _may_have_image:
            wait_secs = _MSG_IMAGE_COALESCE_SECS
        else:
            _is_user = existing.get("context") == "user"
            wait_secs = _MSG_COALESCE_USER_SECS if _is_user else _MSG_COALESCE_SECS

        timer = _threading.Timer(wait_secs, _msg_buf_flush, args=(user_id,))
        timer.daemon = True
        existing["timer"] = timer

    timer.start()


def _handle_missing_ecount_name(text: str) -> str | None:
    """內部群指令「缺品名」：列出 specs 裡 Ecount 查不到品名的貨號"""
    if text.strip() not in ("缺品名", "無品名"):
        return None
    import json as _json
    from scripts.import_specs import OUTPUT
    from services.ecount import ecount_client
    try:
        specs = _json.loads(OUTPUT.read_text(encoding="utf-8")) if OUTPUT.exists() else {}
    except Exception:
        return "❌ 讀取 specs.json 失敗"
    ecount_client._ensure_product_cache()
    missing = []
    for code, s in specs.items():
        item = ecount_client.get_product_cache_item(code)
        if not item or not (item.get("name") or "").strip():
            spec_name = s.get("name", "")
            missing.append(f"  {code}：{spec_name}")
    if not missing:
        return "✅ 所有規格品項都有 Ecount 品名"
    return f"⚠️ Ecount 無品名（{len(missing)} 筆）：\n" + "\n".join(missing)


_ANALYTICS_RE = re.compile(r"^(分析報告|銷售排行|滯銷品|補貨預測|價位分析|品類分析|客戶分析|月趨勢|產品趨勢|庫存周轉|客戶流失|不叫貨|全部訂單|可通知客戶)$")
_NEW_PROD_SUGGEST_RE = re.compile(r"^新品建議\s+(.+?)\s+(\d+)元?$")


def _handle_analytics_command(text: str) -> str | None:
    t = text.strip()
    m = _ANALYTICS_RE.match(t)
    if m:
        from services.analytics import (
            full_report, top_sellers, slow_movers, restock_forecast,
            price_band_analysis, category_analysis, customer_analysis,
        )
        cmd = m.group(1)
        if cmd == "分析報告":
            return full_report()
        elif cmd == "銷售排行":
            top = top_sellers(30, 15)
            if not top:
                return "沒有銷售資料"
            lines = ["🔥 近30天銷售排行 TOP 15"]
            for i, p in enumerate(top, 1):
                lines.append(f"{i:2}. {p['code']} {p['name'][:18]} 出{p['total_out']}個 日均{p['daily_avg']} [{p['category']}]")
            return "\n".join(lines)
        elif cmd == "滯銷品":
            slow = slow_movers(60, 10)
            if not slow:
                return "沒有滯銷品"
            lines = [f"⚠️ 滯銷品（60天無出庫，庫存≥10）共 {len(slow)} 項"]
            for p in slow[:15]:
                lines.append(f"  {p['code']} {p['name'][:18]} 庫存{p['stock']} 上次{p['last_sale']} [{p['category']}]")
            return "\n".join(lines)
        elif cmd == "補貨預測":
            forecast = restock_forecast(30)
            if not forecast:
                return "沒有補貨預測資料"
            lines = ["🚨 補貨預測（按剩餘天數排序）"]
            for p in forecast[:15]:
                emoji = "🔴" if p["days_left"] <= 7 else "🟡" if p["days_left"] <= 14 else "🟢"
                lines.append(f"{emoji} {p['code']} {p['name'][:15]} 庫存{p['stock']} 日出{p['daily_avg_out']} 剩{p['days_left']}天")
            return "\n".join(lines)
        elif cmd == "價位分析":
            pb = price_band_analysis(90)
            if not pb:
                return "沒有價位資料"
            lines = ["💰 價位帶表現（近90天）"]
            for b in pb:
                top_name = b["top3"][0]["name"][:12] if b["top3"] else ""
                lines.append(f"  {b['band']:>8}  {b['product_count']}品  出{b['total_qty']}個  ${b['total_amount']:,}  冠軍:{top_name}")
            return "\n".join(lines)
        elif cmd == "品類分析":
            cats = category_analysis(90)
            if not cats:
                return "沒有品類資料"
            lines = ["📦 品類銷售（近90天）"]
            for c in cats:
                lines.append(f"  {c['category']:>10}  {c['product_count']}品  {c['pct']}%  ${c['total_amount']:,}")
            return "\n".join(lines)
        elif cmd == "客戶分析":
            custs = customer_analysis(90, 15)
            if not custs:
                return "沒有客戶資料"
            lines = ["👥 客戶 TOP 15（近90天）"]
            for c in custs:
                interval = f"每{c['avg_interval_days']}天" if c['avg_interval_days'] > 0 else "單次"
                lines.append(f"  {c['name'][:8]}  ${c['total_amount']:,}  {c['order_count']}次  {interval}  愛買:{c['fav_category']}")
            return "\n".join(lines)
        elif cmd == "月趨勢":
            from services.analytics import monthly_trend
            data = monthly_trend()
            if not data:
                return "沒有月趨勢資料"
            lines = ["📈 月銷售趨勢"]
            for d in data:
                lines.append(f"  {d['month']}  ${d['amount']:,}  {d['orders']}筆  {d['customers']}客")
            return "\n".join(lines)
        elif cmd == "產品趨勢":
            from services.analytics import product_trend
            data = product_trend(90)
            lines = ["📊 產品趨勢（近90天前半 vs 後半）"]
            if data["growing"]:
                lines.append("\n🚀 成長品：")
                for p in data["growing"][:5]:
                    lines.append(f"  {p['code']} {p['name'][:15]} {p['before']}→{p['after']} +{p['growth']}%")
            if data["declining"]:
                lines.append("\n📉 衰退品：")
                for p in data["declining"][:5]:
                    lines.append(f"  {p['code']} {p['name'][:15]} {p['before']}→{p['after']} -{p['decline']}%")
            if not data["growing"] and not data["declining"]:
                lines.append("沒有明顯成長或衰退的品項")
            return "\n".join(lines)
        elif cmd == "庫存周轉":
            from services.analytics import stock_turnover
            data = stock_turnover(90)
            if not data:
                return "沒有庫存周轉資料"
            lines = ["🔄 庫存周轉率 TOP 15（近90天，越高賣越快）"]
            for d in data[:15]:
                lines.append(f"  {d['code']} {d['name'][:15]} 出{d['total_out']} 庫{d['stock']} 周轉{d['turnover']}x")
            return "\n".join(lines)
        elif cmd == "客戶流失":
            from services.analytics import customer_churn
            data = customer_churn(60)
            if not data:
                return "沒有流失客戶"
            lines = [f"⚠️ 流失風險客戶（60天未下單，共{len(data)}位）"]
            for c in data[:15]:
                lines.append(f"  {c['name'][:8]}  ${c['total_amount']:,}  上次{c['last_order']}  已{c['inactive_days']}天")
            return "\n".join(lines)

        elif cmd == "全部訂單":
            from handlers.internal import _load_unfulfilled, _load_unclaimed
            from handlers.internal import _unfulfilled_needs_refresh, _refresh_unfulfilled
            from handlers.internal import _unclaimed_needs_refresh, _refresh_unclaimed
            from handlers.inventory import _check_preorder
            from services.rebate import _get_base_name

            if _unfulfilled_needs_refresh():
                _refresh_unfulfilled()
            if _unclaimed_needs_refresh():
                _refresh_unclaimed()

            uf = _load_unfulfilled()
            uc = _load_unclaimed()

            # 按客戶 base name 分組
            from collections import defaultdict as _dft
            uf_by = _dft(list)
            uc_by = _dft(list)
            for o in uf:
                uf_by[_get_base_name(o.get("customer", ""))].append(o)
            for o in uc:
                uc_by[_get_base_name(o.get("customer", ""))].append(o)

            all_custs = sorted(set(list(uf_by.keys()) + list(uc_by.keys())))

            ready = []    # 完全備好（只有未取，沒有未備貨，排除預購未備貨）
            pending = []  # 還有未備貨

            for cust in all_custs:
                uf_items = uf_by.get(cust, [])
                uc_items = uc_by.get(cust, [])
                # 非預購的未備貨
                non_po_uf = [o for o in uf_items if not _check_preorder(o.get("code", ""))]
                uc_count = len(uc_items)
                uf_count = len(non_po_uf)
                po_uf_count = len(uf_items) - len(non_po_uf)

                if uc_count > 0 and uf_count == 0:
                    po_note = f"（+預購{po_uf_count}筆）" if po_uf_count else ""
                    ready.append(f"  {cust}：{uc_count}筆未取{po_note}")
                elif uf_count > 0 or uc_count > 0:
                    parts = []
                    if uc_count: parts.append(f"{uc_count}筆未取")
                    if uf_count: parts.append(f"{uf_count}筆未備貨")
                    if po_uf_count: parts.append(f"{po_uf_count}筆預購未備")
                    pending.append(f"  {cust}：{'、'.join(parts)}")

            lines = [f"📋 全部訂單（{len(all_custs)}位客戶）"]
            lines.append(f"\n✅ 完全備好可取貨（{len(ready)}位）：")
            if ready:
                lines.extend(ready)
            else:
                lines.append("  （無）")
            lines.append(f"\n⏳ 還有未備貨（{len(pending)}位）：")
            if pending:
                lines.extend(pending)
            else:
                lines.append("  （無）")
            return "\n".join(lines)

        elif cmd == "可通知客戶":
            from handlers.internal import _load_unfulfilled, _load_unclaimed
            from handlers.internal import _unfulfilled_needs_refresh, _refresh_unfulfilled
            from handlers.internal import _unclaimed_needs_refresh, _refresh_unclaimed
            from handlers.inventory import _check_preorder
            from services.rebate import _get_base_name
            from collections import defaultdict as _dft2

            if _unfulfilled_needs_refresh():
                _refresh_unfulfilled()
            if _unclaimed_needs_refresh():
                _refresh_unclaimed()

            uf = _load_unfulfilled()
            uc = _load_unclaimed()

            uf_by = _dft2(list)
            uc_by = _dft2(list)
            for o in uf:
                uf_by[_get_base_name(o.get("customer", ""))].append(o)
            for o in uc:
                uc_by[_get_base_name(o.get("customer", ""))].append(o)

            ready = []
            for cust in sorted(set(list(uf_by.keys()) + list(uc_by.keys()))):
                uc_items = uc_by.get(cust, [])
                uf_items = uf_by.get(cust, [])
                non_po_uf = [o for o in uf_items if not _check_preorder(o.get("code", ""))]
                po_uf_count = len(uf_items) - len(non_po_uf)

                if uc_items and not non_po_uf:
                    po_note = f"（+預購{po_uf_count}筆）" if po_uf_count else ""
                    ready.append(f"  {cust}：{len(uc_items)}筆{po_note}")

            if not ready:
                return "目前沒有可通知的客戶（都還有未備貨品項）"
            lines = [f"📩 可通知取貨（{len(ready)}位）："]
            lines.extend(ready)
            return "\n".join(lines)

        elif cmd == "不叫貨":
            from services.analytics import do_not_restock
            data = do_not_restock()
            if not data:
                return "目前沒有建議不叫貨的品項"
            lines = [f"🚫 不建議叫貨（共{len(data)}品）"]
            for d in data[:15]:
                lines.append(f"  {d['code']} {d['name'][:15]} 庫{d['stock']} | {d['reasons']}")
            return "\n".join(lines)

    m = _NEW_PROD_SUGGEST_RE.match(t)
    if m:
        from services.analytics import new_product_suggestion
        cat, price = m.group(1), int(m.group(2))
        s = new_product_suggestion(cat, price)
        lines = [f"📋 新品建議：{cat} {price}元"]
        lines.append(f"同品類 {s['same_category_count']} 個品項，月均銷量 {s['same_category_avg_monthly']} 個")
        lines.append(f"同價位 {s['same_priceband_count']} 個品項，月均銷量 {s['same_priceband_avg_monthly']} 個")
        lines.append(f"\n💡 建議首批進貨：{s['suggested_qty']} 個")
        if s["top_in_category"]:
            lines.append(f"\n同品類熱銷：")
            for p in s["top_in_category"][:3]:
                lines.append(f"  {p['code']} {p['name'][:15]} {p['price']}元 銷{p['qty']}個")
        return "\n".join(lines)

    # ── 採購建議（貨號或PO文）──
    if t.startswith("採購建議"):
        content = t.replace("採購建議", "").strip()
        if not content:
            return "請輸入貨號或 PO 文，例如：\n採購建議 P0154\n或：\n採購建議\nP0154\nT1234"
        codes = _PROD_CODE_RE.findall(content)
        if codes and len(content.replace("\n", " ").split()) <= len(codes) + 2:
            return _analyze_purchase_by_codes(codes)
        else:
            return _analyze_purchase(content)

    return None


def _analyze_purchase_by_codes(codes: list[str]) -> str:
    """用貨號自動找 PO 文 + 圖片，然後分析"""
    from handlers.internal import _format_po, _get_raw_po_block
    from services.ecount import ecount_client

    results = []
    for raw_code in codes:
        code = raw_code.upper()
        # 找 PO 文
        po_block = _get_raw_po_block(code)
        if not po_block:
            # 沒有 PO 文，用 Ecount 品名
            item = ecount_client.get_product_cache_item(code)
            name = (item.get("name") if item else None) or code
            po_block = f"品名：{name}\n編號：{code}"

        # 找圖片
        has_photo = False
        for suffix in ["A.jpg", "B.jpg", ".jpg"]:
            if (_PRODUCT_PHOTO_DIR / f"{code}{suffix}").exists():
                has_photo = True
                break

        result = _analyze_purchase(po_block)
        if has_photo:
            result += f"\n📷 有產品照片"
        results.append(result)

    return "\n\n{'='*30}\n\n".join(results) if results else "找不到任何產品資料"


def _analyze_purchase(po_text: str) -> str:
    """用 Claude 分析師角色 + 銷售數據，建議採購數量"""
    from services.analytics import (
        _classify, top_sellers, price_band_analysis,
        category_analysis, new_product_suggestion,
    )
    import re as _re

    # 從 PO 文提取品名和價格
    name_m = _re.search(r'品名[：:]\s*(.+)', po_text)
    price_m = _re.search(r'(?:價格|售價|批發價|優惠價)[：:]*\s*(\d+)\s*元?', po_text)
    prod_name = name_m.group(1).strip() if name_m else po_text.split('\n')[0][:20]
    price = int(price_m.group(1)) if price_m else 0

    # 用整段 PO 文分類（品名可能跨行，如「品名：最強品牌\nC3膠囊行動電源」）
    category = _classify(po_text)
    if category == "其他":
        category = _classify(prod_name)

    # 收集分析數據
    data_parts = []

    # 本品自身歷史銷量（最重要！已經賣過的就用實際數據）
    _own_code = _re.search(r'[A-Za-z]{1,3}-?\d{3,6}', po_text)
    if _own_code:
        _own_cd = _own_code.group(0).upper()
        import sqlite3 as _sq_own
        _conn_own = _sq_own.connect(str(Path(__file__).parent / "data" / "sales_detail.db"))
        _own_rows = _conn_own.execute(
            "SELECT substr(date,1,7) as m, SUM(qty), SUM(amount), COUNT(DISTINCT customer) "
            "FROM sales_detail WHERE prod_cd=? AND customer NOT LIKE '%民享%' GROUP BY m ORDER BY m",
            (_own_cd,)
        ).fetchall()
        _own_total = _conn_own.execute(
            "SELECT SUM(qty), SUM(amount), COUNT(DISTINCT customer) "
            "FROM sales_detail WHERE prod_cd=? AND customer NOT LIKE '%民享%'",
            (_own_cd,)
        ).fetchone()
        _conn_own.close()
        if _own_rows:
            months_str = "\n".join(f"  {r[0]}: 出{int(r[1])}個 ${int(r[2]):,} ({int(r[3])}位客戶)" for r in _own_rows)
            data_parts.append(
                f"【⚠️ 此產品本身的歷史銷量（最重要！）】\n"
                f"貨號：{_own_cd}\n"
                f"總計：出{int(_own_total[0])}個 ${int(_own_total[1]):,} {int(_own_total[2])}位客戶\n"
                f"月別明細：\n{months_str}\n"
                f"月均：{int(_own_total[0]) // max(len(_own_rows), 1)}個"
            )

    # 同品類分析
    if category != "其他":
        s = new_product_suggestion(category, price)
        # 列出所有同品類品項的完整銷量（讓 Claude 看到分佈）
        all_items_str = "\n".join(
            f"  {p['name'][:20]} {p['price']}元 → 銷{p['qty']}個" for p in s['top_in_category']
        )
        data_parts.append(
            f"【同品類「{category}」全部品項銷量（近3個月）】\n"
            f"品項數：{s['same_category_count']} 個\n"
            f"月均銷量：{s['same_category_avg_monthly']} 個/品\n"
            f"各品銷量明細：\n{all_items_str}"
        )
    if price > 0:
        s2 = new_product_suggestion(category if category != "其他" else "行動電源", price)
        data_parts.append(
            f"【同價位帶分析】\n"
            f"品項數：{s2['same_priceband_count']} 個\n"
            f"月均銷量：{s2['same_priceband_avg_monthly']} 個/品"
        )

    # 同名稱/關鍵字的歷史銷量（找類似產品）
    import sqlite3
    _db_path = Path(__file__).parent / "data" / "sales_detail.db"
    if _db_path.exists():
        _conn = sqlite3.connect(str(_db_path))
        # 從 PO 文提取關鍵字（品名+Ecount品名優先，再補 PO 文其他行）
        _STOP_WORDS = {"原裝", "精品", "公司", "正版", "品質", "保障", "產品", "包裝", "尺寸",
                       "重量", "價格", "現貨", "不多", "建議", "標準", "編號", "品名", "附贈",
                       "精美", "海報", "起批", "凸面", "印刷", "超質", "輕巧", "實用", "設計",
                       "快速", "充電", "合適", "大貨", "盒起", "一張", "黃金", "金級", "級吸",
                       "立馬", "馬吸", "吸淨", "吸力", "強勁", "動力", "高效", "續航",
                       "渦輪", "核心", "極效"}
        # 優先用 Ecount 品名 + PO文品名行
        _ecount_name = ""
        _code_m = _re.search(r'[A-Za-z]{1,3}-?\d{3,6}', po_text)
        if _code_m:
            from services.ecount import ecount_client as _ec_kw
            _ec_item = _ec_kw.get_product_cache_item(_code_m.group(0).upper())
            if _ec_item:
                _ecount_name = _ec_item.get("name", "")
        _priority_text = f"{_ecount_name} {prod_name}"
        _full_text = f"{_priority_text} {po_text}"
        _cn_chars = _re.findall(r'[\u4e00-\u9fff]+', _priority_text)
        _cn_chars += _re.findall(r'[\u4e00-\u9fff]+', po_text)
        _keywords = []
        for seg in _cn_chars:
            if len(seg) >= 2:
                for i in range(len(seg) - 1):
                    w = seg[i:i+2]
                    if w not in _keywords and w not in _STOP_WORDS:
                        _keywords.append(w)
        _similar = []
        for kw in _keywords[:8]:
            rows = _conn.execute(
                "SELECT prod_cd, prod_name, SUM(qty) as total, unit_price "
                "FROM sales_detail WHERE prod_name LIKE ? AND customer NOT LIKE '%民享%' "
                "GROUP BY prod_cd ORDER BY total DESC LIMIT 5",
                (f"%{kw}%",)
            ).fetchall()
            for r in rows:
                if r[0] not in [s[0] for s in _similar]:
                    _similar.append(r)
        if _similar:
            sim_str = "\n".join(
                f"  {r[1][:20]} ({r[0]}) {int(r[3])}元 → 實際銷{int(r[2])}個" for r in _similar[:8]
            )
            data_parts.append(
                f"【⚠️ 同類型/同名稱產品的實際歷史銷量（最重要參考）】\n{sim_str}"
            )
        _conn.close()

    # 同品類庫存水平
    import json as _j_ap
    _avail_path_ap = Path(__file__).parent / "data" / "available.json"
    if _avail_path_ap.exists():
        _avail_ap = _j_ap.loads(_avail_path_ap.read_text(encoding="utf-8"))
        stock_levels = []
        for p in s.get('top_in_category', []) if category != "其他" else []:
            d = _avail_ap.get(p['code'])
            if isinstance(d, dict):
                stock_levels.append(f"  {p['name'][:15]} → 目前庫存{d.get('available',0)}個")
        if stock_levels:
            data_parts.append(
                f"【同品類目前庫存水平（參考進貨量）】\n" + "\n".join(stock_levels)
            )

    # 品類整體表現
    cats = category_analysis(90)
    cat_data = next((c for c in cats if c["category"] == category), None)
    if cat_data:
        data_parts.append(
            f"【品類整體表現（近90天）】\n"
            f"品類：{category}\n"
            f"銷售佔比：{cat_data['pct']}%\n"
            f"品項數：{cat_data['product_count']}\n"
            f"總出貨：{cat_data['total_qty']} 個\n"
            f"總金額：${cat_data['total_amount']:,}"
        )

    # 價位帶表現
    pb = price_band_analysis(90)
    if price > 0:
        for b in pb:
            band = b["band"]
            if ((band == "50以下" and price <= 50) or
                (band == "51-100" and 51 <= price <= 100) or
                (band == "101-150" and 101 <= price <= 150) or
                (band == "151-200" and 151 <= price <= 200) or
                (band == "201-300" and 201 <= price <= 300) or
                (band == "300以上" and price > 300)):
                data_parts.append(
                    f"【價位帶 {band} 表現】\n"
                    f"品項數：{b['product_count']}\n"
                    f"平均每品出貨：{b['avg_qty_per_product']} 個\n"
                    f"冠軍：{b['top3'][0]['name'][:15] if b['top3'] else 'N/A'}"
                )
                break

    # 用 Google 搜蝦皮市場價格
    try:
        import httpx
        _search_name = _re.sub(r'[（）\(\)\[\]【】\(\)]', ' ', prod_name).strip()
        # 取品名核心詞（去掉大/原等前綴）
        _core_name = _re.sub(r'^[\(（]?[大原小][\)）]?\s*', '', _search_name).strip()
        _search_q = f"site:shopee.tw {_core_name}"
        _resp = httpx.get(
            "https://html.duckduckgo.com/html/",
            params={"q": _search_q},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=10,
        )
        # 從搜尋結果提取價格（$數字 或 數字元）
        _prices = _re.findall(r'\$\s*([\d,]+)', _resp.text)
        _prices += _re.findall(r'(?<!\d)([\d,]{2,6})\s*元', _resp.text)
        # 過濾合理價格範圍（進價的 0.5~5 倍）
        _min_p = max(price * 0.5, 30) if price > 0 else 30
        _max_p = max(price * 5, 2000) if price > 0 else 5000
        _price_nums = sorted(set(
            int(p.replace(",", "")) for p in _prices
            if _min_p < int(p.replace(",", "")) < _max_p
        ))
        if _price_nums:
            _mid = _price_nums[len(_price_nums)//2]
            data_parts.append(
                f"【蝦皮市場零售價（Google 搜尋）】\n"
                f"搜尋：{_core_name}\n"
                f"找到 {len(_price_nums)} 個價格：${min(_price_nums):,} ~ ${max(_price_nums):,}\n"
                f"中位數：${_mid:,}　進價：${price:,}　利潤空間：{round((_mid-price)/price*100) if price else '?'}%"
            )
    except Exception:
        pass

    analytics_context = "\n\n".join(data_parts)

    # 呼叫 Claude 分析師
    from services.claude_ai import _CLAUDE_CMD, _TIMEOUT
    import subprocess, os

    prompt = f"""你是台灣頂尖的娃娃機商品採購分析師，擁有 10 年以上經驗。你的分析必須 100% 數據導向，不猜測、不灌水。
根據以下的銷售數據和PO文，給出精準的採購建議。

【新品PO文】
{po_text}

【銷售數據分析】
{analytics_context}

請分析：
0. 市場零售價對比（如有提供蝦皮搜尋結果，分析進價 vs 零售價的利潤空間）
1. 這個品類和價位的市場表現如何
2. 同類產品的銷售速度（重點看熱銷品的實際銷量，不要只看平均值）
3. 建議首批進貨數量（具體數字，參考同品類+同價位帶的熱銷品銷量，不要太保守）
4. 風險評估（高/中/低）
5. 一句話建議

重要：
- 如果有「此產品本身的歷史銷量」，這是最最重要的參考，直接用月均銷量推算
- 如果是全新產品（無自身銷量），參考「同名稱/同類型產品的歷史銷量」
- 如果連同名的都沒有，才用同品類平均值
- 建議數量要合理：已經賣過的產品，不要建議超出歷史月均的 2 倍
- 你是頂級採購分析師，分析要精準、數據導向、不說廢話
回覆格式要簡潔清楚，用繁體中文，適合在 LINE 群組閱讀。"""

    try:
        env = {**os.environ, "PYTHONIOENCODING": "utf-8"}
        result = subprocess.run(
            [_CLAUDE_CMD, "-p", "-", "--tools", ""],
            input=prompt.encode("utf-8"),
            capture_output=True, timeout=120, env=env,
            cwd="C:\\Users\\bear\\AppData\\Local\\Temp",
        )
        answer = result.stdout.decode("utf-8", errors="replace").strip()
        if answer and result.returncode == 0:
            return f"🧠 採購分析報告\n{'='*20}\n{answer}"
    except Exception as e:
        print(f"[purchase-analysis] Claude 分析失敗: {e}", flush=True)

    # Claude 失敗 fallback：用程式算
    if category != "其他" and price > 0:
        s = new_product_suggestion(category, price)
        return (
            f"🧠 採購建議（自動）\n"
            f"品類：{category} | 價位：{price}元\n"
            f"同品類 {s['same_category_count']} 品，月均 {s['same_category_avg_monthly']} 個\n"
            f"💡 建議首批：{s['suggested_qty']} 個"
        )
    return "⚠️ 無法分析，請確認 PO 文包含品名和價格"


def _dispatch_internal_fallback(combined: str, group_id: str, line_api, staff_id: str = "") -> str | None:
    """內部群組 fallback dispatch chain"""
    # 含 @ tag 的訊息是對話，不需要 bot 回覆
    if "@" in combined:
        return None
    return (
        _handle_analytics_command(combined)
        or handle_spec_inquiry_reply(group_id, combined, line_api)
        or handle_spec_inquiry_qty(group_id, combined, line_api)
        or handle_ad_update_trigger(combined, group_id, line_api)
        or _handle_pending_list_command(combined)
        or _handle_staff_resolve(combined)
        or _handle_visit_resolve(combined)
        or _handle_visit_query_command(combined)
        or _handle_spec_rebuild_command(combined)
        or _handle_bot_notify_command(combined)
        or handle_internal_showcase_push(combined, line_api)
        or handle_internal_contact_group_push(combined, line_api)
        or handle_internal_recommend_push(combined, line_api)
        or handle_internal_label_queue(combined)
        or handle_internal_tag_push(combined, line_api)
        or handle_internal_add_customer(combined)
        or handle_internal_notify_register(combined, line_api)
        or handle_internal_arrival(combined, line_api)
        or handle_ambiguous_resolve(group_id, combined)
        or handle_name_order_confirm(group_id, combined)
        or handle_new_customer_confirm(group_id, combined)
        or handle_internal_cart(combined, line_api, staff_id=staff_id)
        or handle_internal_order(combined, line_api, group_id=group_id)
        or handle_internal_set_rebate_target(combined)
        or handle_internal_rebate_push(combined, line_api)
        or handle_internal_rebate(combined, group_id)
        or handle_internal_unfulfilled(combined, group_id)
        or handle_internal_unclaimed(combined, group_id)
        or handle_internal_ready_for_pickup(combined, group_id)
        or handle_internal_ad_query(combined, group_id)
        or handle_internal_customer_orders(combined, group_id)
        or handle_internal_consumable(combined, group_id)
        or handle_internal_competitor_price(combined)
        or handle_internal_product_po_photo(combined, line_api)
        or handle_internal_product_photo(combined, line_api)
        or handle_internal_spec_query(combined)
        or handle_internal_product_info(combined, group_id)
        or handle_internal_price_query(combined)
        or handle_internal_inventory(combined, group_id)
        or handle_internal_spec_inquiry(combined, group_id)
        or handle_internal_product_info_by_name(combined, group_id)
    )


def _flush_media_only(
    user_id: str,
    entry: dict,
    _all_media: list,
    img_e: dict | None,
    media_e: dict | None,
) -> None:
    """無文字、只有媒體時的處理（取代舊 _img_buf_flush + _media_buf_flush）"""
    context     = entry["context"]
    group_id    = entry.get("group_id")
    reply_token = entry.get("reply_token")
    line_api    = _line_api

    if context == "group":
        # ── 如果有「補圖/加圖/存圖」等待中 → 直接處理 ──
        # 先查 group_id 的 state（不同人傳圖也能觸發），再查 user_id
        _st = state_manager.get(group_id) if group_id else None
        if not _st or _st.get("action") not in ("pending_add_img", "pending_save_img"):
            _st = state_manager.get(user_id)
        if _st and _st.get("action") == "pending_add_img":
            _add_code = _st["prod_code"]
            _add_gid = _st.get("group_id", group_id)
            state_manager.clear(group_id)
            state_manager.clear(user_id)
            ack = handle_internal_add_images(_add_code, _all_media)
            if ack:
                _send_reply(reply_token, _add_gid, ack, line_api)
            return
        if _st and _st.get("action") == "pending_save_img":
            _save_code = _st["prod_code"]
            _save_gid = _st.get("group_id", group_id)
            state_manager.clear(group_id)
            state_manager.clear(user_id)
            ack = handle_internal_save_images(_save_code, _all_media)
            if ack:
                _send_reply(reply_token, _save_gid, ack, line_api)
            return
        # ── upload session 進行中（per-group）→ 直接追加 ──
        # 先查 group_id 的上架 session
        _upload_st = state_manager.get(group_id) if group_id else None
        if _upload_st and _upload_st.get("action") == "uploading":
            for m in _all_media:
                handle_internal_upload_add_media(group_id, m["msg_id"], m["type"])
            _upload_timer_reset(group_id, group_id, reply_token)
            print(f"[msg-buf] upload session 進行中，媒體已追加", flush=True)
            return
        # ── 單張圖片 → 商品識別；多張/含影片 → 提示 ──
        if len(_all_media) == 1 and _all_media[0]["type"] == "image":
            reply_text = handle_internal_image(user_id, _all_media[0]["msg_id"], line_api)
        else:
            n = len(_all_media)
            reply_text = (
                f"收到 {n} 個檔案，請補上指令：\n"
                f"• 上架（圖片 + PO文一起存）\n"
                f"• 存圖 Z3432（只存圖片）"
            )
        _send_reply(reply_token, group_id, reply_text, line_api)
    else:
        # 1:1 客戶
        # 只取第一張圖片做商品識別
        # 先辨識，存結果到 state，再等 5 秒看有沒有文字跟上
        if img_e:
            _img_mid = img_e.get("msg_id")
            _identified = _img_identify_from_buf(_img_mid) if _img_mid else None
            if _identified:
                print(f"[msg-buf] 圖片先到，辨識出 {_identified}，等文字 5 秒", flush=True)
                state_manager.set(user_id, {
                    "action":   "image_waiting_text",
                    "prod_cd":  _identified,
                    "msg_id":   _img_mid,
                    "reply_token": reply_token,
                })
                # 5 秒後如果沒文字，直接回覆
                def _delayed_image_reply(_uid=user_id, _mid=_img_mid, _rt=reply_token):
                    _st = state_manager.get(_uid)
                    if _st and _st.get("action") == "image_waiting_text":
                        state_manager.clear(_uid)
                        _reply = handle_image_product(_uid, _mid, _line_api)
                        _send_reply(_rt, _uid, _reply, _line_api)
                _t = _threading.Timer(5.0, _delayed_image_reply)
                _t.daemon = True
                _t.start()
            else:
                # 辨識失敗 → 直接走完整流程（含 Claude 圖片辨識）
                reply_text = handle_image_product(user_id, img_e["msg_id"], line_api)
                _send_reply(reply_token, user_id, reply_text, line_api)


def _msg_buf_flush(user_id: str) -> None:
    """per-user 鎖 wrapper：確保同一客戶的 flush 不會並行"""
    with _user_flush_locks_lock:
        if user_id not in _user_flush_locks:
            _user_flush_locks[user_id] = _threading.Lock()
        user_lock = _user_flush_locks[user_id]
    # trylock：如果已在處理中，不排隊等待（讓新訊息繼續累積在 buffer 裡）
    acquired = user_lock.acquire(blocking=False)
    if not acquired:
        # 前一則還在處理，把這次的 timer 延後 3 秒再試
        _t = _threading.Timer(3.0, _msg_buf_flush, args=(user_id,))
        _t.daemon = True
        _t.start()
        return
    try:
        _msg_buf_flush_inner(user_id)
    finally:
        user_lock.release()

# keep old name as alias so inspect.getsource(_txt_buf_flush) still works
_txt_buf_flush = _msg_buf_flush


def _msg_buf_flush_inner(user_id: str) -> None:
    # 如果前一則正在處理中，新訊息可能剛加進來，先檢查 buffer 有沒有被更新
    # （防止 per-user lock 排隊導致每則分開處理）
    """
    Timer 到期後觸發：合併所有緩衝文字 + 媒體，統一處理並回覆。
    """
    with _msg_buffer_lock:
        entry = _msg_buffer.pop(user_id, None)
    if not entry:
        return

    combined       = "\n".join(entry["lines"]) if entry["lines"] else ""
    context        = entry["context"]
    group_id       = entry.get("group_id")
    reply_token    = entry.get("reply_token")
    quoted_msg_id  = entry.get("quoted_msg_id")

    # ── 從統一 buffer 建立相容變數（img_e / media_e）──────────────────
    _all_media = entry.get("media", [])
    _images    = [m for m in _all_media if m["type"] == "image"]
    img_e = (
        {
            "msg_id":    _images[0]["msg_id"],
            "msg_ids":   [i["msg_id"] for i in _images],
            "context":   context,
            "group_id":  group_id,
            "reply_token": reply_token,
        }
        if _images else None
    )
    media_e = (
        {
            "media":       _all_media,
            "group_id":    group_id,
            "reply_token": reply_token,
        }
        if _all_media else None
    )

    print(f"[msg-buf-flush] user={user_id[:10]}... text={len(entry['lines'])}行 media={len(_all_media)}個 img_e={'有' if img_e else '無'} context={context}", flush=True)

    # ── 無文字、只有媒體 → 走「媒體獨立處理」路徑 ──────────────────
    if not combined and _all_media:
        _flush_media_only(user_id, entry, _all_media, img_e, media_e)
        return

    line_api = _line_api

    if True:  # preserve indentation block
        # ── 回覆輔助：優先 reply_message，失敗才 push（共用 _send_reply）──
        _reply_token_used = [False]

        def _send_group_ack(text_or_tuple) -> None:
            nonlocal reply_token
            from linebot.v3.messaging import ImageMessage as _ImgMsg
            # 支援 tuple: (text, [image_urls])
            image_urls = []
            if isinstance(text_or_tuple, tuple):
                text, image_urls = text_or_tuple[0], text_or_tuple[1]
            else:
                text = text_or_tuple
            if len(text) > 4990:
                text = text[:4950] + "\n\n...（內容過長，已截斷）"
            messages = [TextMessage(text=text)]
            for url in image_urls[:4]:
                messages.append(_ImgMsg(original_content_url=url, preview_image_url=url))
            if reply_token and not _reply_token_used[0]:
                try:
                    # 每次建新的 client 避免連線過期
                    with ApiClient(_configuration) as _fresh_client:
                        _fresh_api = MessagingApi(_fresh_client)
                        _fresh_api.reply_message(ReplyMessageRequest(
                            reply_token=reply_token,
                            messages=messages[:5],
                        ))
                    _reply_token_used[0] = True
                    return
                except Exception as _reply_err:
                    print(f"[txt-buf] reply_message 失敗: {_reply_err}", flush=True)
                    reply_token = None  # 標記 token 已失效
            print(f"[txt-buf] 內部群 reply 失敗，不回應", flush=True)

        # ════ 內部群組 ════
        if context == "group":
            # 上架 session：per-group（認指令不認人）
            upload_state = state_manager.get(group_id)
            print(f"[txt-buf] 內部群 flush: user={user_id[:10]}... upload_state(group)={upload_state.get('action') if upload_state else None} combined={combined[:30]!r}", flush=True)
            # 訂單 state：per-group（任何人都能接下「客戶名 N個」）
            group_order_state = state_manager.get(group_id)

            # ── 新增品項：完整訊息 → 優先立即處理（不論是否在 session 中）──
            if _NEW_PROD_TRIGGER_RE.match(combined.strip()) and _split_new_product_entries(combined):
                _new_prod_timer_cancel(user_id)
                state_manager.clear(user_id)
                ack = handle_internal_new_product(combined)
                if ack:
                    _send_group_ack(ack)
                return

            # ── 新增品項 Session 進行中 ──────────────────────────────────
            if upload_state and upload_state.get("action") == "new_product_session":
                if _INTERNAL_UPLOAD_FINISH_RE.match(combined.strip()):   # 「完成」
                    _new_prod_timer_cancel(user_id)
                    lines = upload_state.get("lines", [])
                    state_manager.clear(user_id)
                    ack = handle_internal_new_product("\n".join(lines)) if lines else None
                elif _INTERNAL_UPLOAD_CANCEL_RE.match(combined.strip()):   # 「取消」
                    _new_prod_timer_cancel(user_id)
                    state_manager.clear(user_id)
                    ack = "❌ 已取消新增品項"
                else:
                    upload_state["lines"].append(combined)
                    state_manager.set(user_id, upload_state)
                    _new_prod_timer_reset(user_id, group_id, reply_token)
                    ack = None  # 靜默等待
                if ack:
                    _send_group_ack(ack)
                return

            # ── 補圖/存圖 Session 進行中（查 group_id + user_id）──────────
            _pending_img_state = state_manager.get(group_id) if group_id else None
            if not _pending_img_state or _pending_img_state.get("action") not in ("pending_add_img", "pending_save_img"):
                _pending_img_state = upload_state if upload_state and upload_state.get("action") in ("pending_add_img", "pending_save_img") else None
            if _pending_img_state:
                if _INTERNAL_UPLOAD_FINISH_RE.match(combined.strip()):
                    _action = _pending_img_state["action"]
                    _code = _pending_img_state["prod_code"]
                    _collected = _pending_img_state.get("media", [])
                    if media_e:
                        _collected.extend(media_e["media"])
                    state_manager.clear(group_id)
                    state_manager.clear(user_id)
                    if not _collected:
                        ack = f"❌ {_code} 沒有收到圖片"
                    elif _action == "pending_add_img":
                        ack = handle_internal_add_images(_code, _collected)
                    else:
                        ack = handle_internal_save_images(_code, _collected)
                elif _INTERNAL_UPLOAD_CANCEL_RE.match(combined.strip()):
                    _code = _pending_img_state.get("prod_code", "")
                    _n = len(_pending_img_state.get("media", []))
                    state_manager.clear(group_id)
                    state_manager.clear(user_id)
                    ack = f"❌ {_code} 已取消（丟棄 {_n} 張）" if _n else f"❌ {_code} 已取消"
                else:
                    # 非完成指令 → 累積圖片
                    if media_e:
                        _existing = _pending_img_state.get("media", [])
                        _existing.extend(media_e["media"])
                        _pending_img_state["media"] = _existing
                        _state_key = group_id if state_manager.get(group_id) else user_id
                        state_manager.set(_state_key, _pending_img_state)
                    ack = None
                if ack:
                    _send_group_ack(ack)
                return

            # ── 批次上架 Session 進行中（per-group）─────────────────────
            if upload_state and upload_state.get("action") == "uploading":
                if _INTERNAL_UPLOAD_FINISH_RE.match(combined.strip()):
                    _upload_timer_cancel(group_id)
                    # 補進統一 buffer 中等待的圖片/影片
                    if media_e:
                        for _mi in media_e["media"]:
                            handle_internal_upload_add_media(group_id, _mi["msg_id"], _mi["type"])
                    ack = handle_internal_upload_finish(group_id)
                elif _INTERNAL_UPLOAD_CANCEL_RE.match(combined.strip()):
                    _upload_timer_cancel(group_id)
                    ack = handle_internal_upload_cancel(group_id)
                else:
                    # 收到文字也重置 timer，避免文字和 auto-finish 同時觸發兩個回覆
                    _upload_timer_reset(group_id, group_id, reply_token)
                    ack = handle_internal_upload_text(group_id, combined)
                    # 補進跟文字一起到達的圖片/影片（屬於新 PO文）
                    if media_e:
                        for _mi in media_e["media"]:
                            handle_internal_upload_add_media(group_id, _mi["msg_id"], _mi["type"])
                if ack:
                    _send_group_ack(ack)
                return

            # ── 觸發批次上架 Session ─────────────────────────────────────
            _first_line = combined.split('\n')[0].strip()
            if _first_line in _INTERNAL_UPLOAD_TRIGGERS:
                ack = handle_internal_upload_start(group_id)
                if ack:
                    _send_group_ack(ack)
                _remaining = combined[len(_first_line):].strip()
                if _remaining:
                    # 修 Bug2：_remaining 最後一行若是「完成」→ 拆開，先存 PO 文再 finish
                    _rem_lines = _remaining.splitlines()
                    _finish_idx = next(
                        (i for i, l in enumerate(_rem_lines)
                         if _INTERNAL_UPLOAD_FINISH_RE.match(l.strip())),
                        -1
                    )
                    if _finish_idx >= 0:
                        _po_part = "\n".join(_rem_lines[:_finish_idx]).strip()
                        if _po_part:
                            handle_internal_upload_text(group_id, _po_part)
                        # 補進統一 buffer 中的媒體 — 利用 after_text 按順序分配給各 PO文組
                        if media_e:
                            _raw_lines = entry.get("lines", [])
                            # 找出每行原始文字裡含貨號的行號（在 raw_lines 中的 index）
                            from handlers.internal import _PROD_CODE_RE as _PCR
                            _code_line_indices = []
                            for _li, _ln in enumerate(_raw_lines):
                                _mc = _PCR.search(_ln)
                                if _mc:
                                    _code_line_indices.append(_li)
                            if len(_code_line_indices) > 1:
                                # 多組 PO文：按 after_text 分配圖片
                                from storage.state import state_manager as _sm
                                _st = _sm.get(group_id)
                                if _st:
                                    _grps = _st.get("groups", [])
                                    # 為每組建立 media list（按 after_text 切割）
                                    # 圖片在某個 PO文之後、下一個 PO文之前 → 屬於該 PO文
                                    for _mi in media_e["media"]:
                                        _at = _mi.get("after_text", 0)
                                        # 找出這張圖屬於哪個 code（最後一個 line_idx <= after_text 的）
                                        _owner_idx = 0
                                        for _ci, _cli in enumerate(_code_line_indices):
                                            if _cli < _at:
                                                _owner_idx = _ci
                                            else:
                                                break
                                        # _owner_idx 對應 groups 中的第幾組
                                        if _owner_idx < len(_grps):
                                            _grps[_owner_idx].setdefault("media", []).append(_mi)
                                        else:
                                            # 屬於 current（最後一組）
                                            handle_internal_upload_add_media(group_id, _mi["msg_id"], _mi["type"])
                                    _st["groups"] = _grps
                                    _sm.set(group_id, _st)
                            else:
                                # 單一貨號：全部圖片給 current
                                for _mi in media_e["media"]:
                                    handle_internal_upload_add_media(group_id, _mi["msg_id"], _mi["type"])
                        ack2 = handle_internal_upload_finish(group_id)
                    else:
                        ack2 = handle_internal_upload_text(group_id, _remaining)
                        # 補進 media_buf（圖片/影片在 txt_buf 15s 內到達的情況）
                        if media_e:
                            for _mi in media_e["media"]:
                                handle_internal_upload_add_media(group_id, _mi["msg_id"], _mi["type"])
                            _upload_timer_reset(group_id, group_id, reply_token)
                    if ack2:
                        _send_group_ack(ack2)
                elif media_e:
                    # 只有 "上架" 沒有 PO文，但有圖片（快速傳送情境）
                    for _mi in media_e["media"]:
                        handle_internal_upload_add_media(group_id, _mi["msg_id"], _mi["type"])
                    _upload_timer_reset(group_id, group_id, reply_token)
                return

            # ── 存圖 Z3432 或 Z3432存圖（替換舊圖 + 圖片/影片）──────────
            _save_img_m = _INTERNAL_SAVE_IMG_RE.search(combined)
            if _save_img_m and media_e:
                _code = (_save_img_m.group(1) or _save_img_m.group(2) or "").upper()
                ack = handle_internal_save_images(_code, media_e["media"])
                if ack:
                    _send_group_ack(ack)
                return
            elif _save_img_m and not media_e:
                _save_code = (_save_img_m.group(1) or _save_img_m.group(2) or "").upper()
                # 用 group_id 存 state（內部群任何人傳圖都能觸發）
                state_manager.set(group_id, {
                    "action": "pending_save_img",
                    "prod_code": _save_code,
                    "group_id": group_id,
                })
                return

            # ── 加圖 Z3432 或 Z3432加圖（保留舊圖，追加新圖片/影片）──────
            _add_img_m = _INTERNAL_ADD_IMG_RE.search(combined)
            if _add_img_m and media_e:
                _code = (_add_img_m.group(1) or _add_img_m.group(2) or "").upper()
                ack = handle_internal_add_images(_code, media_e["media"])
                if ack:
                    _send_group_ack(ack)
                return
            elif _add_img_m and not media_e:
                # 文字先到，圖片還沒到 → 用 group_id 存 state
                _add_code = (_add_img_m.group(1) or _add_img_m.group(2) or "").upper()
                state_manager.set(group_id, {
                    "action": "pending_add_img",
                    "prod_code": _add_code,
                    "group_id": group_id,
                })
                return

            # ── 沒貨/有貨 Z3432（秒殺標記/取消）─────────────────────────
            _sold_m = _INTERNAL_SOLD_OUT_RE.search(combined)
            if _sold_m:
                _so_code = (_sold_m.group(1) or _sold_m.group(2) or "").upper()
                ack = handle_internal_mark_sold_out(_so_code)
                if ack:
                    _send_group_ack(ack)
                return
            _restock_m = _INTERNAL_RESTOCK_RE.search(combined)
            if _restock_m:
                _rs_code = (_restock_m.group(1) or _restock_m.group(2) or "").upper()
                ack = handle_internal_unmark_sold_out(_rs_code)
                if ack:
                    _send_group_ack(ack)
                return

            # ── 存文（純文字 → PO文.txt）────────────────────────────────
            if "存文" in combined:
                ack = handle_internal_save_text(combined)
                if ack:
                    _send_group_ack(ack)
                return

            # ── 既有：圖片識別（從 media_e 取第一張 或 img_e）──────────
            img_pc = None
            source_msg_id = None
            if media_e and media_e["media"] and media_e["media"][0]["type"] == "image":
                source_msg_id = media_e["media"][0]["msg_id"]
            elif img_e:
                source_msg_id = img_e["msg_id"]
            if source_msg_id:
                img_pc = _img_identify_from_buf(source_msg_id)
                # pHash/OCR 都失敗 → fallback 走 Claude vision（內部群專屬，客戶 1:1 已在 service.py 走過）
                if not img_pc:
                    from services.vision import download_image as _dl_im
                    from services.claude_ai import ask_claude_image as _cai
                    _im_bytes = _dl_im(source_msg_id)
                    if _im_bytes:
                        _claude_reply = _cai(_im_bytes, user_id=user_id)
                        if _claude_reply:
                            _codes = re.findall(r'[A-Za-z]{1,3}-?\d{3,6}', _claude_reply)
                            if _codes:
                                img_pc = _codes[0].upper()
                                print(f"[txt-buf] Claude vision 辨識 → 內部群 {img_pc}", flush=True)

            if img_pc:
                print(f"[txt-buf] 圖片+文字 → 內部群 {img_pc}")
            elif source_msg_id:
                print("[txt-buf] 媒體+文字 → 內部群，圖片識別失敗（含 Claude fallback）")

            # 圖片辨識成功 → 查庫存資訊回覆（不設 state，要下單打完整格式）
            ack = None
            if img_pc:
                from services.ecount import ecount_client as _ec
                _ec._ensure_product_cache()
                _gi = _ec.lookup(img_pc)
                if _gi:
                    from handlers.internal import _format_po, _fmt_stock_lines
                    _po = _format_po(img_pc)
                    _stock = _fmt_stock_lines(_gi, img_pc)
                    ack = f"{_po}\n{_stock}"
                else:
                    # lookup 失敗（available.json 過期等）→ 只回品名+貨號
                    _cache = _ec.get_product_cache_item(img_pc)
                    _name = (_cache.get("name") if _cache else None) or img_pc
                    ack = f"📦 {_name}（{img_pc}）\n（庫存資料暫時無法取得）"

            if not ack:
                    # ── 新增品項觸發 ──
                    if _NEW_PROD_TRIGGER_RE.match(combined.strip()):
                        # 訊息已含品項資料 → 直接處理，不開 session
                        if _split_new_product_entries(combined):
                            ack = handle_internal_new_product(combined)
                        else:
                            state_manager.set(user_id, {
                                "action": "new_product_session",
                                "lines":  [combined],
                            })
                            _new_prod_timer_reset(user_id, group_id, reply_token)
                            ack = "📝 品項建立模式，請依序輸入各品項資料\n完成後傳「完成」，或等待 30 秒自動處理"
                    elif source_msg_id:
                        # 圖片辨識失敗（pHash+OCR+Claude 都沒命中）→ 告知辨識失敗
                        ack = "⚠️ 圖片辨識失敗，請再傳一次或用「存圖 XXXX」+ 圖片指定貨號（XXXX 是該產品貨號）"
                    else:
                        ack = _dispatch_internal_fallback(combined, group_id, line_api, staff_id=user_id)
            else:
                try:
                    # ── 新增品項觸發 ──
                    if _NEW_PROD_TRIGGER_RE.match(combined.strip()):
                        # 訊息已含品項資料 → 直接處理，不開 session
                        if _split_new_product_entries(combined):
                            ack = handle_internal_new_product(combined)
                        else:
                            state_manager.set(user_id, {
                                "action": "new_product_session",
                                "lines":  [combined],
                            })
                            _new_prod_timer_reset(user_id, group_id, reply_token)
                            ack = "📝 品項建立模式，請依序輸入各品項資料\n完成後傳「完成」，或等待 30 秒自動處理"
                    else:
                        print(f"[txt-buf] dispatch_internal_fallback 開始: {combined!r}", flush=True)
                        ack = _dispatch_internal_fallback(combined, group_id, line_api, staff_id=user_id)
                        print(f"[txt-buf] dispatch_internal_fallback 結果: {ack!r}", flush=True)
                except Exception as _ge:
                    print(f"[txt-buf] 內部群處理例外: {type(_ge).__name__}: {_ge}", flush=True)
                    import traceback; traceback.print_exc()
                    ack = f"❌ 處理時發生錯誤：{_ge}"
            if ack:
                _send_group_ack(ack)

        # ════ 1:1 客戶 ════
        else:
            try:
                current_state = state_manager.get(user_id)

                # 文字含收據/明細等 → 不走圖片辨識（收據上有貨號會誤匹配）
                _DOC_KW = ["收據", "明細", "發票", "對帳", "帳單"]
                if any(kw in combined for kw in _DOC_KW):
                    img_e = None  # 清掉圖片，不辨識

                # 1:1 多引用處理：每行 tag 不同圖片 + 數量 → 批次加購物車
                _line_quotes = entry.get("line_quotes", [])
                _has_multi_quotes = (
                    len(_line_quotes) > 1
                    and sum(1 for q in _line_quotes if q) >= 2
                )
                if _has_multi_quotes and not img_e:
                    from handlers.ordering import extract_quantity as _eq_mq
                    _added_items = []
                    for _li, (_line_text, _lq) in enumerate(zip(entry["lines"], _line_quotes)):
                        if not _lq:
                            continue
                        _lq_code = lookup_sent_image(_lq)
                        if not _lq_code:
                            continue
                        _lq_qty = _eq_mq(_line_text) or 1
                        from services.ecount import ecount_client as _ec_mq
                        _lq_item = _ec_mq.get_product_cache_item(_lq_code)
                        _lq_name = (_lq_item.get("name") if _lq_item else "") or _lq_code
                        from storage import cart as _cart_mq
                        _cart_mq.set_item(user_id, _lq_code, _lq_name, _lq_qty)
                        _added_items.append(f"  • {_lq_name}（{_lq_code}）× {_lq_qty}")
                        print(f"[quote-multi] {_lq_code} × {_lq_qty}", flush=True)
                    _has_unknown_quotes = any(
                        lq and not lookup_sent_image(lq)
                        for lq in _line_quotes if lq
                    )
                    if _added_items:
                        state_manager.clear(user_id)
                        reply_text = tone.cart_item_added(_cart_mq.get_cart(user_id))
                        _send_reply(reply_token, user_id, reply_text, line_api)
                        return
                    elif _has_unknown_quotes:
                        issue_store.add(user_id, "quote_unknown", f"客戶引用了無法辨識的圖片：{combined[:50]}")
                        from handlers.hours import _is_open_now as _io_q, next_open_reply as _nor_q
                        from datetime import datetime as _dt_q
                        import pytz as _pz_q
                        _now_q = _dt_q.now(_pz_q.timezone(settings.BUSINESS_TZ))
                        _q_reply = "稍等一下唷～等等處理嘿" if _io_q(_now_q) else _nor_q()
                        _send_reply(reply_token, user_id, _q_reply, line_api)
                        return

                # ── tag 清單 + 確認/送出/好了 → 直接結帳（略過圖片/貨號追溯）──
                if quoted_msg_id and not img_e:
                    from storage import cart as _cart_qck
                    _CHECKOUT_QUICK_KW = set(CHECKOUT_KEYWORDS) | set(AFFIRMATIVE_KEYWORDS)
                    _txt_strip = combined.strip()
                    _is_checkout_word = (_txt_strip in _CHECKOUT_QUICK_KW or
                                         any(kw in _txt_strip for kw in
                                             ["確認訂單", "送出訂單", "幫我送出", "幫忙送出",
                                              "就這樣送出", "就這些送出", "這樣就好"]))
                    if _is_checkout_word and not _cart_qck.is_empty(user_id):
                        # 已經在地址選擇/聯絡資訊等結帳中間狀態 → 不再重複觸發 checkout
                        _qck_state = state_manager.get(user_id) or {}
                        _qck_action = _qck_state.get("action", "")
                        if _qck_action in ("awaiting_address_selection_checkout",
                                           "awaiting_group_address_confirm",
                                           "awaiting_contact_info_checkout",
                                           "awaiting_address_selection"):
                            print(f"[quote] tag 清單 + 結帳詞，但已在 {_qck_action}，忽略", flush=True)
                            return
                        print(f"[quote] tag 清單 + 結帳關鍵字 → 直接結帳: {_txt_strip!r}", flush=True)
                        from handlers.ordering import handle_checkout
                        _qck_reply = handle_checkout(user_id, line_api)
                        _send_reply(reply_token, user_id, _qck_reply, line_api)
                        return

                # 1:1 圖片識別：優先從文字提取貨號，沒有才辨識圖片
                # 若客戶引用了某張圖片，把引用的圖片當作 img_e 來辨識
                if quoted_msg_id and not img_e:
                    _quoted_pc = _img_identify_from_buf(quoted_msg_id)
                    if _quoted_pc:
                        print(f"[quote] 引用圖片辨識成功: {quoted_msg_id} → {_quoted_pc}", flush=True)
                        img_e = {"msg_id": quoted_msg_id}
                    else:
                        # 可能是 bot 發的圖片（無法下載）→ 查對應表
                        _sent_pc = lookup_sent_image(quoted_msg_id)
                        if _sent_pc:
                            print(f"[quote] 引用 bot 發的圖片: {quoted_msg_id} → {_sent_pc}", flush=True)
                            img_e = {"msg_id": quoted_msg_id, "bot_sent_code": _sent_pc}
                        else:
                            # 不是圖片 → 查客戶自己之前傳的文字（引用文字場景）
                            _quoted_text = lookup_incoming_text(quoted_msg_id)
                            _quoted_codes = []
                            if _quoted_text:
                                _quoted_codes = list(dict.fromkeys(
                                    c.upper() for c in _PROD_CODE_RE.findall(_quoted_text)
                                ))
                            if _quoted_codes:
                                print(f"[quote] 引用客戶文字 {quoted_msg_id} → 抓到貨號 {_quoted_codes}", flush=True)
                                # 把第一個貨號當「已辨識產品」走後續正常流程（同 bot_sent_code 路徑）
                                img_e = {"msg_id": quoted_msg_id, "bot_sent_code": _quoted_codes[0]}
                            else:
                                print(f"[quote] 引用辨識失敗（非圖片、文字也查不到貨號）: {quoted_msg_id}", flush=True)
                                # 查不到對應產品 → 回覆稍等 + 記待處理
                                issue_store.add(user_id, "quote_unknown", f"客戶引用了無法辨識的訊息：{combined[:50]}")
                                from handlers.hours import _is_open_now as _io_q, next_open_reply as _nor_q
                                from datetime import datetime as _dt_q
                                import pytz as _pz_q
                                _now_q = _dt_q.now(_pz_q.timezone(settings.BUSINESS_TZ))
                                _q_reply = "稍等一下唷～等等處理嘿" if _io_q(_now_q) else _nor_q()
                                _send_reply(reply_token, user_id, _q_reply, line_api)
                                return

                img_pcs = []
                _text_codes = _PROD_CODE_RE.findall(combined) if img_e else []
                if _text_codes:
                    # 文字裡有貨號，直接用
                    for _tc in _text_codes:
                        _tc_upper = _tc.upper()
                        if _tc_upper not in img_pcs:
                            img_pcs.append(_tc_upper)
                elif img_e and img_e.get("bot_sent_code"):
                    # bot 發的圖片被 tag → 直接用已記錄的產品代碼
                    img_pcs.append(img_e["bot_sent_code"])
                elif img_e:
                    # 文字沒貨號，才用圖片辨識（含引用圖片）
                    msg_ids = img_e.get("msg_ids", [img_e["msg_id"]] if img_e.get("msg_id") else [])
                    for _mid in msg_ids:
                        _pc = _img_identify_from_buf(_mid)
                        if _pc and _pc not in img_pcs:
                            img_pcs.append(_pc)
                img_pc = img_pcs[0] if img_pcs else None

                # ── 多張圖但辨識失敗：跟客戶說清楚，別裝單張 ──────────────
                _img_count = len(img_e.get("msg_ids", [])) if img_e else 0
                _each_m_chk = re.search(r'各\s*(\d+)', combined) if combined else None
                if _img_count >= 2 and len(img_pcs) < _img_count and _each_m_chk and not current_state:
                    # 有 N 張圖、客戶講「各 X」、但只認出 < N 款 → 登記真人處理
                    issue_store.add(user_id, "multi_img_partial",
                                    f"客戶傳 {_img_count} 張圖+「{combined}」，只認出 {len(img_pcs)} 款：{img_pcs}")
                    print(f"[txt-buf] 多圖辨識不全：{_img_count} 張 / 認出 {len(img_pcs)}，轉真人", flush=True)
                    reply_text = f"收到～您傳了 {_img_count} 張圖，我這邊還在確認中，稍等一下嘿 小幫手馬上為您處理"
                    _send_reply(reply_token, user_id, reply_text, line_api)
                    reply_text = None
                    img_pc = None
                    img_pcs = []

                # ── 多張圖片 + 「各X」→ 批次加入購物車 ──────────────
                _each_m = re.search(r'各\s*(\d+)\s*(?:個|箱|件|盒|套|組)?', combined) if img_pcs else None
                if len(img_pcs) > 1 and _each_m and not current_state:
                    _each_qty = int(_each_m.group(1))
                    from services.ecount import ecount_client as _ec
                    from storage import cart as cart_store
                    _added = []
                    for _pc in img_pcs:
                        _info = _ec.lookup(_pc)
                        _pn = (_info.get("name") if _info else None) or _pc
                        cart_store.add_item(user_id, _pc, _pn, _each_qty)
                        _added.append(f"• {_pn} × {_each_qty}")
                    state_manager.set(user_id, {"action": "awaiting_multi_img_confirm"})
                    reply_text = (
                        "收到！已加入購物車：\n"
                        + "\n".join(_added)
                        + "\n\n如果沒有問題我就送出訂單囉"
                    )
                    _send_reply(reply_token, user_id, reply_text, line_api)
                    reply_text = None
                    img_pc = None  # 已處理，不走單張流程

                # ── 圖片 + 文字：規則判斷意圖，再決定處理方式 ──────────────
                # 有圖片時優先處理圖片（清除舊 state，避免被舊狀態阻擋）
                if img_pc and current_state:
                    state_manager.clear(user_id)
                    current_state = None
                    print(f"[txt-buf] 有圖片，清除舊 state: {current_state}", flush=True)
                if img_pc and not current_state:
                    # 秒殺擋下 — 內部群已標記「沒貨 XXXX」→ 直接回擋客戶下單
                    from storage import sold_out as _so
                    if _so.is_sold_out(img_pc):
                        print(f"[sold-out] 擋下客戶下單 {img_pc} user={user_id[:10]}...", flush=True)
                        _send_reply(reply_token, user_id, tone.sold_out_secret_kill(), line_api)
                        return
                    from services.ecount import ecount_client as _ec
                    from handlers.inventory import _check_preorder
                    _ui   = _ec.lookup(img_pc)
                    _uqty = _ui.get("qty") if _ui else None
                    _un   = (_ui.get("name") if _ui else None) or img_pc

                    # 從文字判斷客戶意圖
                    _txt_lower = combined.lower()
                    _delivery_kw = any(k in combined for k in [
                        "什麼時候到", "何時到", "到貨了嗎", "到了嗎", "到貨時間",
                        "什麼時候會到", "幾時到", "到貨沒", "出貨了嗎", "出貨沒",
                        "寄了嗎", "送到了嗎", "等很久", "到貨", "催貨",
                    ])
                    _inv_kw  = any(k in combined for k in ["有嗎", "有沒有", "庫存", "缺貨", "還有", "有貨", "剩幾"])
                    _reserve_kw = any(k in combined for k in [
                        "幫我留", "幫留", "留一個", "留一隻", "留一支", "留給我",
                        "要留", "先留", "留著", "留下",
                    ])
                    _price_kw = any(k in combined for k in ["多少", "幾元", "幾錢", "價格", "價錢", "多少錢", "售價"])
                    _qty_m   = re.search(r'(\d+)\s*(?:個|箱|件|盒|套|組|支|隻)', combined)
                    # 也支援中文數字（八個、十二箱等）
                    from handlers.ordering import extract_quantity as _ext_qty
                    _ext_qty_val = _ext_qty(combined)

                    if _delivery_kw:
                        # 問到貨/出貨 → 轉真人處理（Bot 無法查訂單進度）
                        print(f"[txt-buf] 圖片+文字 → 問到貨時間 {img_pc}", flush=True)
                        issue_store.add(user_id, "delivery_query",
                                        f"（傳圖詢問到貨）{_un}（{img_pc}）：{combined}")
                        reply_text = tone.urgent_order_ack()
                        _send_reply(reply_token, user_id, reply_text, line_api)
                        reply_text = None
                    elif _inv_kw:
                        # 問庫存 → 直接用辨識到的貨號查（不走 handle_inventory 的文字解析）
                        print(f"[txt-buf] 圖片+文字 → 問庫存 {img_pc}", flush=True)
                        _inv_reply = None
                        if _uqty and _uqty > 0:
                            from handlers.inventory import _query_single_product
                            _inv_reply = _query_single_product(user_id, img_pc, line_api)
                        else:
                            # 沒貨 → 判斷預購或缺貨
                            if _check_preorder(img_pc):
                                state_manager.set(user_id, {
                                    "action":    "awaiting_quantity",
                                    "prod_cd":   img_pc,
                                    "prod_name": _un,
                                })
                                _inv_reply = tone.preorder_ask_qty(_un)
                                print(f"[txt-buf] 圖片+文字 → 問庫存，預購 {img_pc}", flush=True)
                            else:
                                state_manager.set(user_id, {
                                    "action":    "awaiting_quantity",
                                    "prod_cd":   img_pc,
                                    "prod_name": _un,
                                })
                                _inv_reply = tone.out_of_stock_ask_qty(_un)
                                print(f"[txt-buf] 圖片+文字 → 問庫存，缺貨走購物車 {img_pc}", flush=True)
                        if _inv_reply:
                            _send_reply(reply_token, user_id, _inv_reply, line_api)
                        reply_text = None
                        return  # 已處理完畢，不走後面的 dispatch
                    elif _reserve_kw:
                        # 幫我留/要留 → 加購物車（下單意圖，優先於 price_kw，因為 PO 文常含「價格」誤觸）
                        _rsv_qty = _ext_qty_val or (int(_qty_m.group(1)) if _qty_m else None) or 1
                        from storage import cart as _cart_rsv
                        _existing = next((it for it in _cart_rsv.get_cart(user_id)
                                          if it["prod_cd"].upper() == img_pc.upper()), None)
                        if _existing:
                            _cart_rsv.set_item(user_id, img_pc, _un, _rsv_qty)
                            print(f"[txt-buf] 圖片+文字 → 要留，改數量 {img_pc} {_existing['qty']}→{_rsv_qty}", flush=True)
                        else:
                            _cart_rsv.add_item(user_id, img_pc, _un, _rsv_qty)
                            print(f"[txt-buf] 圖片+文字 → 要留，加購物車 {img_pc} x{_rsv_qty}", flush=True)
                        reply_text = tone.cart_item_added(_cart_rsv.get_cart(user_id))
                        _send_reply(reply_token, user_id, reply_text, line_api)
                        return
                    elif _price_kw:
                        # 問價格 → 走 price handler
                        print(f"[txt-buf] 圖片+文字 → 問價格 {img_pc}", flush=True)
                        reply_text = handle_price(user_id, img_pc)
                        if reply_text:
                            _send_reply(reply_token, user_id, reply_text, line_api)
                        reply_text = None
                    elif (_qty_m or _ext_qty_val) and (_uqty and _uqty > 0):
                        # 直接說要幾個 + 有貨
                        # 同貨號 → 預設「改成 N 個」(set)；除非客戶明說「再/再加/多/多加/加買」才累加 (add)
                        _direct_qty = _ext_qty_val or int(_qty_m.group(1))
                        _add_more = bool(re.search(r'(?:再|多|加買)\s*\d|再加|多加|還要', combined))
                        from storage import cart as _cart_direct
                        _existing = next((it for it in _cart_direct.get_cart(user_id)
                                          if it["prod_cd"].upper() == img_pc.upper()), None)
                        if _existing and not _add_more:
                            _cart_direct.set_item(user_id, img_pc, _un, _direct_qty)
                            print(f"[txt-buf] 圖片+文字 → 改數量 {img_pc} {_existing['qty']}→{_direct_qty}", flush=True)
                        else:
                            _cart_direct.add_item(user_id, img_pc, _un, _direct_qty)
                            print(f"[txt-buf] 圖片+文字 → 直接下單 {img_pc} x{_direct_qty} ({'累加' if _existing else '新增'})", flush=True)
                        reply_text = tone.cart_item_added(_cart_direct.get_cart(user_id))
                        _send_reply(reply_token, user_id, reply_text, line_api)
                        return
                    elif (_qty_m or _ext_qty_val) and not (_uqty and _uqty > 0):
                        # 缺貨/預購 + 有數量
                        _direct_qty = _ext_qty_val or int(_qty_m.group(1))
                        _is_po = _check_preorder(img_pc)
                        _add_more = bool(re.search(r'(?:再|多|加買)\s*\d|再加|多加|還要', combined))
                        from storage import cart as _cart_oos
                        _existing = next((it for it in _cart_oos.get_cart(user_id)
                                          if it["prod_cd"].upper() == img_pc.upper()), None)
                        if _existing and not _add_more:
                            _cart_oos.set_item(user_id, img_pc, _un, _direct_qty)
                            print(f"[txt-buf] 圖片+文字 → 改數量({'預購' if _is_po else '缺貨'}) {img_pc} {_existing['qty']}→{_direct_qty}", flush=True)
                        else:
                            _cart_oos.add_item(user_id, img_pc, _un, _direct_qty)
                            print(f"[txt-buf] 圖片+文字 → 直接下單({'預購' if _is_po else '缺貨'}) {img_pc} x{_direct_qty} ({'累加' if _existing else '新增'})", flush=True)
                        reply_text = tone.cart_item_added(_cart_oos.get_cart(user_id))
                        _send_reply(reply_token, user_id, reply_text, line_api)
                        return
                    else:
                        # 意圖不明 → 交給 Claude 指令引擎判斷
                        from services.claude_ai import ask_claude_command as _acc_img
                        _cmd = _acc_img(combined, user_id=user_id,
                                        product_code=img_pc, product_name=_un)
                        if _cmd:
                            _cmd_result = _execute_claude_command(user_id, _cmd, line_api, original_text=combined)
                            if _cmd_result:
                                _send_reply(reply_token, user_id, _cmd_result, line_api)
                                return
                        # Claude 也判斷不了 → fallback 到 awaiting_quantity
                        state_manager.set(user_id, {
                            "action":    "awaiting_quantity",
                            "prod_cd":   img_pc,
                            "prod_name": _un,
                        })
                        if _uqty and _uqty > 0:
                            print(f"[txt-buf] 圖片+文字 → 有貨 {img_pc}，等待數量", flush=True)
                        elif _check_preorder(img_pc):
                            reply_text = tone.preorder_ask_qty(_un)
                            _send_reply(reply_token, user_id, reply_text, line_api)
                            reply_text = None
                        else:
                            reply_text = tone.out_of_stock_ask_qty(_un)
                            _send_reply(reply_token, user_id, reply_text, line_api)
                            reply_text = None
                        current_state = state_manager.get(user_id)

                elif img_e:
                    # 有新圖片但辨識失敗 → 清掉舊 state，用 Claude 辨識
                    if current_state:
                        state_manager.clear(user_id)
                        current_state = None
                        print(f"[txt-buf] 有新圖片，清掉舊 state", flush=True)
                    print("[txt-buf] 圖片+文字 → 1:1，圖片識別失敗，嘗試 Claude", flush=True)
                    # 嘗試用 Claude 辨識圖片（附上 OCR 結果作為提示）
                    from services.claude_ai import ask_claude_image
                    from services.vision import download_image as _dl_img, ocr_extract_candidates
                    _img_mid = img_e.get("msg_id") or (img_e.get("msg_ids", [None])[0])
                    _img_data = _dl_img(_img_mid) if _img_mid else None
                    _ocr_hint = ""
                    if _img_data:
                        _ocr_candidates = ocr_extract_candidates(_img_data)
                        if _ocr_candidates:
                            _ocr_hint = f"\nOCR 讀到的文字：{'、'.join(_ocr_candidates[:20])}"
                    _claude_combined = combined + _ocr_hint
                    _claude_img_reply = ask_claude_image(_img_data, _claude_combined, user_id=user_id) if _img_data else None
                    if _claude_img_reply:
                        _ci_codes_raw = _PROD_CODE_RE.findall(_claude_img_reply)
                        # 去重
                        _ci_codes = list(dict.fromkeys(c.upper() for c in _ci_codes_raw))
                        # 過濾：只保留 Ecount 實際存在的貨號（避免把品牌型號如 DG556 當貨號）
                        from services.ecount import ecount_client as _ec_filter
                        _ci_codes = [c for c in _ci_codes if _ec_filter.get_product_cache_item(c)]
                        from handlers.ordering import extract_quantity as _eq_ci
                        _ci_qty = _eq_ci(combined)
                        if len(_ci_codes) == 1 and _ci_qty:
                            # 單一產品 + 有數量 → 直接加購物車
                            _ci_cd = _ci_codes[0]
                            from services.ecount import ecount_client as _ec_ci
                            _ci_item = _ec_ci.get_product_cache_item(_ci_cd)
                            _ci_name = (_ci_item.get("name") if _ci_item else None) or _ci_cd
                            from storage import cart as _cart_ci
                            _cart_ci.add_item(user_id, _ci_cd, _ci_name, _ci_qty)
                            reply_text = tone.cart_item_added(_cart_ci.get_cart(user_id))
                            print(f"[claude-ai] 圖片辨識+數量 → 加購物車: {_ci_cd} x{_ci_qty}", flush=True)
                            _send_reply(reply_token, user_id, reply_text, line_api)
                        elif len(_ci_codes) > 1:
                            # 多個產品 → 回覆 Claude 的列表讓客戶選，不自動加購物車
                            print(f"[claude-ai] 圖片辨識出多個產品: {_ci_codes}，讓客戶選", flush=True)
                            _send_reply(reply_token, user_id, _claude_img_reply, line_api)
                        elif _ci_codes and not _ci_qty:
                            # Claude 辨識出產品但沒數量 → 設 state 等數量
                            _ci_cd = _ci_codes[0].upper()
                            from services.ecount import ecount_client as _ec_ci2
                            _ci_item2 = _ec_ci2.get_product_cache_item(_ci_cd)
                            _ci_name2 = (_ci_item2.get("name") if _ci_item2 else None) or _ci_cd
                            state_manager.set(user_id, {
                                "action":    "awaiting_quantity",
                                "prod_cd":   _ci_cd,
                                "prod_name": _ci_name2,
                            })
                            print(f"[claude-ai] 圖片辨識後設 awaiting_quantity: {_ci_cd}", flush=True)
                            _send_reply(reply_token, user_id, _claude_img_reply, line_api)
                        else:
                            # Claude 也辨識不出貨號 → 記待處理 + 根據營業時間送回覆
                            issue_store.add(user_id, "image_query", f"（圖片+文字，Claude 無法辨識）{combined[:30]}")
                            from handlers.hours import _is_open_now as _io_img, next_open_reply as _nor_img
                            from datetime import datetime as _dt_img
                            import pytz as _pz_img
                            _now_img = _dt_img.now(_pz_img.timezone(settings.BUSINESS_TZ))
                            if not _io_img(_now_img):
                                _claude_img_reply = _nor_img()
                            _send_reply(reply_token, user_id, _claude_img_reply, line_api)
                            print(f"[claude-ai] 圖片無貨號 → 送回覆（上班={_io_img(_now_img)}）+ 記待處理", flush=True)
                        # Claude 回覆含「確認」「稍後」→ 登記待處理 + 清掉 awaiting_quantity
                        _unsure_img_kw = ["確認一下", "稍後回覆", "幫您確認", "幫你確認", "稍後", "查一下"]
                        if _claude_img_reply and any(k in _claude_img_reply for k in _unsure_img_kw):
                            state_manager.clear(user_id)  # 不該同時等數量又說幫確認
                            issue_store.add(user_id, "claude_unsure", f"圖片+文字需確認：{combined[:50]}")
                            print(f"[claude-ai] 圖片回覆含確認語，清 state + 記待處理", flush=True)
                        # chat_history 由 _send_reply 在送出成功後自動記（避免假紀錄）
                        return
                    issue_store.add(user_id, "image_query", "（圖片+文字，圖片無法辨識）")

                # ── 凍結判斷：有待處理問題 → 完全靜默，等真人標記完成 ──────
                if not current_state and issue_store.has_pending_issue(user_id):
                    print(f"[frozen] {user_id[:10]}... 有待處理問題，靜默", flush=True)
                    return

                # 記錄客戶訊息（供 Claude 前後文）
                from services.claude_ai import add_chat_history
                add_chat_history(user_id, "user", combined)

                if is_payment_message(combined):
                    reply_text = handle_payment(user_id, combined)
                elif current_state and current_state.get("action") == "image_waiting_text":
                    # 圖片先到已辨識，文字後到 → 合併處理
                    _iwt_code = current_state["prod_cd"]
                    state_manager.clear(user_id)
                    print(f"[msg-buf] image_waiting_text + 文字到了，產品={_iwt_code} 文字={combined[:30]!r}", flush=True)
                    from services.ecount import ecount_client as _ec_iwt
                    _iwt_item = _ec_iwt.lookup(_iwt_code)
                    _iwt_name = (_iwt_item.get("name") if _iwt_item else "") or _iwt_code
                    # 交給 Claude 指令引擎判斷
                    from services.claude_ai import ask_claude_command as _acc_iwt
                    _iwt_cmd = _acc_iwt(combined, user_id=user_id,
                                        product_code=_iwt_code, product_name=_iwt_name)
                    if _iwt_cmd:
                        reply_text = _execute_claude_command(user_id, _iwt_cmd, line_api, original_text=combined)
                    if not reply_text:
                        # Claude 指令引擎沒回應 → 用產品資訊回覆
                        reply_text = handle_image_product(user_id, current_state.get("msg_id", ""), line_api)
                elif current_state:
                    reply_text = _handle_stateful(user_id, combined, current_state, line_api)
                    if reply_text is None:
                        # state 被清除且回 None → 重新走正常 dispatch
                        intent     = detect_intent(combined)
                        reply_text = _dispatch(user_id, combined, intent, line_api)
                else:
                    intent     = detect_intent(combined)
                    reply_text = _dispatch(user_id, combined, intent, line_api)

                if reply_text:
                    _send_reply(reply_token, user_id, reply_text, line_api)
            except Exception as _ce:
                print(f"[txt-buf] 1:1 客戶處理例外: {_ce}", flush=True)
                import traceback; traceback.print_exc()


def _img_identify_from_buf(msg_id: str) -> str | None:
    """
    從緩衝的 msg_id 下載圖片 → pHash → OCR，回傳 prod_code（或 None）。
    供文字 handler 在「圖片+文字合併」時使用。
    """
    from services.vision import download_image, identify_product, identify_product_weak, ocr_extract_candidates
    from services.ecount import ecount_client as _ec

    img_bytes = download_image(msg_id)
    if not img_bytes:
        return None

    prod_code = identify_product(img_bytes)
    if not prod_code:
        # 嚴格 pHash 沒中 → 試弱命中（diff ≤ 10），仍然比 OCR 可靠
        prod_code = identify_product_weak(img_bytes)
    if not prod_code:
        # 只嘗試「貨號格式」(字母+數字) 或中文詞，跳過純英文短詞（OCR 雜訊）
        _CODE_OR_ZH = re.compile(r'(?:[A-Za-z]\d{2,}|[\u4e00-\u9fff]{2,})')
        for candidate in ocr_extract_candidates(img_bytes):
            if not _CODE_OR_ZH.search(candidate):
                continue
            matched = _ec._resolve_product_code(candidate)
            if matched:
                prod_code = matched
                break
    return prod_code


# ── 離峰時段判斷（00:00 ~ 10:00）──────────────────────
def _in_quiet_hours() -> bool:
    """離峰時段已停用，永遠回傳 False"""
    return False


# 離峰時段仍直接回覆的意圖（只有營業時間查詢，其餘全部靜默入佇列）
_QUIET_HOURS_DIRECT_INTENTS = {
    Intent.BUSINESS_HOURS,   # 營業時間（客戶問今天有開嗎，隨時可回）
}


async def _cart_cleanup_loop():
    """每天 23:00 清理超過 48 小時的購物車"""
    while True:
        now = datetime.now()
        # 計算到今天（或明天）23:00 的秒數
        target = now.replace(hour=23, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        wait_secs = (target - now).total_seconds()
        await asyncio.sleep(wait_secs)
        try:
            from storage import cart as _cart_clean
            removed = _cart_clean.cleanup_expired(max_age_hours=48)
            if removed:
                print(f"[cart-cleanup] 清理 {removed} 個過期購物車", flush=True)
        except Exception as e:
            print(f"[cart-cleanup] 清理失敗: {e}", flush=True)


async def _refresh_data_loop():
    """每 2 小時檢查一次資料庫是否需要刷新"""
    while True:
        await asyncio.sleep(2 * 3600)
        try:
            await asyncio.to_thread(check_and_refresh)
        except Exception as e:
            print(f"[scheduler] 資料庫刷新失敗: {e}")


async def _competitor_sync_loop():
    """每天 01:00 同步 dingshang.com.tw + 產出 Excel 對比表到 資料/"""
    while True:
        now = datetime.now()
        target = now.replace(hour=1, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        try:
            from scripts.competitor_dingshang_sync import sync as _ds_sync
            result = await asyncio.to_thread(_ds_sync, True)
            print(f"[competitor-sync] dingshang 完成 total={result['total']} new={len(result['new_pids'])} updated={result['updated']}", flush=True)
            from scripts.competitor_dingshang_to_excel import _build as _ds_excel
            out_file = await asyncio.to_thread(_ds_excel)
            print(f"[competitor-sync] Excel 已輸出：{out_file}", flush=True)
        except Exception as e:
            print(f"[competitor-sync] 失敗: {e}", flush=True)


async def _process_queued_messages():
    """處理離峰佇列中所有未處理的訊息（使用 push_message 補發回覆）"""
    msgs = queue_store.get_unprocessed()
    if not msgs:
        print("[queue] 無待處理的離峰訊息")
        return

    print(f"[queue] 開始補處理 {len(msgs)} 則離峰訊息...")

    line_api = _line_api

    if True:  # preserve indentation block
        for msg in msgs:
            try:
                uid = msg["user_id"]
                if msg["msg_type"] == "text":
                    intent = detect_intent(msg["content"])
                    reply_text = _dispatch(uid, msg["content"], intent, line_api)
                elif msg["msg_type"] == "image":
                    reply_text = handle_image_product(uid, msg["msg_id"], line_api)
                else:
                    reply_text = None

                if reply_text:
                    if _push_quota_exhausted:
                        print(f"[queue] 月額度已用完，跳過 push {uid[:10]}...", flush=True)
                    else:
                        try:
                            line_api.push_message(
                                PushMessageRequest(
                                    to=uid,
                                    messages=[TextMessage(text=reply_text)],
                                )
                            )
                            print(f"[queue] OK {uid[:10]}... ({msg['msg_type']})", flush=True)
                        except Exception as _push_e:
                            if _is_quota_429(_push_e):
                                _mark_push_exhausted()
                            else:
                                print(f"[queue] push_message 失敗 (跳過): {_push_e}", flush=True)
            except Exception as e:
                print(f"[queue] 處理失敗 uid={msg['user_id'][:10]}...: {e}", flush=True)
            finally:
                queue_store.mark_processed(msg["id"])

    print(f"[queue] 補處理完成，共 {len(msgs)} 則")


async def _rebate_sync_loop():
    """每日凌晨 1:00 自動同步回饋金資料"""
    while True:
        now = datetime.now()
        target = now.replace(hour=1, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        try:
            import subprocess as _sp
            _python = _sys.executable
            _root = str(Path(__file__).parent)
            _flags = _sp.CREATE_NO_WINDOW if _sys.platform == "win32" else 0

            # 同步回饋金
            print("[rebate] 凌晨自動同步本月資料...")
            proc = await asyncio.to_thread(
                _sp.run, [_python, "-m", "scripts.sync_rebate"],
                cwd=_root, capture_output=True, timeout=180, creationflags=_flags,
            )
            if proc.stdout:
                print(proc.stdout.decode("utf-8", errors="replace"), flush=True)
            if proc.returncode != 0 and proc.stderr:
                print(proc.stderr.decode("utf-8", errors="replace"), flush=True)

            # 每月 1~3 日額外同步上月資料
            if now.day <= 3:
                print("[rebate] 月初，額外同步上月資料...")
                proc = await asyncio.to_thread(
                    _sp.run, [_python, "-m", "scripts.sync_rebate", "--last-month"],
                    cwd=_root, capture_output=True, timeout=180, creationflags=_flags,
                )
                if proc.stdout:
                    print(proc.stdout.decode("utf-8", errors="replace"), flush=True)

            # 同步未處理訂單 + 未取訂單
            print("[unfulfilled] 凌晨自動同步未處理+未取訂單...")
            proc = await asyncio.to_thread(
                _sp.run, [_python, "-m", "scripts.sync_unfulfilled"],
                cwd=_root, capture_output=True, timeout=180, creationflags=_flags,
            )
            if proc.stdout:
                print(proc.stdout.decode("utf-8", errors="replace"), flush=True)

        except Exception as e:
            print(f"[rebate/unfulfilled] 自動同步失敗: {e}")
            _notify_sync_failure("回饋金/未處理訂單同步", str(e))


async def _sync_ecount_customers_loop():
    """每日 13:00 自動同步 Ecount 客戶名單"""
    while True:
        now = datetime.now()
        target = now.replace(hour=13, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        try:
            import subprocess as _sp
            _python = _sys.executable
            _root = str(Path(__file__).parent)
            _flags = _sp.CREATE_NO_WINDOW if _sys.platform == "win32" else 0
            print("[cust-sync] 每日 13:00 自動同步 Ecount 客戶名單...")
            proc = await asyncio.to_thread(
                _sp.run, [_python, "-m", "scripts.sync_cust_from_web"],
                cwd=_root, capture_output=True, timeout=300, creationflags=_flags,
            )
            if proc.returncode == 0:
                print("[cust-sync] 同步完成", flush=True)
            else:
                stderr = proc.stderr.decode("utf-8", errors="replace")[-500:] if proc.stderr else ""
                print(f"[cust-sync] 同步失敗: {stderr}", flush=True)
                _notify_sync_failure("Ecount 客戶名單同步", stderr)
        except Exception as e:
            print(f"[cust-sync] 同步失敗: {e}", flush=True)
            _notify_sync_failure("Ecount 客戶名單同步", str(e))


_PICKUP_LOCK_PATH = Path(__file__).parent / "data" / ".pickup_notify.lock"

async def _pickup_notify_loop():
    """每天 16:00 和 22:00 檢查：客戶訂單全部備好 → 通知取貨（公休日不跑）"""
    _NOTIFY_HOURS = [16, 22]
    while True:
        now = datetime.now()
        # 找下一個觸發時間
        targets = []
        for h in _NOTIFY_HOURS:
            t = now.replace(hour=h, minute=0, second=0, microsecond=0)
            if t > now:
                targets.append(t)
        if not targets:
            # 今天的都過了 → 明天第一個
            tomorrow = now + timedelta(days=1)
            targets = [tomorrow.replace(hour=_NOTIFY_HOURS[0], minute=0, second=0, microsecond=0)]
        target = min(targets)
        await asyncio.sleep((target - now).total_seconds())
        # 公休日不跑
        if datetime.now().isoweekday() not in settings.business_days_list():
            print("[pickup-notify] 今天公休，跳過", flush=True)
            continue
        # 用檔案鎖防止多進程同時跑
        import time as _t_lock
        _lock_file = _PICKUP_LOCK_PATH
        if _lock_file.exists():
            _lock_age = _t_lock.time() - _lock_file.stat().st_mtime
            if _lock_age < 120:  # 2 分鐘內有其他進程跑過
                print(f"[pickup-notify] 另一個進程 {int(_lock_age)} 秒前已執行，跳過", flush=True)
                continue
        _lock_file.write_text(str(_t_lock.time()))
        try:
            await asyncio.to_thread(_check_and_notify_pickup)
        except Exception as e:
            print(f"[pickup-notify] 失敗: {e}", flush=True)


def _check_and_notify_pickup():
    """
    掃 notify_store 裡每個客戶的登記品項，檢查庫存是否到齊：
    - 可售庫存 > 0 → 到貨 ✓
    - 可售庫存 = 0 且 未出貨 = 庫存數量（貨在倉庫被訂單佔住）→ 到貨 ✓
    全部品項都到齊 → 通知客戶取貨
    有 LINE ID → push 客戶，沒有 → 推內部群
    """
    from storage.notify import notify_store
    from storage.customers import customer_store
    from config import settings
    from linebot.v3.messaging import MessagingApi, PushMessageRequest, TextMessage, ApiClient
    from collections import defaultdict
    import json, random

    # 通知前先更新所有資料
    print("[pickup-notify] 更新資料中...", flush=True)
    import time as _t

    _ADMIN_UID = settings.ADMIN_LINE_UID

    def _notify_admin_sync_fail(reason: str):
        """同步失敗時通知管理員"""
        print(f"[pickup-notify] ⚠️ {reason}，跳過到貨通知", flush=True)
        try:
            from linebot.v3.messaging import PushMessageRequest
            from linebot.v3.messaging import Configuration as _MsgCfg
            with ApiClient(_MsgCfg(access_token=settings.LINE_CHANNEL_ACCESS_TOKEN)) as _ac:
                _api = MessagingApi(_ac)
                _api.push_message(PushMessageRequest(
                    to=_ADMIN_UID,
                    messages=[TextMessage(text=f"⚠️ 到貨通知排程失敗\n{reason}\n今日到貨通知已跳過")],
                ))
        except Exception:
            pass

    # 1. 庫存同步（Excel 下載）
    avail_path = Path(__file__).parent / "data" / "available.json"
    _avail_ok = False
    try:
        import subprocess as _sp_pn
        _python_pn = _sys.executable
        _root_pn = str(Path(__file__).parent)
        _flags_pn = _sp_pn.CREATE_NO_WINDOW if _sys.platform == "win32" else 0
        _proc_avail = _sp_pn.run([_python_pn, "-m", "scripts.auto_sync_unfulfilled"],
                   cwd=_root_pn, timeout=180, creationflags=_flags_pn,
                   capture_output=True, text=True)
        _stdout_avail = _proc_avail.stdout or ""
        if "更新" in _stdout_avail and "筆" in _stdout_avail:
            _avail_ok = True
            print("[pickup-notify] 庫存已更新", flush=True)
        elif avail_path.exists() and _t.time() - avail_path.stat().st_mtime < 60 * 60:
            # 最近 1 小時內有更新過，視為 OK
            _avail_ok = True
            print("[pickup-notify] 庫存使用既有資料（1 小時內）", flush=True)
        else:
            print(f"[pickup-notify] 庫存同步可能失敗: {_stdout_avail[-200:]}", flush=True)
    except Exception as _e_avail:
        print(f"[pickup-notify] 庫存同步失敗: {_e_avail}", flush=True)

    if not _avail_ok:
        _notify_admin_sync_fail("庫存 Excel 同步失敗")
        return

    # 2. 未備貨 + 已備貨未取（Excel 下載）
    _unfulfilled_path = Path(__file__).parent / "data" / "unfulfilled_orders.json"
    _unfulfilled_ok = False
    try:
        _proc_pn = _sp_pn.run([_python_pn, "-m", "scripts.sync_unfulfilled"],
                   cwd=_root_pn, timeout=120, creationflags=_flags_pn,
                   capture_output=True, text=True,
                   encoding="utf-8", errors="replace")
        _stdout_pn = _proc_pn.stdout or ""
        if "✓" in _stdout_pn and "已存" in _stdout_pn:
            _unfulfilled_ok = True
            print("[pickup-notify] 未備貨+未取已更新", flush=True)
        elif _unfulfilled_path.exists() and _t.time() - _unfulfilled_path.stat().st_mtime < 10 * 60:
            _unfulfilled_ok = True
            print("[pickup-notify] 未備貨使用既有資料（1 小時內）", flush=True)
        else:
            print(f"[pickup-notify] 未備貨同步可能失敗: {_stdout_pn[-200:]}", flush=True)
    except Exception as _e_sync:
        print(f"[pickup-notify] 未備貨同步失敗: {_e_sync}", flush=True)
        if _unfulfilled_path.exists() and _t.time() - _unfulfilled_path.stat().st_mtime < 10 * 60:
            _unfulfilled_ok = True
            print("[pickup-notify] 未備貨使用既有資料（1 小時內）", flush=True)

    if not _unfulfilled_ok:
        _notify_admin_sync_fail("未備貨 Excel 同步失敗")
        return

    avail = json.loads(avail_path.read_text(encoding="utf-8"))

    # 取所有待通知的登記
    pending = notify_store.get_pending(source=None)
    if not pending:
        print("[pickup-notify] 沒有待通知的登記", flush=True)
        return

    # 按 user_id 分組
    by_user = defaultdict(list)
    for p in pending:
        by_user[p["user_id"]].append(p)

    def _is_ready(prod_cd: str) -> bool:
        d = avail.get(prod_cd.upper())
        if not d:
            return False
        if isinstance(d, dict):
            available = d.get("available", 0)
            unfilled = d.get("unfilled", 0)
            balance = d.get("balance", 0)
        else:
            return d > 0
        if available > 0:
            return True
        if available == 0 and balance > 0 and unfilled == balance:
            return True
        return False

    print(f"[pickup-notify] 檢查 {len(by_user)} 位客戶的 {len(pending)} 筆登記...", flush=True)

    no_line_id = []
    notified_count = 0
    notified_list = []  # 已自動通知的客戶清單

    from linebot.v3.messaging import Configuration as _MsgConfig
    _cfg_pickup = _MsgConfig(access_token=settings.LINE_CHANNEL_ACCESS_TOKEN)
    with ApiClient(_cfg_pickup) as api_client:
        line_api = MessagingApi(api_client)

        # 預載未備貨+預購資料（避免每個客戶重複查）
        from handlers.internal import _load_unfulfilled, _unfulfilled_needs_refresh, _refresh_unfulfilled
        from handlers.inventory import _check_preorder
        if _unfulfilled_needs_refresh():
            _refresh_unfulfilled()
        unfulfilled = _load_unfulfilled()

        for uid, items in by_user.items():
            # 分離：預購品 vs 非預購品
            preorder_items = [p for p in items if _check_preorder(p["prod_code"])]
            non_preorder_items = [p for p in items if not _check_preorder(p["prod_code"])]

            # 沒有非預購品 → 全是預購，等到貨再說
            if not non_preorder_items:
                continue

            # 檢查非預購品是否全部到齊
            all_non_preorder_ready = all(_is_ready(p["prod_code"]) for p in non_preorder_items)
            if not all_non_preorder_ready:
                continue

            # 檢查客戶是否還有非預購的未備貨訂單
            # 收集所有可能的客戶名稱（real_name、display_name、Ecount 客戶名）
            cust = customer_store.get_by_line_id(uid)
            _name_candidates = set()
            if uid.startswith("ecount:"):
                _name_candidates.add(uid.split(":", 1)[1])
            if cust:
                for _nk in ("real_name", "display_name", "chat_label"):
                    _nv = (cust.get(_nk) or "").strip()
                    if _nv:
                        _name_candidates.add(_nv)
                # 也用 Ecount 客戶代碼查 Ecount 客戶名
                _ec_cd = (cust.get("ecount_cust_cd") or "").strip()
                if _ec_cd:
                    from handlers.internal import _load_ec_customers
                    for _ec in _load_ec_customers():
                        if _ec.get("code") == _ec_cd:
                            _ec_name = (_ec.get("name") or "").strip()
                            if _ec_name:
                                _name_candidates.add(_ec_name)
                            break
            non_preorder_uf = [
                o for o in unfulfilled
                if any(nc and nc in o.get("customer", "") for nc in _name_candidates)
                and not _check_preorder(o.get("code", ""))
            ]
            if non_preorder_uf:
                continue

            # 非預購品全部到齊 → 通知（只列非預購品，預購品留著等）
            cust_name = (cust.get("real_name") or cust.get("display_name") or uid[:10]) if cust else uid[:10]

            # 只列非預購品項
            item_list = "\n".join(
                f"  • {p['prod_name'][:20]} × {p.get('qty_wanted', 1)}" for p in non_preorder_items
            )
            # 如果還有預購品，加提示
            po_note = ""
            if preorder_items:
                po_names = "、".join(p['prod_name'][:10] for p in preorder_items)
                po_note = f"\n\n（{po_names} 是預購品，到貨會再通知您）"
            _auto_note = "\n\n此訊息為自動通知！如已經取貨～請自動忽略嘿～感謝您～"
            notify_msg = random.choice([
                f"老闆您好～您訂的貨都到齊囉！\n{item_list}\n\n方便的時候可以來取貨哦{po_note}{_auto_note}",
                f"您好～通知您一下，以下品項都到貨了：\n{item_list}\n\n有空過來拿就可以囉～{po_note}{_auto_note}",
                f"老闆～您等的貨都到了！\n{item_list}\n\n歡迎隨時來取貨{po_note}{_auto_note}",
            ])

            if uid and not uid.startswith("_") and not uid.startswith("ecount:"):
                # 再查一次確認還是 pending（防止重複通知）
                _still_pending = [p for p in non_preorder_items
                                  if notify_store.get_status(p["id"]) == "pending"]
                if not _still_pending:
                    print(f"[pickup-notify] ⏭ {cust_name} 已通知過，跳過", flush=True)
                    continue
                # 先標記，避免重啟時重複通知
                for p in _still_pending:
                    notify_store.mark_notified(p["id"])
                try:
                    line_api.push_message(PushMessageRequest(
                        to=uid, messages=[TextMessage(text=notify_msg)]
                    ))
                    print(f"[pickup-notify] ✅ 通知 {cust_name}（{len(_still_pending)} 品項）", flush=True)
                    # 記入 chat_history，讓客戶回覆時 Claude 知道是在回應到貨通知
                    try:
                        from services.claude_ai import add_chat_history
                        add_chat_history(uid, "bot", notify_msg)
                    except Exception:
                        pass
                    # 存到貨 snapshot —— 客戶問「多少錢」時直接回總金額（純商品不含運）
                    try:
                        from storage import arrival_snapshot as _ars
                        from services.ecount import ecount_client as _ec_snap
                        _snap_products = []
                        for p in non_preorder_items:
                            _code = p.get("prod_code", "")
                            _qty = p.get("qty_wanted", 1) or 1
                            _up = 0
                            if _code:
                                try:
                                    _pr = _ec_snap.get_price(_code)
                                    if _pr and _pr.get("price"):
                                        _up = int(_pr["price"])
                                except Exception:
                                    pass
                            _snap_products.append({
                                "code": _code, "name": p.get("prod_name", ""),
                                "qty": _qty, "unit_price": _up,
                            })
                        _ars.set_snapshot(uid, _snap_products)
                    except Exception as _se:
                        print(f"[pickup-notify] snapshot 失敗: {_se}", flush=True)
                    notified_count += 1
                    notified_list.append({
                        "name": cust_name,
                        "items": [{"prod_name": p["prod_name"], "qty": p.get("qty_wanted", 1)} for p in non_preorder_items],
                    })
                except Exception as e:
                    print(f"[pickup-notify] ❌ 通知 {cust_name} 失敗: {e}", flush=True)
            else:
                no_line_id.append((cust_name, non_preorder_items))
                for p in non_preorder_items:
                    notify_store.mark_notified(p["id"])

        # 沒有 LINE ID 的（含 ecount 客戶）→ 整理成一份 push 給管理員
        if no_line_id:
            _lines = ["📦 以下客戶貨到齊了，請手動通知："]
            for _nl_name, _nl_items in no_line_id:
                _item_str = "、".join(f"{p['prod_name'][:15]}×{p.get('qty_wanted',1)}" for p in _nl_items)
                _lines.append(f"  👤 {_nl_name}：{_item_str}")
            try:
                line_api.push_message(PushMessageRequest(
                    to=settings.HELPER_LINE_UID,
                    messages=[TextMessage(text="\n".join(_lines))],
                ))
                print(f"[pickup-notify] 已推送 {len(no_line_id)} 位需手動通知給管理員", flush=True)
            except Exception as _e_adm:
                print(f"[pickup-notify] 推送管理員失敗: {_e_adm}", flush=True)

    # 儲存結果供 admin 查看（保留近 2 天）
    global _pickup_notify_results
    now_ts = datetime.now().isoformat()
    cutoff = (datetime.now() - timedelta(days=2)).isoformat()

    # 追加這次自動通知的到 history，清掉 2 天前的
    old_history = _pickup_notify_results.get("notified_history", [])
    for n in notified_list:
        n["timestamp"] = now_ts
    old_history = [h for h in old_history if h.get("timestamp", "") >= cutoff]
    old_history.extend(notified_list)

    _pickup_notify_results = {
        "timestamp": now_ts,
        "notified_history": old_history,
        "no_line_id": [
            {"name": name, "items": [{"prod_name": p["prod_name"], "qty": p.get("qty_wanted", 1)} for p in items]}
            for name, items in no_line_id
        ],
    }
    print(f"[pickup-notify] 完成，通知 {notified_count} 位客戶，{len(no_line_id)} 位需手動通知", flush=True)


_REBATE_NOTIFY_LOCK_PATH = Path(__file__).parent / "data" / ".rebate_notify.lock"


async def _rebate_notify_loop():
    """每月 20 號 14:00 自動推送回饋金進度（get_approaching_customers 的名單）+ 管理員總表。
    每月只跑一次，lock 檔存 year-month 避免重複。"""
    while True:
        now = datetime.now()
        target = now.replace(hour=14, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())

        now = datetime.now()
        if now.day != 20:
            continue

        month_key = now.strftime("%Y-%m")
        try:
            if _REBATE_NOTIFY_LOCK_PATH.exists():
                stored = _REBATE_NOTIFY_LOCK_PATH.read_text(encoding="utf-8").strip()
                if stored == month_key:
                    print(f"[rebate-notify] {month_key} 本月已執行過，跳過", flush=True)
                    continue
        except Exception:
            pass

        print(f"[rebate-notify] 觸發 20 號自動推送 ({month_key})", flush=True)
        try:
            from services.rebate_push import run as rebate_push_run
            from linebot.v3.messaging import (
                MessagingApi, ApiClient, Configuration as _Cfg,
            )
            with ApiClient(_Cfg(access_token=settings.LINE_CHANNEL_ACCESS_TOKEN)) as _ac:
                _api = MessagingApi(_ac)
                results = await asyncio.to_thread(
                    rebate_push_run, _api, False, True,
                )
            _REBATE_NOTIFY_LOCK_PATH.parent.mkdir(exist_ok=True)
            _REBATE_NOTIFY_LOCK_PATH.write_text(month_key, encoding="utf-8")
            print(
                f"[rebate-notify] 完成：推 {len(results['pushed'])} 位、"
                f"未備註 {len(results['need_designation'])} 組、"
                f"找不到 LINE {len(results['no_line'])} 位",
                flush=True,
            )
        except Exception as e:
            print(f"[rebate-notify] 失敗: {e}", flush=True)


async def _sync_cust_ecount_loop():
    """每日 13:10 自動同步 Ecount ↔ customers.db（手機比對、自動建 Ecount 客戶）"""
    while True:
        now = datetime.now()
        target = now.replace(hour=13, minute=10, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        try:
            import subprocess as _sp
            _python = _sys.executable
            _root = str(Path(__file__).parent)
            _flags = _sp.CREATE_NO_WINDOW if _sys.platform == "win32" else 0
            print("[cust-ecount] 每日 13:10 自動同步 Ecount ↔ customers.db...")
            proc = await asyncio.to_thread(
                _sp.run, [_python, "-m", "scripts.sync_cust_ecount"],
                cwd=_root, capture_output=True, timeout=300, creationflags=_flags,
            )
            if proc.returncode == 0:
                stdout = proc.stdout.decode("utf-8", errors="replace") if proc.stdout else ""
                print(f"[cust-ecount] 同步完成", flush=True)
                if stdout.strip():
                    print(stdout.strip(), flush=True)
            else:
                stderr = proc.stderr.decode("utf-8", errors="replace")[-500:] if proc.stderr else ""
                print(f"[cust-ecount] 同步失敗: {stderr}", flush=True)
                _notify_sync_failure("Ecount ↔ DB 客戶同步", stderr)
        except Exception as e:
            print(f"[cust-ecount] 同步失敗: {e}", flush=True)
            _notify_sync_failure("Ecount ↔ DB 客戶同步", str(e))


async def _queue_processor_loop():
    """每日 10:00 處理離峰佇列"""
    while True:
        now = datetime.now()
        target = now.replace(hour=10, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        try:
            await _process_queued_messages()
        except Exception as e:
            print(f"[queue] 處理失敗: {e}")


async def _check_restock_notifications():
    """
    檢查所有等待到貨通知的記錄，查詢 Ecount 庫存：
    有貨 → push 通知客戶 + 標記 notified。
    """
    from services.ecount import ecount_client

    # 先同步庫存，確保資料是最新的
    try:
        print("[notify] 先同步庫存資料...", flush=True)
        ecount_client._sync_and_wait()
        print("[notify] 庫存同步完成", flush=True)
    except Exception as _sync_e:
        print(f"[notify] 庫存同步失敗（繼續使用既有資料）: {_sync_e}", flush=True)

    pending = notify_store.get_pending(source=None)  # 全部（客戶+內部群登記）
    if not pending:
        print("[notify] 無等待通知記錄")
        return

    print(f"[notify] 開始檢查 {len(pending)} 筆到貨通知...")
    notified_count = 0

    line_api = _line_api

    if True:  # preserve indentation block
        for record in pending:
            try:
                result = ecount_client.lookup(record["prod_code"])
                qty = result.get("qty") if result else None
                balance = result.get("balance") or 0 if result else 0
                unfilled = result.get("unfilled") or 0 if result else 0
                # 條件1: 可售>0 || 條件2: 可售=0 但倉庫有貨且全部被訂單佔住（balance>0 且 balance==unfilled）
                _should_notify = (qty and qty > 0) or (qty == 0 and balance > 0 and balance == unfilled)
                if _should_notify:
                    source = record.get("source", "customer")
                    qty_wanted = record.get("qty_wanted", 1)

                    if source == "staff":
                        # 內部群登記：用訂購格式通知
                        # 換算箱數顯示
                        item = ecount_client.get_product_cache_item(record["prod_code"])
                        box_qty = (item.get("box_qty") or 0) if item else 0
                        prod_unit = (item.get("unit") or "") if item else ""
                        if box_qty > 1 and qty_wanted >= box_qty and qty_wanted % box_qty == 0:
                            qty_display = f"{qty_wanted // box_qty}箱"
                        else:
                            qty_display = f"{qty_wanted}{prod_unit or '個'}"
                        msg = (
                            f"老闆您好，您之前訂的貨已經到了\n"
                            f"{record['prod_name']}（{record['prod_code']}）× {qty_display}"
                        )
                    else:
                        # 客戶自己登記：用原本的到貨通知格式
                        msg = tone.restock_back_in_stock(
                            name=record["prod_name"],
                            code=record["prod_code"],
                        )

                    # 決定推送目標
                    _target_uid = record["user_id"]
                    _is_ecount_only = _target_uid.startswith("ecount:")
                    if _is_ecount_only:
                        # 無 LINE ID → 推到內部群
                        _target_uid = settings.LINE_GROUP_ID
                        _cust_name = record["user_id"].replace("ecount:", "")
                        msg = (
                            f"📬 到貨通知\n"
                            f"客戶：{_cust_name}\n"
                            f"{record['prod_name']}（{record['prod_code']}）× {qty_display if source == 'staff' else qty_wanted}"
                        )

                    if _push_quota_exhausted:
                        print(f"[notify] 月額度已用完，跳過 push {record['user_id'][:15]}...", flush=True)
                        continue
                    try:
                        line_api.push_message(
                            PushMessageRequest(
                                to=_target_uid,
                                messages=[TextMessage(text=msg)],
                            )
                        )
                    except Exception as _notify_push_e:
                        if _is_quota_429(_notify_push_e):
                            _mark_push_exhausted()
                            continue
                        raise
                    notify_store.mark_notified(record["id"])
                    notified_count += 1
                    _dest = "內部群" if _is_ecount_only else record["user_id"][:10]
                    print(
                        f"[notify] OK 已通知 {_dest}... "
                        f"-> {record['prod_name']} 庫存={qty} (source={source})"
                    )
                else:
                    print(
                        f"[notify] 仍無貨：{record['prod_name']}（{record['prod_code']}）"
                    )
            except Exception as e:
                print(f"[notify] FAIL 通知失敗 id={record['id']}: {e}", flush=True)

    print(f"[notify] 完成，共通知 {notified_count} 筆")


async def _sync_failure_notify_loop():
    """每日 17:00 統一通知今日排程失敗"""
    while True:
        now = datetime.now()
        target = now.replace(hour=17, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        if not _sync_failures:
            continue
        try:
            lines = [f"⚠️ 今日排程失敗 {len(_sync_failures)} 項："]
            for t, name, err in _sync_failures:
                lines.append(f"• {t} {name}\n  {err}")
            msg = "\n".join(lines)
            if not _push_quota_exhausted:
                _line_api.push_message(PushMessageRequest(
                    to=settings.ADMIN_LINE_UID,
                    messages=[TextMessage(text=msg)],
                ))
                print(f"[sync-fail] 已通知管理員，{len(_sync_failures)} 項失敗", flush=True)
        except Exception as e:
            print(f"[sync-fail] 通知失敗: {e}", flush=True)
        _sync_failures.clear()


async def _restock_notify_loop():
    """每日 21:00 執行到貨通知"""
    while True:
        now = datetime.now()
        target = now.replace(hour=21, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())
        try:
            await _check_restock_notifications()
        except Exception as e:
            print(f"[notify] 排程執行失敗: {e}")
            _notify_sync_failure("到貨通知檢查", str(e))


async def _followup_loop():
    """每小時檢查對話狀態：24h 提醒 / 48h 清除"""
    await asyncio.sleep(60)   # 啟動後 1 分鐘才第一次跑（避免啟動擠塞）
    while True:
        _check_quota_reset()
        try:
            from handlers.followup import check_and_followup
            result = await asyncio.to_thread(check_and_followup, _line_api)
            if result["reminded"] or result["expired"]:
                print(f"[followup] 提醒 {result['reminded']} 人，清除 {result['expired']} 筆過期狀態")
        except Exception as e:
            print(f"[followup] 排程執行失敗: {e}")
            _notify_sync_failure("Followup 提醒檢查", str(e))
        await asyncio.sleep(3600)   # 每小時一次


def _startup_verify():
    """啟動時確認各模組功能是否正確載入，結果印至 log"""
    import inspect
    checks = []
    try:
        from handlers.internal import handle_ambiguous_resolve
        checks.append("✅ handle_ambiguous_resolve")
    except Exception as e:
        checks.append(f"❌ handle_ambiguous_resolve ({e})")
    try:
        src = inspect.getsource(
            __import__("services.ecount", fromlist=["EcountClient"]).EcountClient._ensure_product_cache
        )
        if "_ZX_RE" in src:
            checks.append("✅ Ecount Z+英文過濾")
        else:
            checks.append("❌ Ecount Z+英文過濾（未套用）")
    except Exception as e:
        checks.append(f"❌ Ecount Z+英文過濾 ({e})")
    try:
        src2 = inspect.getsource(_msg_buf_flush_inner)
        checks.append("✅ 內部群 try/except" if "內部群處理例外" in src2 else "❌ 內部群 try/except（未套用）")
        checks.append("✅ ambiguous dispatch" if "handle_ambiguous_resolve" in src2 else "❌ ambiguous dispatch（未套用）")
    except Exception as e:
        checks.append(f"❌ dispatch 檢查 ({e})")
    print("[startup] 功能確認：", flush=True)
    for c in checks:
        print(f"  {c}", flush=True)


@asynccontextmanager
async def lifespan(app: FastAPI):
    _startup_verify()               # 確認各模組是否正確載入
    queue_store.init()              # 建立離峰佇列 table
    notify_store._init_db()         # 確保到貨通知 table 存在
    visit_store.init()              # 建立客戶到店記錄 table
    # check_and_refresh 改背景執行，避免 Ecount 連不上時阻塞啟動
    import threading as _t
    _t.Thread(target=check_and_refresh, daemon=True).start()
    state_manager.restore_from_db() # 從 SQLite 恢復對話狀態
    _restore_txt_buffer()           # 恢復 reload 前未處理的文字 buffer
    asyncio.create_task(_refresh_data_loop())
    # asyncio.create_task(_queue_processor_loop())  # 離峰佇列已停用
    # asyncio.create_task(_restock_notify_loop())  # 已停用，改用 14:00 _pickup_notify_loop
    # asyncio.create_task(_followup_loop())  # 24h 提醒已停用
    asyncio.create_task(_midnight_inventory_check_loop())
    asyncio.create_task(_rebate_sync_loop())
    asyncio.create_task(_sync_ecount_customers_loop())
    asyncio.create_task(_sync_cust_ecount_loop())
    asyncio.create_task(_sync_failure_notify_loop())
    asyncio.create_task(_cart_cleanup_loop())
    asyncio.create_task(_pickup_notify_loop())
    asyncio.create_task(_rebate_notify_loop())
    asyncio.create_task(_competitor_sync_loop())
    # 離峰佇列已停用
    yield
    # ── shutdown：持久化未處理的文字 buffer ──
    _persist_txt_buffer()


# ── 文字 buffer 持久化（防 reload 丟訊息）────────────────────────
_TXT_BUF_PERSIST_PATH = _BASE_DIR / "data" / "txt_buffer_pending.json"


def _persist_txt_buffer():
    """shutdown 時把未處理的 _msg_buffer 存到 JSON（只存文字部分，媒體無法序列化）"""
    import json as _json
    with _msg_buffer_lock:
        pending = {}
        for uid, entry in _msg_buffer.items():
            if not entry.get("lines"):
                continue  # 只有媒體、沒文字 → 不持久化
            try:
                entry["timer"].cancel()
            except Exception:
                pass
            pending[uid] = {
                "lines":       entry["lines"],
                "context":     entry["context"],
                "group_id":    entry.get("group_id"),
            }
    if pending:
        try:
            _TXT_BUF_PERSIST_PATH.write_text(
                _json.dumps(pending, ensure_ascii=False), encoding="utf-8")
            print(f"[shutdown] 已保存 {len(pending)} 筆未處理的文字 buffer", flush=True)
        except Exception as e:
            print(f"[shutdown] buffer 保存失敗: {e}", flush=True)


def _restore_txt_buffer():
    """startup 時恢復未處理的 _msg_buffer 並立即 flush"""
    import json as _json
    if not _TXT_BUF_PERSIST_PATH.exists():
        return
    try:
        pending = _json.loads(_TXT_BUF_PERSIST_PATH.read_text(encoding="utf-8"))
        _TXT_BUF_PERSIST_PATH.unlink()  # 讀完就刪
        if not pending:
            return
        print(f"[startup] 恢復 {len(pending)} 筆未處理的文字 buffer", flush=True)
        for uid, entry in pending.items():
            # 塞回 buffer 然後立即 flush
            with _msg_buffer_lock:
                _msg_buffer[uid] = {
                    "lines":       entry["lines"],
                    "media":       [],
                    "context":     entry["context"],
                    "group_id":    entry.get("group_id"),
                    "reply_token": None,  # reload 後 token 已過期
                    "quoted_msg_id": None,
                }
                # 不設 timer，直接啟動 flush
                _msg_buffer[uid]["timer"] = _threading.Timer(0, lambda: None)
            _threading.Thread(target=_msg_buf_flush, args=(uid,), daemon=True).start()
    except Exception as e:
        print(f"[startup] buffer 恢復失敗: {e}", flush=True)


app = FastAPI(title="LINE Customer Service Bot", lifespan=lifespan)

_SERVER_START_TIME = datetime.now()

@app.get("/health")
def health():
    """Server 啟動時間確認用，更新後呼叫此端點驗證是否為新版"""
    import os
    return {
        "status": "ok",
        "started_at": _SERVER_START_TIME.strftime("%Y-%m-%d %H:%M:%S"),
        "pid": os.getpid(),
    }

# ── 靜態檔案（產品圖片等）────────────────────────────────────────────
_STATIC_DIR = _BASE_DIR / "static"
_PRODUCTS_IMG_DIR = _STATIC_DIR / "products"
_PRODUCTS_IMG_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/static", StaticFiles(directory=str(_STATIC_DIR)), name="static")
# 產品照片靜態路由（供 LINE push image 用）
_PRODUCT_PHOTO_STATIC = Path(r"H:\其他電腦\我的電腦\小蠻牛\產品照片")
if _PRODUCT_PHOTO_STATIC.exists():
    app.mount("/product-photo", StaticFiles(directory=str(_PRODUCT_PHOTO_STATIC)), name="product-photo")


# ── Admin 介面密碼保護（HTTP Basic Auth）─────────────
@app.middleware("http")
async def admin_auth_middleware(request: Request, call_next):
    if request.url.path.startswith("/admin"):
        auth = request.headers.get("Authorization", "")
        authorized = False
        if auth.startswith("Basic "):
            try:
                decoded  = base64.b64decode(auth[6:]).decode("utf-8")
                username, _, password = decoded.partition(":")
                ok_user = secrets.compare_digest(username, settings.ADMIN_USER)
                ok_pass = secrets.compare_digest(password, settings.ADMIN_PASS)
                authorized = ok_user and ok_pass
            except Exception:
                pass
        if not authorized:
            return Response(
                content="需要登入",
                status_code=401,
                headers={"WWW-Authenticate": 'Basic realm="Admin"'},
            )
    return await call_next(request)


_PHONE_RE = re.compile(r"09\d{8}")


def _user_phone(user_id: str) -> str:
    """從 customer_store 取得客戶手機號碼，供 save_order HP_NO 使用。"""
    info = customer_store.get_by_line_id(user_id)
    return (info or {}).get("phone", "") or ""
_ADDR_RE  = re.compile(
    r"[\u4e00-\u9fff]{2,5}[市縣][\u4e00-\u9fff]{1,5}[區鄉鎮市]"
    r"[\u4e00-\u9fff\d\-]+[路街][^\n,，。]{0,40}"
)


def _auto_save_contact_info(user_id: str, text: str) -> None:
    """從訊息自動擷取電話/住址並儲存到客戶資料"""
    phones = _PHONE_RE.findall(text)
    if phones:
        customer_store.update_phone(user_id, phones[0])
    addr = _ADDR_RE.search(text)
    if addr:
        customer_store.update_address(user_id, addr.group(0).strip())

_YES_KW = set(AFFIRMATIVE_KEYWORDS) | {"好了", "下單"}  # FSM 答 yes 情境：含 CHECKOUT 的「好了/下單」
_NO_KW = {"不", "否", "no", "NO", "不要", "不用", "取消", "算了", "不訂", "不對", "錯了", "不是", "不行", "等不了", "太久", "換一個", "其他"}

_configuration = Configuration(access_token=settings.LINE_CHANNEL_ACCESS_TOKEN)
_api_client = ApiClient(_configuration)
_line_api = MessagingApi(_api_client)
_webhook_handler = WebhookHandler(settings.LINE_CHANNEL_SECRET)


def _resolve_one(kind: str, item_id: int) -> str:
    """標記單一項目並回傳結果文字"""
    if kind == "P":
        ok = payment_store.resolve(item_id)
        return f"✅ #P{item_id} 轉帳確認完成" if ok else f"⚠️ #P{item_id} 找不到或已標記過"
    elif kind == "R":
        ok = restock_store.update_status(item_id, "confirmed")
        return f"✅ #R{item_id} 調貨處理完成" if ok else f"⚠️ #R{item_id} 找不到（可能已處理過）"
    elif kind == "D":
        ok = delivery_store.resolve(item_id)
        return f"✅ #D{item_id} 配送詢問處理完成" if ok else f"⚠️ #D{item_id} 找不到（可能已處理過）"
    elif kind == "I":
        # 先取 user_id，標記對話紀錄
        _issue = issue_store.get_by_id(item_id)
        ok = issue_store.resolve(item_id)
        if ok and _issue:
            from services.claude_ai import add_chat_history
            _issue_uid = _issue.get("user_id", "")
            add_chat_history(_issue_uid, "bot", "（以上問題已由真人客服處理完成）")
            # 背景去 LINE OA 讀真人回覆，存到 chat_history
            if _issue_uid:
                import threading as _t_resolve
                _OA_SKIP = {"已讀", "讀", ""}
                def _sync_resolve_chat():
                    try:
                        _cust = customer_store.get_by_line_id(_issue_uid)
                        _cname = (_cust.get("chat_label") or _cust.get("real_name") or _cust.get("display_name") or "") if _cust else ""
                        if _cname:
                            from services.line_oa_chat import read_chat_sync
                            _msgs = read_chat_sync(_cname, max_messages=15)
                            if _msgs:
                                _staff = [m for m in _msgs if m["role"] == "staff"
                                          and m["text"].replace("已讀", "").strip()
                                          and m["text"].strip() not in _OA_SKIP]
                                for m in _staff[-5:]:
                                    _clean_text = m['text'].replace("已讀", "").strip()
                                    if _clean_text:
                                        add_chat_history(_issue_uid, "bot", f"（真人回覆）{_clean_text}")
                                print(f"[resolve] LINE OA 對話已同步：{_cname} {len(_staff)} 則真人回覆", flush=True)
                    except Exception as e:
                        print(f"[resolve] LINE OA 對話同步失敗: {e}", flush=True)
                _t_resolve.Thread(target=_sync_resolve_chat, daemon=True).start()
        return f"✅ #I{item_id} 問題處理完成" if ok else f"⚠️ #I{item_id} 找不到（可能已處理過）"
    elif kind == "Q":
        ok = pending_store.mark_answered(item_id)
        return f"✅ #Q{item_id} 商品查詢已回覆" if ok else f"⚠️ #Q{item_id} 找不到（可能已處理過）"
    return ""


_PENDING_LIST_RE = re.compile(r'^(清單|待確認|待處理|清單查詢|待確認清單|待處理清單)$')

def _get_ngrok_url_sync() -> str | None:
    """同步取得 ngrok 目前公開網址（daemon thread 可用）"""
    try:
        import httpx
        r = httpx.get("http://localhost:4040/api/tunnels", timeout=2.0)
        tunnels = r.json().get("tunnels", [])
        for t in tunnels:
            if t.get("proto") == "https":
                return t["public_url"]
        if tunnels:
            return tunnels[0]["public_url"]
    except Exception:
        pass
    return None


def _handle_visit_query_command(text: str) -> str | None:
    """內部群「誰要來 / 哪些客人要來」→ 回傳到店預約清單"""
    if not is_visit_query(text):
        return None
    return handle_visit_query()


def _handle_visit_resolve(text: str) -> str | None:
    """內部群「✅ V1」→ 標記客人已到店"""
    if not re.search(r"[✅☑️√v]", text) and "已到" not in text:
        return None
    m = re.search(r"[Vv]\s*(\d+)", text)
    if not m:
        return None
    vid = int(m.group(1))
    ok = visit_store.mark_visited(vid)
    return f"✅ #V{vid} 已到店" if ok else f"找不到 #V{vid}"


def _handle_pending_list_command(text: str) -> str | None:
    """
    偵測「清單」/「待確認」等指令 → 回傳待處理清單文字 + admin 網址。
    """
    if not _PENDING_LIST_RE.match(text.strip()):
        return None
    body = build_pending_text()
    if not body:
        return "目前沒有待處理項目"
    url = _get_ngrok_url_sync()
    suffix = f"\n\n🌐 管理介面：{url}/admin" if url else ""
    return body + suffix


def _handle_staff_resolve(text: str) -> str | None:
    """
    內部群組人工標記完成指令。

    支援格式：
      ✅ I2              → 單筆
      ✅ I2 I3 P1        → 多筆
      I1-I6已處理        → 範圍（同類型）
      全部已處理          → 全部未處理項目一次標記
    """
    # 沒有標記觸發詞（✅/已處理/已完成）→ 直接略過，避免誤判貨號連字號為範圍
    if not _RESOLVE_TRIGGER_RE.search(text):
        return None

    results = []

    # ── 全部已處理 ────────────────────────────────────
    if _RESOLVE_ALL_RE.search(text):
        count = 0
        for p in payment_store.get_pending():
            payment_store.resolve(p["id"]); count += 1
        for r in restock_store.get_unresolved():
            restock_store.update_status(r["id"], "confirmed"); count += 1
        for d in delivery_store.get_pending():
            delivery_store.resolve(d["id"]); count += 1
        for i in issue_store.get_pending():
            issue_store.resolve(i["id"]); count += 1
        for q in pending_store.get_pending():
            pending_store.mark_answered(q["id"]); count += 1
        return f"✅ 全部 {count} 筆待處理項目已標記完成" if count else "目前無待處理項目"

    # ── 範圍標記：I1-I6 / I1~I6 ─────────────────────
    for m in _RESOLVE_RANGE_RE.finditer(text):
        kind  = m.group(1).upper()
        start = int(m.group(2))
        end   = int(m.group(3))
        if start > end:
            start, end = end, start
        if end - start > 50:   # 防止誤操作，最多 50 筆
            results.append(f"⚠️ 範圍太大（{start}~{end}），最多一次 50 筆")
            continue
        for i in range(start, end + 1):
            results.append(_resolve_one(kind, i))

    # ── 範圍標記後不再處理單筆（避免重複）────────────
    if results:
        return "\n".join(results)

    # ── 單筆 / 多筆 ──────────────────────────────────
    has_trigger = bool(_RESOLVE_TRIGGER_RE.search(text))
    if not has_trigger:
        return None

    items = _RESOLVE_ITEM_RE.findall(text)
    if not items:
        return None

    for kind_raw, id_raw in items:
        results.append(_resolve_one(kind_raw.upper(), int(id_raw)))

    return "\n".join(results) if results else None


def _handle_spec_rebuild_command(text: str) -> str | None:
    """「規格更新」→ 立即觸發 import_specs + image_hashes 重建（背景非同步）"""
    if text.strip() not in ("規格更新", "更新規格", "重建規格", "specs rebuild"):
        return None
    from services.refresh import trigger_rebuild
    trigger_rebuild()
    return "🔄 規格DB 更新中（背景執行），完成後自動生效，約需 10~30 秒"


def _handle_bot_notify_command(text: str) -> str | None:
    """
    偵測內部群組呼叫 bot 名字的到貨通知代客登記指令。
    格式：「新北小蠻牛 王小明 AA001 需要到貨通知」

    1. 確認前綴含「新北小蠻牛」或「小蠻牛」
    2. 確認含到貨通知關鍵字
    3. 解析客戶姓名 + 貨號 → 查 DB + Ecount → 登記 notify_store
    回傳回覆文字，或 None（不是此指令）。
    """
    m_bot = _BOT_NAME_RE.match(text)
    if not m_bot:
        return None

    remainder = text[m_bot.end():].strip()
    if not any(kw in remainder for kw in _STAFF_NOTIFY_KW):
        return None

    # ── 提取貨號 ──────────────────────────────────────
    m_prod = _PROD_CODE_RE_STAFF.search(remainder)
    if not m_prod:
        return "請加上貨號哦\n格式：新北小蠻牛 王小明 AA001 需要到貨通知"

    prod_code_raw = m_prod.group(0).upper()

    # ── 提取客戶姓名（移除貨號 + 關鍵字後剩下的文字）──
    name_part = _PROD_CODE_RE_STAFF.sub("", remainder)
    for kw in _STAFF_NOTIFY_KW:
        name_part = name_part.replace(kw, "")
    cust_name = name_part.strip()

    if not cust_name:
        return "請加上客戶姓名哦\n格式：新北小蠻牛 王小明 AA001 需要到貨通知"

    # ── 查 Ecount 確認產品 ─────────────────────────────
    from services.ecount import ecount_client
    result = ecount_client.lookup(prod_code_raw)
    if not result:
        return f"找不到貨號「{prod_code_raw}」，請確認一下哦"

    prod_name = result.get("name") or prod_code_raw

    # ── 查客戶 DB ─────────────────────────────────────
    matches = customer_store.search_by_name(cust_name, real_name_only=True)
    if not matches:
        return (
            f"找不到「{cust_name}」的資料\n"
            f"（客戶需要先傳訊息給我才能登記到貨通知唷）"
        )

    # 取第一個有 LINE user_id 的客戶
    matched = next((m for m in matches if m.get("line_user_id")), None)
    if not matched:
        return (
            f"「{cust_name}」尚未和我互動過，無法自動推送\n"
            f"（請客戶先傳一則訊息給我唷）"
        )

    # ── 登記到貨通知 ───────────────────────────────────
    from storage.notify import notify_store
    notify_store.add(matched["line_user_id"], prod_code_raw, prod_name, 1)

    display = (matched.get("real_name") or matched.get("display_name") or cust_name).strip()
    print(
        f"[notify] 人工登記 → {display}（{matched['line_user_id'][:10]}...）"
        f"：{prod_name}（{prod_code_raw}）"
    )
    return (
        f"✅ 已登記！\n"
        f"👤 {display}\n"
        f"📦 「{prod_name}」（{prod_code_raw}）\n"
        f"到貨後會自動推播通知哦"
    )


@app.post("/admin/push-summary")
async def push_summary():
    """手動觸發待處理清單推送"""
    await asyncio.to_thread(send_pending_summary)
    return {"status": "ok"}


@app.post("/admin/generate-ad")
async def admin_generate_ad():
    """觸發廣告圖更新（同內部群 `廣告圖更新` 指令）"""
    from handlers.ad_maker import handle_ad_update_trigger
    result = await asyncio.to_thread(handle_ad_update_trigger, "廣告圖更新", settings.LINE_GROUP_ID)
    return {"status": "ok", "message": result or "已啟動"}


@app.post("/admin/process-queue")
async def process_queue():
    """手動觸發離峰佇列補處理"""
    pending = queue_store.count_unprocessed()
    await _process_queued_messages()
    return {"status": "ok", "processed": pending}


# ══════════════════════════════════════════════════════════════════════════
# 客戶下單網頁 API（LIFF）
# ══════════════════════════════════════════════════════════════════════════

@app.get("/api/shop/products")
async def shop_products():
    """取得可售庫存 > 10 的商品，按台型分類"""
    import json as _json_shop
    from services.ecount import ecount_client as _ec_shop

    _ec_shop._ensure_product_cache()
    avail_path = Path(__file__).parent / "data" / "available.json"
    specs_path = Path(__file__).parent / "data" / "specs.json"

    avail = _json_shop.loads(avail_path.read_text(encoding="utf-8")) if avail_path.exists() else {}
    specs = _json_shop.loads(specs_path.read_text(encoding="utf-8")) if specs_path.exists() else {}

    media_dir = Path(r"H:\其他電腦\我的電腦\小蠻牛\產品照片")
    base_url = "https://xmnline.duckdns.org/product-photo"

    products = []
    for code, data in avail.items():
        if not isinstance(data, dict):
            continue
        qty = data.get("available", 0)
        # 排除耗材
        if code.upper().startswith("HH"):
            continue
        # 先取品名判斷是否為野獸國/美好（庫存 > 0 就上）
        _cache_pre = _ec_shop.get_product_cache_item(code)
        _name_pre = (_cache_pre.get("name") if _cache_pre else "") or ""
        _is_special = "野獸國" in _name_pre or "美好" in _name_pre
        # 檢查是否為預購品
        from handlers.inventory import _check_preorder
        _is_preorder = _check_preorder(code)
        if _is_special or _is_preorder:
            if qty <= 0 and not _is_preorder:
                continue
        else:
            if qty <= 10:
                continue

        cache = _ec_shop.get_product_cache_item(code)
        name = (cache.get("name") if cache else "") or code
        price = (cache.get("price") if cache else 0) or 0
        unit = (cache.get("unit") if cache else "") or "個"

        spec = specs.get(code, {})
        machine = spec.get("machine", [])
        machine_label = "、".join(machine) if machine else "通用"

        # 找第一張圖片（列表頁用，點進去再載入全部）
        image = ""
        if media_dir.exists():
            for suffix in ["A", "", "B", "C", "D", "E"]:
                for ext in [".jpg", ".jpeg", ".png"]:
                    f = media_dir / f"{code}{suffix}{ext}"
                    if f.exists():
                        _mtime = int(f.stat().st_mtime)
                        image = f"{base_url}/{f.name}?v={_mtime}"
                        break
                if image:
                    break

        products.append({
            "code": code,
            "name": name,
            "price": int(price) if price else 0,
            "unit": unit,
            "image": image,
            "machine": machine_label,
            "size": spec.get("size", ""),
            "weight": spec.get("weight", ""),
            "preorder": qty <= 0,
        })

    # 加上品類標籤
    from services.analytics import _classify
    _CATEGORY_ORDER = [
        "藍牙耳機", "音響", "行動電源", "電腦周邊", "手錶",
        "遙控車/飛機", "合金模型", "雷射/燈光",
        "三麗鷗/IP", "娃娃/絨毛", "玩具", "盲盒",
        "涼風扇", "打火機/噴火槍", "工具類", "釣具",
        "生活用品", "美容/保養", "車載用品",
        "耗材", "零食飲料", "飾品配件", "節慶商品",
        "暖風/保暖", "香薰", "其他",
    ]
    _cat_rank = {c: i for i, c in enumerate(_CATEGORY_ORDER)}

    for p in products:
        p["category"] = _classify(p["name"])
        p["cat_rank"] = _cat_rank.get(p["category"], 999)

    # 按品類排序，同品類按價格排
    products.sort(key=lambda x: (x["cat_rank"], -x["price"]))

    return {"products": products, "total": len(products)}


@app.get("/api/shop/images/{code}")
async def shop_images(code: str):
    """取得單一產品的所有圖片（最多 4 張）"""
    media_dir = Path(r"H:\其他電腦\我的電腦\小蠻牛\產品照片")
    base_url = "https://xmnline.duckdns.org/product-photo"
    images = []
    if media_dir.exists():
        _img_exts = {".jpg", ".jpeg", ".png"}
        _found = sorted(
            [f for f in media_dir.iterdir()
             if f.is_file() and f.suffix.lower() in _img_exts
             and f.stem.upper().startswith(code.upper())],
            key=lambda f: f.stem.upper()
        )
        for f in _found[:4]:
            _mtime = int(f.stat().st_mtime)
            images.append(f"{base_url}/{f.name}?v={_mtime}")
    return {"code": code, "images": images}


@app.get("/api/shop/profile")
async def shop_profile(uid: str = "", name: str = ""):
    """LIFF 用：取得客戶 real_name + ecount_cust_cd"""
    cust = None
    if uid:
        cust = customer_store.get_by_line_id(uid)
    if not cust and name:
        # userId 查不到 → 用 displayName 或 real_name 查
        matches = customer_store.search_by_name(name, real_name_only=False)
        if matches:
            cust = matches[0]
    _ec_cd = (cust.get("ecount_cust_cd") or "") if cust else ""
    # 主表空 → fallback 查子表（多地址客戶可能只有子表有代碼）
    if cust and not _ec_cd:
        try:
            _codes = customer_store.get_ecount_codes_by_db_id(cust.get("id"))
            if _codes:
                _ec_cd = (_codes[0].get("ecount_cust_cd") or "").strip()
        except Exception:
            pass
    return {
        "real_name": (cust.get("real_name") or cust.get("display_name") or "") if cust else "",
        "ecount_cust_cd": _ec_cd,
    }


@app.post("/api/shop/order")
async def shop_order(request: Request):
    """LIFF 下單：建立 Ecount 訂單 + 登記到貨通知"""
    import json as _json_order
    body = await request.json()

    line_user_id = body.get("user_id", "")
    items = body.get("items", [])  # [{"code": "Z3555", "qty": 2}, ...]

    if not line_user_id or not items:
        return {"ok": False, "error": "缺少 user_id 或 items"}

    from services.ecount import ecount_client as _ec_order
    from storage.notify import notify_store
    from storage.customers import customer_store as _cs_order

    # 取客戶 Ecount 代碼
    cust = _cs_order.get_by_line_id(line_user_id)
    cust_name = (cust.get("real_name") or cust.get("display_name") or "") if cust else ""
    cust_code = (cust.get("ecount_cust_cd") or "") if cust else ""

    if not cust_code:
        # 沒有 Ecount 代碼 → 用預設
        cust_code = settings.ECOUNT_DEFAULT_CUST_CD

    phone = (cust.get("phone") or "") if cust else ""

    # 建立訂單
    order_items = []
    for item in items:
        code = item["code"].upper()
        qty = int(item.get("qty", 1))
        cache = _ec_order.get_product_cache_item(code)
        name = (cache.get("name") if cache else "") or code
        order_items.append({"prod_cd": code, "qty": qty, "name": name})

    slip_no = _ec_order.save_order(
        cust_code=cust_code,
        items=[{"prod_cd": i["prod_cd"], "qty": i["qty"]} for i in order_items],
        phone=phone,
    )

    if not slip_no:
        return {"ok": False, "error": "訂單建立失敗"}

    # 全部登記到貨通知
    for i in order_items:
        notify_store.add(
            user_id=line_user_id,
            prod_code=i["prod_cd"],
            prod_name=i["name"],
            source="staff",
            qty_wanted=i["qty"],
        )

    return {
        "ok": True,
        "slip_no": slip_no,
        "customer": cust_name,
        "items": [{"code": i["prod_cd"], "name": i["name"], "qty": i["qty"]} for i in order_items],
    }


@app.post("/admin/full-report")
async def push_full_report(days: int = 3):
    """推送完整報表（已處理 + 未處理）到內部群組，預設顯示 3 天內已處理記錄"""
    await asyncio.to_thread(send_full_report, days=days)
    return {"status": "ok", "days": days}


@app.get("/admin/full-report")
async def get_full_report(days: int = 3):
    """直接回傳完整報表文字（不推送到 LINE，方便瀏覽器查看）"""
    report = await asyncio.to_thread(build_full_report, days=days)
    return {"report": report, "days": days}


@app.post("/admin/check-restock-notify")
async def check_restock_notify():
    """已停用 — 改用 pickup-notify 排程"""
    return {"status": "disabled", "message": "舊通知系統已停用，改用 16:00/22:00 排程"}


@app.post("/admin/update-customer-label")
async def update_customer_label(
    line_user_id: str = "", db_id: int = 0, label: str = ""
):
    """
    手動設定客戶標籤（顯示在待處理清單的名稱）。
    優先用 line_user_id；若空則用 db_id（適用於 CSV 匯入無 LINE ID 的客戶）。
    label 傳空字串表示清除。

    用法：POST /admin/update-customer-label?line_user_id=Uxxxxxxx&label=王小明
          POST /admin/update-customer-label?db_id=42&label=王小明
    """
    if line_user_id:
        ok = customer_store.update_chat_label(line_user_id, label)
        cust = customer_store.get_by_line_id(line_user_id)
    elif db_id:
        ok = customer_store.update_chat_label_by_db_id(db_id, label)
        cust = customer_store.get_by_db_id(db_id)
    else:
        raise HTTPException(status_code=400, detail="需要 line_user_id 或 db_id")
    if not ok:
        raise HTTPException(status_code=404, detail=f"找不到客戶 (line_user_id={line_user_id!r} db_id={db_id})")
    display = (cust or {}).get("display_name", "")
    print(f"[admin] 客戶標籤更新: uid={line_user_id or 'N/A'} db_id={db_id} → label={label!r} (display_name={display!r})")
    return {"status": "ok", "chat_label": label or None, "display_name": display}


@app.post("/admin/update-customer-tags")
async def update_customer_tags(db_id: int, tags: str = ""):
    """
    更新客戶分類標籤（VIP/野獸國/標準/中句/K霸）。
    tags 為逗號分隔的標籤字串，例如 "VIP,野獸國"。
    傳空字串表示清除所有標籤。

    用法：POST /admin/update-customer-tags?db_id=42&tags=VIP,野獸國
    """
    tag_list = [t.strip() for t in tags.split(",") if t.strip()] if tags.strip() else []
    ok = customer_store.update_tags_by_db_id(db_id, tag_list)
    if not ok:
        raise HTTPException(status_code=404, detail=f"找不到 db_id={db_id}")
    cust = customer_store.get_by_db_id(db_id)
    display = (cust or {}).get("display_name", "")
    print(f"[admin] 客戶分類標籤更新: db_id={db_id} → tags={tag_list!r} (display_name={display!r})")
    return {"status": "ok", "db_id": db_id, "tags": tag_list, "display_name": display}


@app.post("/admin/update-customer-real-name")
async def update_customer_real_name(line_user_id: str, real_name: str = ""):
    """
    手動更新客戶真實姓名（real_name）。
    real_name 傳空字串表示清除。

    用法：POST /admin/update-customer-real-name?line_user_id=Uxxxxxxx&real_name=王家文
    """
    import sqlite3 as _sq
    from storage.customers import DB_PATH as _CUST_DB
    with _sq.connect(str(_CUST_DB)) as _conn:
        cur = _conn.execute(
            "UPDATE customers SET real_name=? WHERE line_user_id=?",
            (real_name.strip() or None, line_user_id)
        )
        _conn.commit()
    if not cur.rowcount:
        raise HTTPException(status_code=404, detail=f"找不到 line_user_id={line_user_id}")
    print(f"[admin] 真實姓名更新: {line_user_id} → real_name={real_name!r}")
    return {"status": "ok", "line_user_id": line_user_id, "real_name": real_name.strip() or None}


# ── 客戶分類標籤設定 ──────────────────────────────────────────────────────

@app.get("/admin/tags-config")
async def get_tags_config():
    """取得目前所有分類標籤清單"""
    from storage.tags_config import load_tags
    return {"tags": load_tags()}


@app.post("/admin/tags-config/add")
async def add_tags_config(tag: str):
    """
    新增一個分類標籤。
    用法：POST /admin/tags-config/add?tag=新標籤
    """
    from storage.tags_config import add_tag
    tag = tag.strip()
    if not tag:
        raise HTTPException(status_code=400, detail="標籤名稱不可為空")
    if len(tag) > 20:
        raise HTTPException(status_code=400, detail="標籤名稱不可超過 20 字")
    tags = add_tag(tag)
    print(f"[admin] 新增分類標籤: {tag!r}")
    return {"status": "ok", "tags": tags}


@app.post("/admin/tags-config/remove")
async def remove_tags_config(tag: str):
    """
    移除一個分類標籤。
    用法：POST /admin/tags-config/remove?tag=舊標籤
    """
    from storage.tags_config import remove_tag
    tags = remove_tag(tag)
    print(f"[admin] 移除分類標籤: {tag!r}")
    return {"status": "ok", "tags": tags}


@app.get("/admin/product-images")
async def list_product_images(code: str = ""):
    """
    列出指定產品代碼的圖片清單（static/products/{CODE}/）。
    不傳 code 則列出全部產品目錄。
    """
    base = _PRODUCTS_IMG_DIR
    if code:
        folder = base / code.upper()
        if not folder.is_dir():
            return {"code": code.upper(), "images": []}
        imgs = sorted(
            p.name for p in folder.iterdir()
            if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".gif")
        )
        return {"code": code.upper(), "images": imgs}
    else:
        result = {}
        if base.is_dir():
            for d in sorted(base.iterdir()):
                if d.is_dir():
                    imgs = sorted(
                        p.name for p in d.iterdir()
                        if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".gif")
                    )
                    if imgs:
                        result[d.name] = imgs
        return {"products": result}


@app.get("/product-media/{file_path:path}")
async def serve_product_media(file_path: str):
    """
    轉發產品照片資料夾的檔案（圖片 / 影片），供 LINE push 使用的公開 URL。
    實際路徑：settings.PRODUCT_MEDIA_PATH / file_path
    例：GET /product-media/T1202A.jpg
    """
    import sys as _sys
    if _sys.platform == "win32":
        try:
            import ctypes as _ct
            _ct.windll.kernel32.SetErrorMode(0x0001 | 0x8000)
        except Exception:
            pass
    media_file = Path(settings.PRODUCT_MEDIA_PATH) / file_path
    media_file = media_file.resolve()
    media_root = Path(settings.PRODUCT_MEDIA_PATH).resolve()
    if not str(media_file).startswith(str(media_root)):
        raise HTTPException(status_code=403, detail="Forbidden")
    try:
        if not media_file.exists() or not media_file.is_file():
            raise HTTPException(status_code=404, detail=f"找不到檔案: {file_path}")
    except OSError:
        raise HTTPException(status_code=503, detail="產品照片磁碟機未連線")
    # 判斷 Content-Type
    ext = media_file.suffix.lower()
    media_types = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".gif": "image/gif",
        ".mp4": "video/mp4", ".mov": "video/quicktime",
    }
    mt = media_types.get(ext, "application/octet-stream")
    return FileResponse(str(media_file), media_type=mt)


@app.get("/admin/product-media-list")
async def list_product_media(code: str = ""):
    """
    列出產品照片資料夾中匹配指定代碼的所有檔案。
    不傳 code 則列出全部（前 200 個）。
    """
    import sys as _sys
    if _sys.platform == "win32":
        try:
            import ctypes as _ct
            _ct.windll.kernel32.SetErrorMode(0x0001 | 0x8000)
        except Exception:
            pass
    media_dir = Path(settings.PRODUCT_MEDIA_PATH)
    try:
        if not media_dir.is_dir():
            return {"ok": False, "detail": "產品照片磁碟機未連線或路徑不存在"}
    except OSError:
        return {"ok": False, "detail": "磁碟機未連線"}

    exts = {".jpg", ".jpeg", ".png", ".gif", ".mp4", ".mov"}
    if code:
        from handlers.internal import _match_product_media_files
        files = _match_product_media_files(code.upper(), media_dir)
        return {"code": code.upper(), "files": [f.name for f in files]}
    else:
        all_files = sorted(
            f.name for f in media_dir.iterdir()
            if f.is_file() and f.suffix.lower() in exts
        )[:200]
        return {"total": len(all_files), "files": all_files}


@app.post("/admin/update-customer-phone")
async def update_customer_phone(line_user_id: str, phone: str = ""):
    """
    手動更新客戶主電話（強制覆蓋），同時補到 customer_phones 多號表。
    phone 傳空字串表示清除主電話。

    用法：POST /admin/update-customer-phone?line_user_id=Uxxxxxxx&phone=0912345678
    """
    import sqlite3 as _sq
    from storage.customers import DB_PATH as _CUST_DB
    phone_val = phone.strip() or None
    with _sq.connect(str(_CUST_DB)) as _conn:
        cur = _conn.execute(
            "UPDATE customers SET phone=? WHERE line_user_id=?",
            (phone_val, line_user_id)
        )
        _conn.commit()
        if not cur.rowcount:
            raise HTTPException(status_code=404, detail=f"找不到 line_user_id={line_user_id}")
        if phone_val:
            row = _conn.execute(
                "SELECT id FROM customers WHERE line_user_id=?", (line_user_id,)
            ).fetchone()
            if row:
                _conn.execute(
                    "INSERT OR IGNORE INTO customer_phones (customer_id, phone) VALUES (?,?)",
                    (row[0], phone_val)
                )
                _conn.commit()
    print(f"[admin] 電話更新: {line_user_id} → phone={phone_val!r}")
    return {"status": "ok", "line_user_id": line_user_id, "phone": phone_val}


@app.post("/admin/update-customer-ecount-code")
async def update_customer_ecount_code(line_user_id: str, ecount_cust_cd: str = ""):
    """
    手動更新客戶 Ecount 客戶代碼（ecount_cust_cd）。
    傳空字串表示清除。

    用法：POST /admin/update-customer-ecount-code?line_user_id=Uxxxxxxx&ecount_cust_cd=M2509260001
    """
    import sqlite3 as _sq
    from storage.customers import DB_PATH as _CUST_DB
    val = ecount_cust_cd.strip() or None
    with _sq.connect(str(_CUST_DB)) as _conn:
        # 更新主表
        cur = _conn.execute(
            "UPDATE customers SET ecount_cust_cd=? WHERE line_user_id=?",
            (val, line_user_id)
        )
        if not cur.rowcount:
            raise HTTPException(status_code=404, detail=f"找不到 line_user_id={line_user_id}")
        # 同步更新子表（下單讀子表優先）
        if val:
            cust_id = _conn.execute(
                "SELECT id FROM customers WHERE line_user_id=?", (line_user_id,)
            ).fetchone()
            if cust_id:
                existing = _conn.execute(
                    "SELECT id FROM customer_ecount_codes WHERE customer_id=? AND ecount_cust_cd=?",
                    (cust_id[0], val)
                ).fetchone()
                if not existing:
                    _conn.execute(
                        "INSERT INTO customer_ecount_codes (customer_id, ecount_cust_cd, address_label, cust_name) VALUES (?, ?, '', '')",
                        (cust_id[0], val)
                    )
        _conn.commit()
    print(f"[admin] Ecount 代碼更新: {line_user_id} → ecount_cust_cd={val!r}")
    return {"status": "ok", "line_user_id": line_user_id, "ecount_cust_cd": val}


@app.get("/admin/customers")
async def admin_customers(q: str = ""):
    """
    查詢客戶資料（用於取得 line_user_id 再更新標籤）。
    q 可傳姓名或電話，不傳則回傳最近 50 筆。
    """
    if q:
        rows = customer_store.search(q)
    else:
        from storage.customers import DB_PATH as _CUST_DB
        import sqlite3 as _sq
        with _sq.connect(str(_CUST_DB)) as _conn:
            _conn.row_factory = _sq.Row
            rows = [dict(r) for r in _conn.execute(
                "SELECT id, line_user_id, display_name, chat_label, real_name, phone, ecount_cust_cd, tags "
                "FROM customers ORDER BY last_seen DESC LIMIT 50"
            ).fetchall()]
    # 補上子表的 ecount codes（主表沒有時用子表）
    from storage.customers import DB_PATH as _CUST_DB2
    import sqlite3 as _sq2
    with _sq2.connect(str(_CUST_DB2)) as _conn2:
        for r in rows:
            cid = r.get("id")
            if cid and not (r.get("ecount_cust_cd") or "").strip():
                sub_codes = _conn2.execute(
                    "SELECT ecount_cust_cd FROM customer_ecount_codes WHERE customer_id=?",
                    (cid,)
                ).fetchall()
                if sub_codes:
                    r["ecount_cust_cd"] = sub_codes[0][0]
    return {"customers": rows, "count": len(rows)}


@app.post("/admin/merge-customers")
async def merge_customers(keep_id: int, remove_id: int):
    """
    將 remove_id 的客戶資料合併到 keep_id，合併後刪除 remove_id。

    規則：keep 的欄位優先（有值就不蓋掉），remove 有而 keep 沒有的才補入。
    用法：POST /admin/merge-customers?keep_id=268&remove_id=22
    """
    import sqlite3 as _sq
    from storage.customers import DB_PATH as _CUST_DB

    with _sq.connect(str(_CUST_DB)) as conn:
        conn.row_factory = _sq.Row

        keep = conn.execute("SELECT * FROM customers WHERE id=?", (keep_id,)).fetchone()
        remove = conn.execute("SELECT * FROM customers WHERE id=?", (remove_id,)).fetchone()

        if not keep:
            raise HTTPException(status_code=404, detail=f"找不到 keep_id={keep_id}")
        if not remove:
            raise HTTPException(status_code=404, detail=f"找不到 remove_id={remove_id}")

        keep = dict(keep)
        remove = dict(remove)

        # 各欄位：keep 有值就保留，沒有才從 remove 補
        fields = ["phone", "real_name", "address", "note", "ecount_cust_cd",
                  "chat_label", "preferred_ecount_cust_cd"]
        updates = {}
        for f in fields:
            if not keep.get(f) and remove.get(f):
                updates[f] = remove[f]

        if updates:
            set_clause = ", ".join(f"{k}=?" for k in updates)
            conn.execute(
                f"UPDATE customers SET {set_clause} WHERE id=?",
                list(updates.values()) + [keep_id]
            )

        # 搬移 customer_phones
        phones = conn.execute(
            "SELECT phone FROM customer_phones WHERE customer_id=?", (remove_id,)
        ).fetchall()
        for row in phones:
            conn.execute(
                "INSERT OR IGNORE INTO customer_phones (customer_id, phone) VALUES (?,?)",
                (keep_id, row["phone"])
            )

        # 刪除 remove 記錄
        conn.execute("DELETE FROM customers WHERE id=?", (remove_id,))
        conn.commit()

    merged_fields = list(updates.keys()) if updates else []
    print(f"[admin] 合併客戶: keep={keep_id} remove={remove_id} 補入欄位={merged_fields}")
    return {
        "status": "ok",
        "keep_id": keep_id,
        "remove_id": remove_id,
        "merged_fields": merged_fields,
    }


@app.post("/admin/sync-customer-names")
async def sync_customer_names():
    """
    完整版 Ecount 姓名同步（推薦使用這個）：
    用 ecount_cust_cd + 電話號碼 雙重比對，把 Ecount 姓名存入 real_name。
    讓內部群組可以用 Ecount 姓名搜尋到 LINE 客戶（通知登記、代訂單等）。

    需要先有 data/ecount_customers.json（由 scripts/sync_cust_from_web.py 爬取）。
    """
    import json
    from pathlib import Path
    json_path = Path(__file__).parent / "data" / "ecount_customers.json"
    if not json_path.exists():
        return {"status": "error", "message": f"找不到 {json_path.name}，請先執行 scripts/sync_cust_from_web.py"}
    ecount_list = json.loads(json_path.read_text(encoding="utf-8"))
    if not ecount_list:
        return {"status": "error", "message": "ecount_customers.json 內容為空"}
    result = await asyncio.to_thread(customer_store.sync_ecount_names_full, ecount_list)
    total = result["by_code"] + result["by_phone"]
    print(f"[admin] Ecount 姓名同步：{len(ecount_list)} 筆 → 更新 {total} 筆 "
          f"(ecount_cust_cd:{result['by_code']}, 電話:{result['by_phone']}, 跳過:{result['skipped']})")
    return {
        "status": "ok",
        "ecount_total": len(ecount_list),
        "matched_by_code": result["by_code"],
        "matched_by_phone": result["by_phone"],
        "skipped": result["skipped"],
        "total_updated": total,
    }


@app.post("/admin/set-group-address")
async def set_group_address(group_id: str, ecount_cust_cd: str, label: str = ""):
    """
    登記客戶群組預設 Ecount 地址。
    當 bot 收到來自該群組的訊息時，會自動帶入此地址作為訂單預設地址。

    用法：POST /admin/set-group-address?group_id=C1234xxxxx&ecount_cust_cd=M2509260001-2&label=樹林
    """
    customer_store.set_group_address(group_id, ecount_cust_cd, label)
    print(f"[admin] 群組地址設定: {group_id} → {ecount_cust_cd} ({label})")
    return {
        "status": "ok",
        "group_id": group_id,
        "ecount_cust_cd": ecount_cust_cd,
        "label": label,
    }


@app.get("/admin/group-addresses")
async def list_group_addresses():
    """列出所有已登記的客戶群組預設地址"""
    rows = customer_store.list_group_addresses()
    return {"status": "ok", "count": len(rows), "data": rows}


@app.post("/admin/set-customer-preferred-address")
async def set_customer_preferred_address(
    customer_id: int,
    ecount_cust_cd: str,
    label: str = "",
):
    """
    設定某位客戶的個人預設訂單地址。
    叫貨結帳時直接問「是否送到 {label}？」而非列出全部地址選單。
    個人設定優先於群組設定。傳 ecount_cust_cd=（空字串）可清除。

    例：Du 管饒河店
        POST /admin/set-customer-preferred-address
             ?customer_id=XXX&ecount_cust_cd=M2510010009-1&label=饒河

    例：Rachel 管文山店
        POST /admin/set-customer-preferred-address
             ?customer_id=XXX&ecount_cust_cd=M2510010009-4&label=文山
    """
    customer_store.set_preferred_address(customer_id, ecount_cust_cd or None)
    action = "清除" if not ecount_cust_cd else f"設為 {ecount_cust_cd} ({label})"
    print(f"[admin] 客戶 {customer_id} 個人預設地址{action}")
    return {
        "status": "ok",
        "customer_id": customer_id,
        "ecount_cust_cd": ecount_cust_cd or None,
        "label": label,
    }


# ── 根路徑導向 admin ────────────────────────────────────
from fastapi.responses import RedirectResponse

@app.get("/")
async def root():
    return RedirectResponse(url="/admin")

# ── 管理介面 HTML ────────────────────────────────────
@app.get("/admin/analytics")
async def admin_analytics_page():
    """分析儀表板 HTML 頁面"""
    return FileResponse(str(_STATIC_DIR / "analytics.html"))


@app.get("/admin/analytics/data")
async def admin_analytics_data(report: str = "all"):
    """分析 API — 回傳 JSON"""
    from services.analytics import (
        top_sellers, slow_movers, customer_analysis,
        restock_forecast, price_band_analysis, category_analysis,
        monthly_trend, product_trend, stock_turnover, customer_churn, do_not_restock,
    )
    result = {}
    if report in ("all", "top_sellers"):
        result["top_sellers"] = top_sellers(30, 20)
    if report in ("all", "slow_movers"):
        result["slow_movers"] = slow_movers(60, 10)
    if report in ("all", "customers"):
        result["customers"] = customer_analysis(90, 20)
    if report in ("all", "forecast"):
        result["forecast"] = restock_forecast(30)
    if report in ("all", "price_bands"):
        result["price_bands"] = price_band_analysis(90)
    if report in ("all", "categories"):
        result["categories"] = category_analysis(90)
    if report in ("all", "monthly_trend"):
        result["monthly_trend"] = monthly_trend()
    if report in ("all", "product_trend"):
        result["product_trend"] = product_trend(90)
    if report in ("all", "stock_turnover"):
        result["stock_turnover"] = stock_turnover(90)
    if report in ("all", "customer_churn"):
        result["customer_churn"] = customer_churn(60)
    if report in ("all", "do_not_restock"):
        result["do_not_restock"] = do_not_restock()
    return result


@app.get("/admin")
async def admin_ui():
    """管理介面首頁"""
    return FileResponse(_ADMIN_HTML)


@app.get("/admin/ngrok-url")
async def admin_ngrok_url():
    """從 ngrok 本地 API 取得目前的公開網址"""
    import httpx
    try:
        async with httpx.AsyncClient(timeout=2.0) as client:
            r = await client.get("http://localhost:4040/api/tunnels")
            tunnels = r.json().get("tunnels", [])
            for t in tunnels:
                if t.get("proto") == "https":
                    return {"url": t["public_url"]}
            # 沒有 https 就回第一個
            if tunnels:
                return {"url": tunnels[0]["public_url"]}
    except Exception:
        pass
    return {"url": None}


@app.get("/admin/visits")
async def admin_visits():
    """取得預計到店客人清單"""
    return {"visits": visit_store.get_pending()}


@app.post("/admin/visits/{visit_id}/resolve")
async def admin_visit_resolve(visit_id: int):
    """標記客人已到店"""
    ok = visit_store.mark_visited(visit_id)
    return {"ok": ok}


@app.get("/admin/pickup-notify")
async def admin_pickup_notify():
    """取得最近一次到貨通知結果（14:00 排程）"""
    return _pickup_notify_results or {"timestamp": None, "notified_history": [], "no_line_id": []}


@app.post("/admin/pickup-notify/run")
async def admin_pickup_notify_run():
    """手動觸發到貨通知（在 server 內部執行）"""
    import asyncio
    asyncio.get_event_loop().run_in_executor(None, _check_and_notify_pickup)
    return {"ok": True, "message": "到貨通知已觸發，結果稍後可查"}


@app.post("/api/pickup-notify/run")
async def api_pickup_notify_run():
    """手動觸發到貨通知（不需登入）"""
    import asyncio
    asyncio.get_event_loop().run_in_executor(None, _check_and_notify_pickup)
    return {"ok": True, "message": "到貨通知已觸發"}


@app.post("/admin/pickup-notify/{name}/done")
async def admin_pickup_notify_done(name: str):
    """標記需手動通知的客戶為已通知"""
    global _pickup_notify_results
    if not _pickup_notify_results:
        return {"ok": False}
    _pickup_notify_results["no_line_id"] = [
        c for c in _pickup_notify_results.get("no_line_id", [])
        if c["name"] != name
    ]
    return {"ok": True}


@app.get("/admin/carts")
async def admin_carts():
    """取得所有未結帳購物車"""
    from storage import cart as _cart_admin
    result = []
    with _cart_admin._lock:
        for uid, items in _cart_admin._carts.items():
            if not items:
                continue
            cust = customer_store.get_by_line_id(uid)
            name = (cust.get("real_name") or cust.get("display_name") or uid[:15]) if cust else uid[:15]
            result.append({
                "user_id": uid,
                "customer_name": name,
                "items": items,
            })
    return result


@app.delete("/admin/carts/{user_id}")
async def admin_cart_delete(user_id: str):
    """刪除某客戶的購物車"""
    from storage import cart as _cart_del
    _cart_del.clear_cart(user_id)
    return {"ok": True}


@app.get("/admin/seen-groups")
async def admin_seen_groups():
    """列出這次伺服器啟動後收到訊息的所有群組 ID"""
    return {
        "groups": list(_seen_group_ids),
        "configured": {
            "LINE_GROUP_ID": settings.LINE_GROUP_ID or "(未設定)",
            "LINE_GROUP_ID_HQ": settings.LINE_GROUP_ID_HQ or "(未設定)",
            "LINE_GROUP_ID_SHOWCASE": settings.LINE_GROUP_ID_SHOWCASE or "(未設定)",
        }
    }


# ── 群組管理 ─────────────────────────────────────────
@app.get("/admin/groups")
async def admin_groups():
    """列出所有群組（已知+未知）"""
    known = {
        settings.LINE_GROUP_ID: "內部群",
        settings.LINE_GROUP_ID_HQ: "總公司群",
        settings.LINE_GROUP_ID_SHOWCASE: "看貨群",
    }
    registered = customer_store.list_group_addresses()
    reg_map = {r["group_id"]: r for r in registered}

    groups = []
    # Known groups
    for gid, name in known.items():
        if gid:
            groups.append({
                "group_id": gid,
                "type": "system",
                "label": name,
                "ecount_cust_cd": reg_map.get(gid, {}).get("ecount_cust_cd", ""),
            })
    # Registered customer groups
    for r in registered:
        if r["group_id"] not in known:
            groups.append({
                "group_id": r["group_id"],
                "type": "customer",
                "label": r.get("label", ""),
                "ecount_cust_cd": r.get("ecount_cust_cd", ""),
            })
    # Unknown groups (seen but not registered)
    for gid, last_seen in _unknown_groups.items():
        if gid not in known and gid not in reg_map:
            groups.append({
                "group_id": gid,
                "type": "unknown",
                "label": "",
                "ecount_cust_cd": "",
                "last_seen": last_seen,
            })
    return groups


@app.post("/admin/set-group")
async def admin_set_group(group_id: str, ecount_cust_cd: str = "", label: str = ""):
    """設定群組的客戶代碼和標籤"""
    customer_store.set_group_address(group_id, ecount_cust_cd, label)
    # Remove from unknown if it was there
    _unknown_groups.pop(group_id, None)
    return {"ok": True}


# ── Server 重啟 ──────────────────────────────────────
@app.post("/admin/restart")
async def admin_restart():
    """重啟 server（背景執行，不顯示視窗）
    - tray.py 環境：殺掉整個 process tree，watchdog 會自動重啟
    - 獨立執行：用 VBS 先啟動新 server 再退出
    """
    import threading, os, subprocess, signal

    _vbs = _BASE_DIR / "start_server_bg.vbs"
    _lock = _BASE_DIR / "data" / "tray.lock"
    _under_tray = _lock.exists()

    def _do_restart():
        import time
        time.sleep(0.8)  # 等本次 HTTP 回應送出

        if not _under_tray:
            # 獨立模式：先啟動新 server 再退出
            subprocess.Popen(
                ["wscript.exe", str(_vbs)],
                creationflags=0x08000000,
            )
            time.sleep(1.5)

        # 用 taskkill /T 殺掉整個 process tree（包括 reloader + worker + shell）
        # 這樣 port 才會被正確釋放
        my_pid = os.getpid()
        try:
            # 先找到 reloader 的 PID（父進程）
            import ctypes
            kernel32 = ctypes.windll.kernel32
            # 嘗試殺掉佔用 8000 port 的所有 process
            result = subprocess.run(
                ["netstat", "-ano"],
                capture_output=True, text=True,
                creationflags=0x08000000,
            )
            pids_to_kill = set()
            for line in result.stdout.splitlines():
                if ":8000 " in line and "LISTENING" in line:
                    pid = line.strip().split()[-1]
                    if pid.isdigit():
                        pids_to_kill.add(int(pid))

            for pid in pids_to_kill:
                try:
                    subprocess.run(
                        ["taskkill", "/F", "/T", "/PID", str(pid)],
                        creationflags=0x08000000,
                        capture_output=True,
                        timeout=5,
                    )
                except Exception:
                    pass
        except Exception:
            pass

        # 確保自己也退出
        os._exit(0)

    threading.Thread(target=_do_restart, daemon=True).start()
    mode = "tray watchdog" if _under_tray else "VBS"
    print(f"[admin] 重啟指令收到，模式：{mode}，即將重啟...", flush=True)
    return {"ok": True}


# ── Bot 開關 ─────────────────────────────────────────
@app.post("/admin/bot/on")
async def bot_on():
    """啟動機器人（開始處理 LINE 訊息）"""
    global _bot_active
    _bot_active = True
    print("[admin] 機器人已啟動")
    return {"bot_active": True}


@app.post("/admin/bot/off")
async def bot_off():
    """關閉機器人（暫停處理 LINE 訊息）"""
    global _bot_active
    _bot_active = False
    print("[admin] 機器人已關閉")
    return {"bot_active": False}


# ── API 連線狀態 ─────────────────────────────────────
@app.get("/admin/api-status")
async def admin_api_status():
    """
    回傳各服務連線狀態：
      bot_active   — 機器人是否啟用
      line_api     — LINE API 是否正常
      ecount_api   — Ecount ERP 是否可登入
      db_customers — 客戶資料庫是否可讀
      db_orders    — 訂單/調貨/配送資料庫是否可讀
    """
    # LINE API
    try:
        line_ok = await asyncio.to_thread(lambda: bool(_line_api.get_bot_info()))
    except Exception:
        line_ok = False

    # Ecount API
    try:
        from services.ecount import ecount_client as _ec
        sid = await asyncio.to_thread(_ec._ensure_session)
        ecount_ok = bool(sid)
    except Exception:
        ecount_ok = False

    # SQLite DB 健康檢查（使用絕對路徑避免工作目錄問題）
    def _db_ok(rel: str) -> bool:
        p = _BASE_DIR / rel
        if not p.exists():
            return False
        try:
            with sqlite3.connect(str(p)) as conn:
                conn.execute("SELECT 1")
            return True
        except Exception:
            return False

    db_cust_ok = _db_ok("data/customers.db")
    db_orders_ok = all([
        _db_ok("data/payment_confirmations.db"),
        _db_ok("data/restock_requests.db"),
        _db_ok("data/delivery_inquiries.db"),
    ])

    # Server 狀態
    import os as _os
    uptime_sec = int(_time_module.time() - _server_start_time)
    server_info = {"ok": True, "pid": _os.getpid(), "uptime": uptime_sec}

    # ngrok 狀態（查詢 localhost:4040 API）
    ngrok_ok  = False
    ngrok_url = None
    try:
        import httpx as _httpx
        _r = _httpx.get("http://localhost:4040/api/tunnels", timeout=2)
        if _r.status_code == 200:
            tunnels = _r.json().get("tunnels", [])
            for t in tunnels:
                if t.get("proto") == "https":
                    ngrok_url = t.get("public_url")
                    ngrok_ok  = True
                    break
            if not ngrok_ok and tunnels:
                ngrok_ok  = True
                ngrok_url = tunnels[0].get("public_url")
    except Exception:
        pass

    return {
        "bot_active":    _bot_active,
        "line_api":      line_ok,
        "ecount_api":    ecount_ok,
        "db_customers":  db_cust_ok,
        "db_orders":     db_orders_ok,
        "server":        server_info,
        "ngrok":         {"ok": ngrok_ok, "url": ngrok_url},
        "push_quota":    {"exhausted": _push_quota_exhausted, "since": _push_quota_exhausted_at or None},
    }


# ── 庫存情況表 ────────────────────────────────────────
_inventory_check_cache: dict = {"updated_at": None, "items": []}


def _build_inventory_check() -> list[dict]:
    """
    從 Ecount 取所有有庫存（qty > 0）的品項，
    排除 HH 開頭貨號，逐一比對 specs.json（PO文）與產品照片資料夾。
    只回傳「缺規格 or 缺照片」的品項，按庫存量由大到小排序。
    """
    from services.ecount import ecount_client as _ec
    from handlers.internal import _get_media_dir, _match_product_media_files

    # 1. 讀 specs.json
    specs_path = _BASE_DIR / "data" / "specs.json"
    try:
        import json as _j
        specs: dict = _j.loads(specs_path.read_text(encoding="utf-8")) if specs_path.exists() else {}
    except Exception:
        specs = {}

    # 2. 取有庫存品項（Ecount），排除 HH 開頭貨號
    stock_items = _ec.get_all_stock_products()
    if not stock_items:
        return []

    # 3. 取產品照片目錄
    media_dir = _get_media_dir()

    # 3b. 載入原始 PO文.txt 的所有貨號（補充 specs.json 沒有的）
    from handlers.internal import _PO_TXT_PATH
    _po_codes = set()
    try:
        _po_path = Path(_PO_TXT_PATH)
        if _po_path.exists():
            _po_content = None
            for _enc in ("cp950", "big5", "utf-8"):
                try:
                    _po_content = _po_path.read_text(encoding=_enc)
                    break
                except Exception:
                    continue
            if _po_content:
                import re as _re_po
                _po_codes = set(c.upper() for c in _re_po.findall(r'[A-Za-z]{1,3}-?\d{3,6}', _po_content))
    except Exception:
        pass

    # 4. 只保留「缺規格 or 缺照片」的品項
    rows = []
    for item in stock_items:
        code = item["code"].upper()
        if code.startswith("HH"):          # 排除 HH 開頭
            continue
        has_spec  = code in specs or code in _po_codes
        has_photo = bool(media_dir and _match_product_media_files(code, media_dir))
        if has_spec and has_photo:         # 都有 → 跳過
            continue
        rows.append({
            "code":      code,
            "name":      item["name"] or specs.get(code, {}).get("name", ""),
            "qty":       item["qty"],
            "has_spec":  has_spec,
            "has_photo": has_photo,
        })

    # 庫存量由大到小排序
    rows.sort(key=lambda r: -r["qty"])
    return rows


async def _midnight_inventory_check_loop():
    """每天 00:00 自動更新庫存情況表"""
    while True:
        now = datetime.now()
        # 計算距明天 00:00 的秒數
        nxt = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        wait_sec = (nxt - now).total_seconds()
        await asyncio.sleep(wait_sec)
        try:
            items = await asyncio.to_thread(_build_inventory_check)
            _inventory_check_cache["items"]      = items
            _inventory_check_cache["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            print(f"[inventory-check] 每日更新完成，共 {len(items)} 筆有庫存品項")
        except Exception as e:
            print(f"[inventory-check] 每日更新失敗: {e}")


@app.get("/admin/export-specs")
async def admin_export_specs():
    """導出產品規格為 Excel 檔案"""
    import json as _json
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io as _io

    specs_path = _BASE_DIR / "data" / "specs.json"
    if not specs_path.exists():
        raise HTTPException(status_code=404, detail="specs.json 不存在")

    specs = _json.loads(specs_path.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "產品規格"

    # 標題列
    headers = ["產品編號", "品名", "尺寸", "重量", "適用台型", "售價"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 資料列
    row_font = Font(name="微軟正黑體", size=10)
    for row_idx, (code, s) in enumerate(sorted(specs.items()), 2):
        values = [
            s.get("code", ""),
            s.get("name", ""),
            s.get("size", ""),
            s.get("weight", ""),
            "、".join(s.get("machine", [])),
            s.get("price", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = row_font
            cell.border = thin_border

    # 欄寬
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_specs.xlsx"},
    )


@app.get("/admin/export-catalog")
async def admin_export_catalog():
    """導出商品目錄 Excel（含圖片，僅列 available > 10）"""
    import json as _json
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    from handlers.internal import _get_media_dir, _match_product_media_files, _IMG_EXTS

    specs_path = _BASE_DIR / "data" / "specs.json"
    avail_path = _BASE_DIR / "data" / "available.json"
    if not specs_path.exists():
        raise HTTPException(status_code=404, detail="specs.json 不存在")
    if not avail_path.exists():
        raise HTTPException(status_code=404, detail="available.json 不存在")

    specs = _json.loads(specs_path.read_text(encoding="utf-8"))
    avail = _json.loads(avail_path.read_text(encoding="utf-8"))
    media_dir = _get_media_dir()

    rows = []
    for code, s in specs.items():
        a = avail.get(code.upper()) or avail.get(code) or {}
        stock = int(a.get("available", 0) or 0)
        if stock <= 10:
            continue
        rows.append((code, s, stock))
    rows.sort(key=lambda r: r[0])

    wb = Workbook()
    ws = wb.active
    ws.title = "商品目錄"

    headers = ["圖片", "產品編號", "品名", "尺寸", "重量", "售價", "庫存"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    row_font = Font(name="微軟正黑體", size=10)
    THUMB = 80
    for row_idx, (code, s, stock) in enumerate(rows, 2):
        values = [
            "",
            s.get("code", code),
            s.get("name", ""),
            s.get("size", ""),
            s.get("weight", ""),
            s.get("price", ""),
            stock,
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = row_font
            c.border = thin_border
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 65

        if media_dir:
            try:
                files = _match_product_media_files(code, media_dir)
                img_files = sorted(
                    [f for f in files if f.suffix.lower() in _IMG_EXTS]
                )
                if img_files:
                    bio = _io.BytesIO()
                    with PILImage.open(img_files[0]) as im:
                        im = im.convert("RGB")
                        im.thumbnail((THUMB, THUMB))
                        w, h = im.size
                        im.save(bio, format="PNG")
                    bio.seek(0)
                    xl_img = XLImage(bio)
                    xl_img.width = w
                    xl_img.height = h
                    ws.add_image(xl_img, f"A{row_idx}")
            except Exception as e:
                print(f"[export-catalog] {code} 圖片嵌入失敗: {e}")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_catalog.xlsx"},
    )


@app.get("/admin/inventory-check")
async def admin_inventory_check(refresh: bool = False):
    """
    回傳有庫存的品項清單，並標記是否有規格表、是否有產品照片。
    ?refresh=true 強制重新查詢 Ecount。
    """
    if refresh or not _inventory_check_cache["updated_at"]:
        try:
            items = await asyncio.to_thread(_build_inventory_check)
            _inventory_check_cache["items"]      = items
            _inventory_check_cache["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        except Exception as e:
            return {"error": str(e), "items": [], "updated_at": None}
    return {
        "updated_at": _inventory_check_cache["updated_at"],
        "items":      _inventory_check_cache["items"],
    }


@app.get("/admin/new-products")
async def admin_new_products():
    """回傳待審核品項清單"""
    from storage.new_products import new_products_store
    return {"items": new_products_store.get_pending()}

@app.post("/admin/new-products/{item_id}/confirm")
async def admin_confirm_new_product(item_id: int):
    """確認品項（從待審核清單移除）"""
    from storage.new_products import new_products_store
    ok = new_products_store.confirm(item_id)
    return {"ok": ok}

@app.post("/admin/new-products/{item_id}/delete")
async def admin_delete_new_product(item_id: int):
    """棄用品項（直接刪除）"""
    from storage.new_products import new_products_store
    ok = new_products_store.delete(item_id)
    return {"ok": ok}


# ── 接手面板 API ──────────────────────────────────────────────────────────
_TAKEOVER_EXPIRE_MINUTES = 0  # 0 = 不自動過期，只能手動釋放

@app.post("/admin/takeover")
async def admin_takeover(user_id: str, display_name: str = ""):
    """員工接手客戶對話，bot 靜默 60 分鐘"""
    import time as _time
    state_manager.set(user_id, {
        "action":       "human_takeover",
        "taken_at":     _time.time(),
        "display_name": display_name,
    })
    from services.claude_ai import add_chat_history
    add_chat_history(user_id, "bot", "（以下由真人客服接手處理，之前的問題不需要再回覆）")
    print(f"[takeover] 接手: {display_name}（{user_id[:10]}...）")
    return {"status": "ok", "user_id": user_id}

@app.post("/admin/release")
async def admin_release(user_id: str):
    """釋放客戶，bot 恢復回覆。自動從 LINE OA 讀取接管期間的對話。"""
    st = state_manager.get(user_id) or {}
    if st.get("action") == "human_takeover":
        state_manager.clear(user_id)
        from services.claude_ai import add_chat_history

        # 嘗試從 LINE OA Manager 讀取接管期間的對話
        cust = customer_store.get_by_line_id(user_id)
        cust_name = (cust.get("chat_label") or cust.get("real_name") or cust.get("display_name") or "") if cust else ""
        if cust_name:
            try:
                import threading as _t_rel
                _OA_SKIP_TEXT = {"已讀", "讀", ""}
                def _sync_oa_chat():
                    try:
                        print(f"[takeover] 開始從 LINE OA 讀取對話：{cust_name}", flush=True)
                        from services.line_oa_chat import read_chat_sync
                        msgs = read_chat_sync(cust_name, max_messages=20)
                        if msgs:
                            staff_msgs = [m for m in msgs if m["role"] == "staff"
                                          and m["text"].replace("已讀", "").strip()
                                          and m["text"].strip() not in _OA_SKIP_TEXT]
                            for m in staff_msgs[-5:]:
                                _clean_text = m['text'].replace("已讀", "").strip()
                                if _clean_text:
                                    add_chat_history(user_id, "bot", f"（真人回覆）{_clean_text}")
                            print(f"[takeover] LINE OA 對話已同步：{len(staff_msgs)} 則真人回覆", flush=True)
                        else:
                            print(f"[takeover] LINE OA 沒有讀到對話：{cust_name}", flush=True)
                    except Exception as e:
                        import traceback
                        print(f"[takeover] LINE OA 對話同步失敗: {e}", flush=True)
                        traceback.print_exc()
                _t_rel.Thread(target=_sync_oa_chat, daemon=True).start()
            except Exception:
                pass

        add_chat_history(user_id, "bot", "（以上問題已由真人客服處理完成，不需要再回覆之前的問題）")
        print(f"[takeover] 釋放: {user_id[:10]}...")
    return {"status": "ok", "user_id": user_id}

@app.post("/admin/resolve-item")
async def admin_resolve_item(item_type: str, item_id: int):
    """標記單一待處理項目為已處理。item_type: P/R/D/I/Q"""
    result = _resolve_one(item_type, item_id)
    ok = result.startswith("✅")
    return {"status": "ok" if ok else "not_found", "message": result}


@app.post("/admin/resolve-all-pending")
async def admin_resolve_all_pending():
    """標記所有未處理項目為已處理"""
    count = 0
    for p in payment_store.get_pending():
        if payment_store.resolve(p["id"]): count += 1
    for r in restock_store.get_unresolved():
        restock_store.update_status(r["id"], "confirmed"); count += 1
    for d in delivery_store.get_pending():
        if delivery_store.resolve(d["id"]): count += 1
    for i in issue_store.get_pending():
        if issue_store.resolve(i["id"]): count += 1
    for q in pending_store.get_pending():
        pending_store.mark_answered(q["id"]); count += 1
    return {"status": "ok", "resolved_count": count}


@app.get("/admin/takeovers")
async def admin_list_takeovers():
    """列出目前所有接手中的客戶"""
    import time as _time
    now = _time.time()
    results = []
    for uid, st in state_manager.all_states().items():
        if st.get("action") == "human_takeover":
            taken_at  = st.get("taken_at", now)
            elapsed   = int((now - taken_at) / 60)
            if _TAKEOVER_EXPIRE_MINUTES > 0:
                remaining = max(0, _TAKEOVER_EXPIRE_MINUTES - elapsed)
                if remaining == 0:
                    state_manager.clear(uid)
                    continue
            else:
                remaining = -1  # 不過期
            cust = customer_store.get_by_line_id(uid)
            results.append({
                "user_id":       uid,
                "display_name":  st.get("display_name") or (cust.get("real_name") or cust.get("display_name") if cust else uid[:10]),
                "elapsed_min":   elapsed,
                "remaining_min": remaining,
            })
    return results

@app.get("/admin/recent-customers")
async def admin_recent_customers():
    """取得最近 18 位互動的客戶（供接手面板使用）"""
    import sqlite3 as _sq, os as _os
    db_path = _os.path.join(_os.path.dirname(__file__), "data", "customers.db")
    try:
        with _sq.connect(db_path) as conn:
            rows = conn.execute("""
                SELECT line_user_id, real_name, display_name, chat_label
                FROM customers
                WHERE line_user_id IS NOT NULL
                ORDER BY last_seen DESC
                LIMIT 18
            """).fetchall()
        takeover_uids = {uid for uid, st in state_manager.all_states().items()
                         if st.get("action") == "human_takeover"}
        return [
            {
                "user_id":      r[0],
                "name":         r[1] or r[3] or r[2] or r[0][:10],
                "is_takeover":  r[0] in takeover_uids,
            }
            for r in rows
        ]
    except Exception as e:
        return []


# ── 到貨通知管理 ──────────────────────────────────────────
@app.get("/admin/notify")
async def admin_notify_list(q: str = "", status: str = ""):
    """
    取得到貨通知記錄。
    q: 搜尋關鍵字（客戶名、產品編號、品名）
    status: 篩選狀態（pending/notified/cancelled），空=全部
    待通知（pending）最多顯示最新 10 筆。
    """
    from services.ecount import ecount_client
    records = notify_store.get_all()

    def _enrich(r):
        cust = customer_store.get_by_line_id(r["user_id"])
        r["customer_name"] = (cust.get("real_name") or cust.get("display_name") or r["user_id"][:10]) if cust else r["user_id"][:10]
        item = ecount_client.get_product_cache_item(r["prod_code"])
        box_qty = (item.get("box_qty") or 0) if item else 0
        prod_unit = (item.get("unit") or "") if item else ""
        qty = r["qty_wanted"]
        if prod_unit == "箱":
            r["qty_display"] = f"{qty}箱"
            r["unit"] = "箱"
            r["box_qty"] = box_qty or 1
        elif box_qty > 1 and qty >= box_qty and qty % box_qty == 0:
            r["qty_display"] = f"{qty // box_qty}箱"
            r["unit"] = "箱"
            r["box_qty"] = box_qty
        else:
            _u = prod_unit or "個"
            r["qty_display"] = f"{qty}{_u}"
            r["unit"] = _u
            r["box_qty"] = box_qty
        return r

    enriched = []
    for r in records:
        try:
            enriched.append(_enrich(r))
        except Exception as _e:
            print(f"[admin/notify] enrich 失敗 id={r.get('id')}: {_e}", flush=True)
            r["customer_name"] = r.get("user_id", "?")[:20]
            r["qty_display"] = f"{r.get('qty_wanted', 0)}個"
            r["unit"] = "個"
            r["box_qty"] = 0
            enriched.append(r)
    records = enriched

    # 狀態篩選
    if status:
        records = [r for r in records if r["status"] == status]

    # 關鍵字搜尋（客戶名、產品編號、品名）
    if q:
        kw = q.strip().upper()
        records = [r for r in records if (
            kw in r.get("customer_name", "").upper()
            or kw in r.get("prod_code", "").upper()
            or kw in r.get("prod_name", "").upper()
        )]

    # 待通知顯示全部 pending（不截斷）
    if not q and not status:
        pending = [r for r in records if r["status"] == "pending"]
        others = [r for r in records if r["status"] != "pending"]
        records = pending + others

    return records

@app.put("/admin/notify/{notify_id}")
async def admin_notify_update(notify_id: int, request: Request):
    """更新到貨通知"""
    body = await request.json()
    notify_store.update(notify_id, **body)
    return {"ok": True}

@app.delete("/admin/notify/{notify_id}")
async def admin_notify_delete(notify_id: int):
    """刪除到貨通知"""
    notify_store.delete(notify_id)
    return {"ok": True}

# ── 回饋金 ──────────────────────────────────────────────
@app.get("/admin/rebate")
async def admin_rebate(sync: bool = False):
    """取得當月回饋金計算結果，sync=true 時先從 Ecount 同步"""
    if sync:
        try:
            import subprocess as _sp
            _python = _sys.executable
            _root = str(Path(__file__).parent)
            _flags = _sp.CREATE_NO_WINDOW if _sys.platform == "win32" else 0
            proc = await asyncio.to_thread(
                _sp.run, [_python, "-m", "scripts.sync_rebate"],
                cwd=_root, capture_output=True, timeout=180, creationflags=_flags,
            )
            if proc.stdout:
                print(proc.stdout.decode("utf-8", errors="replace"), flush=True)
        except Exception as e:
            print(f"[rebate] 同步失敗: {e}")
    from services.rebate import calculate_rebates
    return await asyncio.to_thread(calculate_rebates)

@app.get("/admin/rebate/approaching")
async def admin_rebate_approaching():
    """1~15日：上月達標客戶；16日起：當月快接近達成"""
    from datetime import datetime as _dt
    day = _dt.now().day
    if day <= 15:
        from services.rebate import get_last_month_achievers
        return await asyncio.to_thread(get_last_month_achievers)
    else:
        from services.rebate import get_approaching_customers
        return await asyncio.to_thread(get_approaching_customers)


@app.post("/webhook")
async def webhook(request: Request):
    signature = request.headers.get("X-Line-Signature", "")
    body = await request.body()

    # ── chat_mode_changed 事件（員工點「使用手動聊天」）────────────────────
    try:
        import json as _json
        _payload = _json.loads(body)
        for _ev in _payload.get("events", []):
            _ev_type = _ev.get("type", "")
            # 入口層級 log：每個 webhook event 都印一行（追未處理訊息用）
            _src_in = _ev.get("source") or {}
            _uid_in = _src_in.get("userId", "") or _src_in.get("groupId", "") or _src_in.get("roomId", "")
            _msg_in = _ev.get("message", {})
            _mtype_in = _msg_in.get("type", "")
            _snippet_in = (_msg_in.get("text", "") or _msg_in.get("id", ""))[:60]
            print(f"[webhook-inbound] uid={_uid_in[:10]}... ev={_ev_type} msg={_mtype_in} text={_snippet_in!r}", flush=True)
            # 詳細記錄非常規事件
            if _ev_type != "message":
                print(f"[webhook] 非message事件: type={_ev_type} ev={_json.dumps(_ev, ensure_ascii=False)[:300]}", flush=True)
            else:
                _msg = _ev.get("message", {})
                _msg_type = _msg.get("type", "")
                if _msg_type not in ("text", "image", "video"):
                    print(f"[webhook] 未處理的message類型: {_msg_type} ev={_json.dumps(_ev, ensure_ascii=False)[:300]}", flush=True)
            if _ev_type == "chat_mode_changed":
                _src  = _ev.get("source") or {}
                _uid  = _src.get("userId", "") or _src.get("groupId", "") or _src.get("roomId", "")
                _mode = _ev.get("mode", "")
                if _uid:
                    if _mode == "standby":
                        state_manager.set(_uid, {"action": "human_takeover"})
                        from services.claude_ai import add_chat_history as _ach
                        _ach(_uid, "bot", "（以下由真人客服接手處理，之前的問題不需要再回覆）")
                        print(f"[chat_mode] 員工接手 {_uid[:10]}... → bot 靜默")
                    elif _mode == "active":
                        if (state_manager.get(_uid) or {}).get("action") == "human_takeover":
                            state_manager.clear(_uid)
                        print(f"[chat_mode] bot 恢復 {_uid[:10]}...")
    except Exception as _e:
        print(f"[webhook] chat_mode parse 錯誤: {_e}")

    try:
        await asyncio.to_thread(_webhook_handler.handle, body.decode("utf-8"), signature)
    except InvalidSignatureError:
        raise HTTPException(status_code=400, detail="Invalid signature")
    return "OK"


@_webhook_handler.add(MessageEvent, message=TextMessageContent)
def on_message(event: MessageEvent):
    # ── 訊息去重 ─────────────────────────────────────────
    msg_id = event.message.id
    now = _time_module.time()
    if msg_id in _processed_msg_ids:
        return  # Already processed
    _processed_msg_ids[msg_id] = now
    # Clean entries older than 60s periodically
    if len(_processed_msg_ids) > 500:
        cutoff = now - 60
        for k in [k for k, v in _processed_msg_ids.items() if v < cutoff]:
            _processed_msg_ids.pop(k, None)

    # 印出來源 ID，方便取得兩個 Group ID
    source_type = event.source.type
    if source_type == "group":
        group_id = event.source.group_id
        # in-memory 記錄所有見過的群組 ID
        if group_id not in _seen_group_ids:
            _seen_group_ids.add(group_id)
        print(f"[GROUP] {group_id}", flush=True)
        # 寫檔備份（含 user_id + display_name）
        try:
            _uid = event.source.user_id
            # 嘗試取得顯示名稱
            _display = ""
            try:
                with ApiClient(_configuration) as _api:
                    _profile = MessagingApi(_api).get_group_member_profile(group_id, _uid)
                    _display = _profile.display_name or ""
            except Exception:
                pass
            log_path = _BASE_DIR / "data" / "group_ids.txt"
            with open(str(log_path), "a", encoding="utf-8") as f:
                from datetime import datetime as _dt
                f.write(f"{_dt.now().strftime('%m-%d %H:%M')} | {group_id} | {_uid} | {_display}\n")
        except Exception as e:
            print(f"[GROUP] 寫檔失敗: {e}", flush=True)

    user_id = event.source.user_id
    text = event.message.text.strip()
    quoted_msg_id = getattr(event.message, 'quoted_message_id', None)
    _store_incoming_text(msg_id, text)

    with ApiClient(_configuration) as api_client:
        line_api = MessagingApi(api_client)

        # 自動記錄客戶 LINE ID + 顯示名稱
        profile = _get_profile_cached(line_api, user_id)
        display_name = profile.display_name if profile else ""
        # 群組訊息且 get_profile 失敗 → 嘗試 get_group_member_profile
        if not display_name and source_type == "group":
            try:
                _gp = line_api.get_group_member_profile(event.source.group_id, user_id)
                display_name = _gp.display_name or ""
            except Exception:
                pass
        try:
            if display_name:  # display_name 空白時不建立空記錄
                customer_store.upsert_from_line(user_id, display_name)
        except Exception as e:
            print(f"[customers] upsert 失敗: {e}")

        # 自動從訊息擷取電話/住址
        try:
            _auto_save_contact_info(user_id, text)
        except Exception as e:
            print(f"[customers] 聯絡資訊擷取失敗: {e}")

        # ── 總公司群組回覆 → 特殊處理，不走一般客戶邏輯 ──────
        if (source_type == "group"
                and settings.LINE_GROUP_ID_HQ
                and event.source.group_id == settings.LINE_GROUP_ID_HQ):
            hq_group_id = event.source.group_id
            # 新增品項 Session 進行中
            _hq_np_state = state_manager.get(user_id)
            if _hq_np_state and _hq_np_state.get("action") == "new_product_session":
                if _INTERNAL_UPLOAD_FINISH_RE.match(text.strip()):
                    _new_prod_timer_cancel(user_id)
                    _lines = _hq_np_state.get("lines", [])
                    state_manager.clear(user_id)
                    ack = handle_internal_new_product("\n".join(_lines)) if _lines else None
                elif _INTERNAL_UPLOAD_CANCEL_RE.match(text.strip()):
                    _new_prod_timer_cancel(user_id)
                    state_manager.clear(user_id)
                    ack = "❌ 已取消新增品項"
                else:
                    _hq_np_state["lines"].append(text)
                    state_manager.set(user_id, _hq_np_state)
                    _new_prod_timer_reset(user_id, hq_group_id, event.reply_token)
                    ack = None
                if ack:
                    line_api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=ack)],
                    ))
                return
            # 新增品項觸發
            if _NEW_PROD_TRIGGER_RE.match(text.strip()):
                # 訊息已含品項資料 → 直接處理，不開 session
                if _split_new_product_entries(text):
                    ack = handle_internal_new_product(text)
                    if ack:
                        line_api.reply_message(ReplyMessageRequest(
                            reply_token=event.reply_token,
                            messages=[TextMessage(text=ack)],
                        ))
                    return
                state_manager.set(user_id, {
                    "action": "new_product_session",
                    "lines":  [text],
                })
                _new_prod_timer_reset(user_id, hq_group_id, event.reply_token)
                ack = "📝 品項建立模式，請依序輸入各品項資料\n完成後傳「完成」，或等待 30 秒自動處理"
                line_api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=ack)],
                ))
                return
            # ── 批次上架 Session 進行中（per-group）──
            _hq_upload_state = state_manager.get(hq_group_id)
            if _hq_upload_state and _hq_upload_state.get("action") == "uploading":
                if _INTERNAL_UPLOAD_FINISH_RE.match(text.strip()):
                    _upload_timer_cancel(hq_group_id)
                    ack = handle_internal_upload_finish(hq_group_id)
                elif _INTERNAL_UPLOAD_CANCEL_RE.match(text.strip()):
                    _upload_timer_cancel(hq_group_id)
                    ack = handle_internal_upload_cancel(hq_group_id)
                else:
                    _upload_timer_reset(hq_group_id, hq_group_id, event.reply_token)
                    ack = handle_internal_upload_text(hq_group_id, text)
                if ack:
                    line_api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=ack)],
                    ))
                return

            # ── 標籤指令 ──
            if text.strip().splitlines()[0].strip().startswith("標籤"):
                ack = handle_internal_label_queue(text)
                if ack:
                    line_api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=ack)],
                    ))
                return

            # ── 觸發批次上架 Session ──
            if text.strip() in _INTERNAL_UPLOAD_TRIGGERS:
                ack = handle_internal_upload_start(hq_group_id)
                if ack:
                    line_api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=ack)],
                    ))
                _upload_timer_reset(hq_group_id, hq_group_id, event.reply_token)
                return

            # 其他 HQ 訊息
            ack = handle_hq_reply(text, line_api)
            if ack:
                line_api.reply_message(
                    ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=ack)],
                    )
                )
            return

        # ── 看貨群 → 只回覆營業時間查詢，其餘全部靜默 ──────────────────
        if (source_type == "group"
                and settings.LINE_GROUP_ID_SHOWCASE
                and event.source.group_id == settings.LINE_GROUP_ID_SHOWCASE):
            if detect_intent(text) == Intent.BUSINESS_HOURS:
                reply = handle_business_hours(text)
                if reply:
                    line_api.reply_message(
                        ReplyMessageRequest(
                            reply_token=event.reply_token,
                            messages=[TextMessage(text=reply)],
                        )
                    )
            return  # 非營業時間查詢 → 靜默不回

        # ── 內部群組 ──────────────────────────────────────────
        if (source_type == "group"
                and settings.LINE_GROUP_ID
                and event.source.group_id == settings.LINE_GROUP_ID):
            # 上架 session 進行中（per-group）→ 跳過 quick_reply，所有文字都當上架內容
            _up_st = state_manager.get(event.source.group_id)
            if _up_st and _up_st.get("action") == "uploading":
                _upload_timer_reset(event.source.group_id, event.source.group_id, event.reply_token)
                print(f"[on_msg] 上架中收到文字，跳過 quick_reply: {text[:30]!r}", flush=True)
                _msg_buf_add(user_id, text=text, context="group",
                             group_id=event.source.group_id,
                             reply_token=event.reply_token)
                return
            # 標籤指令：直接回覆，不走 buffer
            if text.strip().splitlines()[0].strip().startswith("標籤"):
                ack = handle_internal_label_queue(text)
                if ack:
                    _send_reply(event.reply_token, event.source.group_id, ack, line_api)
                return
            # 訊息本身或 buffer 含上架/存文/存圖/加圖/新建品項/新增品項 → 跳過 quick_reply
            _SKIP_QUICK_KW = ("上架", "存圖", "加圖", "存文")
            _skip_quick = any(kw in text for kw in _SKIP_QUICK_KW)
            if not _skip_quick:
                with _msg_buffer_lock:
                    _pending = _msg_buffer.get(user_id)
                    if _pending:
                        _pending_text = "\n".join(_pending["lines"])
                        _skip_quick = any(kw in _pending_text for kw in _SKIP_QUICK_KW)
            if _skip_quick:
                print(f"[on_msg] 含上架指令，跳過 quick_reply: {text[:30]!r}", flush=True)
                _msg_buf_add(user_id, text=text, context="group",
                             group_id=event.source.group_id,
                             reply_token=event.reply_token)
                return
            # 簡單查詢指令：直接 reply 不走 buffer（避免 token 過期）
            from handlers.internal import handle_internal_spec_query as _spec_q
            from handlers.internal import handle_internal_inventory as _inv_q
            from handlers.internal import handle_internal_new_product as _new_prod
            # 含 @ tag 的訊息是對話，不需要 bot 回覆
            _quick_reply = None
            if "@" not in text:
                _quick_reply = (
                    _handle_missing_ecount_name(text)
                    or _new_prod(text)
                    or handle_internal_set_rebate_target(text)
                    or handle_internal_rebate_push(text, line_api)
                    or handle_internal_rebate(text)
                    or handle_internal_unfulfilled(text)
                    or handle_internal_unclaimed(text)
                    or handle_internal_ready_for_pickup(text)
                    or handle_internal_competitor_price(text)
                    or handle_internal_ad_query(text)
                    or _spec_q(text)
                    or _inv_q(text)
                    or handle_internal_order(text, line_api, group_id=event.source.group_id)
                )
            if _quick_reply:
                _send_reply(event.reply_token, event.source.group_id,
                            _quick_reply, line_api)
                return
            # 其餘走文字緩衝，5 秒後統一處理
            _msg_buf_add(user_id, text=text, context="group",
                         group_id=event.source.group_id,
                         reply_token=event.reply_token)
            return

        # ── 機器人關閉中 → 只擋客戶訊息，內部群/總公司群指令不受影響 ──
        if not _bot_active:
            return

        # ── 預設地址偵測（立即寫入 state，flush 時就能用）─────────────
        _cust_for_pref   = customer_store.get_by_line_id(user_id)
        _individual_pref = (
            _cust_for_pref.get("preferred_ecount_cust_cd") if _cust_for_pref else None
        )
        if _individual_pref:
            state_manager.set_group_cust_cd(user_id, _individual_pref)
        elif source_type == "group" and event.source.group_id:
            gid = event.source.group_id
            grp = customer_store.get_group_default(gid)
            if grp:
                state_manager.set_group_cust_cd(user_id, grp["ecount_cust_cd"])
                print(f"[group] 群組預設地址: {grp['ecount_cust_cd']} ({grp.get('label', '')})")
            else:
                print(f"[group] 未知客戶群組: {gid}")
                _unknown_groups[gid] = datetime.now().strftime("%Y-%m-%d %H:%M")

        # ── 真人介入中 → 凍結 bot，不進緩衝 ────────────────────────
        _check_id = group_id if source_type == "group" else user_id
        if (state_manager.get(_check_id) or {}).get("action") == "human_takeover":
            # 雖然靜默，但記錄客戶訊息（讓 Claude 知道接手期間客戶說了什麼）
            from services.claude_ai import add_chat_history as _ach_tk
            _ach_tk(user_id, "user", text)
            print(f"[escalate] 真人介入中，靜默（已記錄）| {_check_id[:10]}...: {text!r}")
            return
        if source_type == "user" and (
            issue_store.has_pending_issue(user_id)
            or delivery_store.has_pending(user_id)
            or pending_store.has_pending(user_id)
        ):
            print(f"[escalate] 真人介入中，靜默 | {user_id[:10]}...: {text!r}")
            return

        # ── 1:1 客戶 → 存文字緩衝，5 秒後統一處理（含圖片合併）────────
        _msg_buf_add(user_id, text=text, context="user",
                     reply_token=event.reply_token,
                     quoted_msg_id=quoted_msg_id)


@_webhook_handler.add(MessageEvent, message=ImageMessageContent)
def on_image_message(event: MessageEvent):
    """
    處理傳來的產品圖片（客戶 or 內部群組）

    收到圖片後先存入緩衝，等待 _IMG_COALESCE_SECS 秒：
    - 若後續有文字（如「5個」「幫XX訂5個」）→ 由 on_message 合併處理
    - 若無後續文字 → timer 到期自動單獨處理（push_message）
    """
    message_id  = event.message.id
    source_type = event.source.type
    user_id     = event.source.user_id

    # ── 內部群組圖片 ─────────────────────────────────────────────────────
    if (source_type == "group"
            and settings.LINE_GROUP_ID
            and event.source.group_id == settings.LINE_GROUP_ID):
        # 批次上架 Session 進行中
        _gid = event.source.group_id
        _st = state_manager.get(_gid)
        if _st and _st.get("action") == "uploading":
            # per-group：直接追加到 group session
            handle_internal_upload_add_media(_gid, message_id, "image")
            _upload_timer_reset(_gid, _gid, event.reply_token)
            return  # 靜默，上架作業全程不回覆，直到完成才通知
        # 一般模式 → 存媒體緩衝，等文字跟上；timer 到期再單獨處理
        _msg_buf_add(user_id, media_msg_id=message_id, media_type="image",
                     context="group", group_id=event.source.group_id,
                     reply_token=event.reply_token)
        return

    # ── 總公司群圖片：上架 Session 進行中 → 追加到 state ──
    if (source_type == "group"
            and settings.LINE_GROUP_ID_HQ
            and event.source.group_id == settings.LINE_GROUP_ID_HQ):
        _hq_gid = event.source.group_id
        _st = state_manager.get(_hq_gid)
        if _st and _st.get("action") == "uploading":
            handle_internal_upload_add_media(_hq_gid, message_id, "image")
            _upload_timer_reset(_hq_gid, _hq_gid, event.reply_token)
            return
        # 非上架 session → 靜默
        return

    # 其他群組靜默
    if source_type == "group":
        return

    # 機器人關閉中 → 靜默不回應（客戶圖片）
    if not _bot_active:
        return

    # ── 真人介入中 → 靜默 ────────────────────────────
    if (state_manager.get(user_id) or {}).get("action") == "human_takeover":
        print(f"[escalate] 真人介入中，圖片靜默 | {user_id[:10]}...")
        return
    if (issue_store.has_pending_issue(user_id)
            or delivery_store.has_pending(user_id)
            or pending_store.has_pending(user_id)):
        print(f"[escalate] 真人介入中，圖片靜默 | {user_id[:10]}...")
        return

    # 自動記錄客戶資料（不影響緩衝邏輯）
    with ApiClient(_configuration) as api_client:
        line_api = MessagingApi(api_client)
        profile = _get_profile_cached(line_api, user_id)
        if profile and profile.display_name:  # display_name 空白時不建立空記錄
            customer_store.upsert_from_line(user_id, profile.display_name)

    # 存入緩衝，等待後續文字（如「5個」「有貨嗎」）
    _msg_buf_add(user_id, media_msg_id=message_id, media_type="image",
                 context="user", reply_token=event.reply_token)


@_webhook_handler.add(MessageEvent, message=VideoMessageContent)
def on_video_message(event: MessageEvent):
    """
    處理傳來的影片（僅內部群組）
    存入媒體緩衝，等待 PO 文字跟上；15 秒後無文字則提示補充。
    """
    source_type = event.source.type
    user_id     = event.source.user_id

    if (source_type == "group"
            and settings.LINE_GROUP_ID
            and event.source.group_id == settings.LINE_GROUP_ID):
        # 批次上架 Session 進行中（per-group）
        _vid_gid = event.source.group_id
        _st = state_manager.get(_vid_gid)
        if _st and _st.get("action") == "uploading":
            handle_internal_upload_add_media(_vid_gid, event.message.id, "video")
            _upload_timer_reset(_vid_gid, _vid_gid, event.reply_token)
            return  # 靜默，上架作業全程不回覆，直到完成才通知
        _msg_buf_add(user_id, media_msg_id=event.message.id, media_type="video",
                     context="group", group_id=event.source.group_id,
                     reply_token=event.reply_token)
        return

    # ── 總公司群影片：上架 Session 進行中 → 追加到 state ──
    if (source_type == "group"
            and settings.LINE_GROUP_ID_HQ
            and event.source.group_id == settings.LINE_GROUP_ID_HQ):
        _hq_vid_gid = event.source.group_id
        _st = state_manager.get(_hq_vid_gid)
        if _st and _st.get("action") == "uploading":
            handle_internal_upload_add_media(_hq_vid_gid, event.message.id, "video")
            _upload_timer_reset(_hq_vid_gid, _hq_vid_gid, event.reply_token)
            return


def _handle_stateful(
    user_id: str, text: str, state: dict, line_api: MessagingApi
) -> str:
    """處理多輪對話狀態"""
    action = state.get("action")

    # ── 等待客戶選擇款式（多筆匹配）──────────────────
    if action == "awaiting_product_clarify":
        from handlers.inventory import _query_single_product
        candidates: list = state.get("candidates", [])   # [(code, name), ...]

        # 支援多選：「3.6.7」「3、6、7」「3 6 7」「1,3,5」
        import re as _re_clarify
        _nums = _re_clarify.findall(r'\d+', text)
        chosen_codes = []
        for n in _nums:
            idx = int(n) - 1
            if 0 <= idx < len(candidates):
                chosen_codes.append(candidates[idx])

        if not chosen_codes:
            # 用品名子字串選
            t = text.strip()
            for code, name in candidates:
                if t in name or name in t:
                    chosen_codes.append((code, name))
                    break

        # 檢查是否要照片
        _want_photo = any(k in text for k in ["照片", "圖片", "圖", "看看", "有沒有照"])

        if len(chosen_codes) == 1 and not _want_photo:
            state_manager.clear(user_id)
            return _query_single_product(user_id, chosen_codes[0][0], line_api)
        elif chosen_codes:
            state_manager.clear(user_id)
            # 組合回覆文字
            from services.ecount import ecount_client as _ec_multi
            lines = [f"好的～以下是您選的 {len(chosen_codes)} 款："]
            for code, name in chosen_codes:
                item = _ec_multi.get_product_cache_item(code)
                price = f"　${int(item['price'])}" if item and item.get("price") and item["price"] > 0 else ""
                lines.append(f"  • {name}（{code}）{price}")
            lines.append(f"\n{tone.boss()}各需要幾個呢？")
            reply_text = "\n".join(lines)

            # 附上圖片（取第一款的圖）
            _img_urls = _get_product_image_urls([c for c, _ in chosen_codes])
            if _img_urls:
                return (reply_text, _img_urls)
            return reply_text
        else:
            return tone.ask_product_clarify(state.get("keyword", ""), candidates)

    # ── 改單/取消偵測（所有 stateful 狀態共用）──────────────
    _CHANGE_QTY_KW = ["改成", "改為", "改"]
    _CANCEL_KW_ORDER = ["取消訂單", "我取消", "不要了", "我不要"]
    _OTHER_CHANGE_KW = ["改訂單", "改數量", "改下單",
                        "減少", "少叫", "多叫", "加訂", "追加",
                        "幫我改", "幫改", "修改訂單"]
    # 購物車有東西 + 「改X個」→ 直接改數量
    if any(kw in text for kw in _CHANGE_QTY_KW):
        _change_qty = extract_quantity(text)
        if _change_qty:
            from storage import cart as _cart_chg
            _chg_cart = _cart_chg.get_cart(user_id)
            if _chg_cart:
                _last = _chg_cart[-1]
                _cart_chg.set_item(user_id, _last["prod_cd"], _last["prod_name"], _change_qty)
                _new_cart = _cart_chg.get_cart(user_id)
                print(f"[stateful] 改數量: {_last['prod_name']} → {_change_qty}", flush=True)
                return tone.cart_item_added(_new_cart)
    if any(kw in text for kw in _CANCEL_KW_ORDER + _OTHER_CHANGE_KW):
        # awaiting_quantity 下的「追加N個」等 → 當成直接指定數量，走下面的 qty 建單流程
        if action == "awaiting_quantity" and extract_quantity(text):
            print(f"[stateful] awaiting_quantity + 追加詞+數字 → 當作 qty，不轉真人: {text!r}", flush=True)
        else:
            state_manager.clear(user_id)
            issue_store.add(user_id, "order_change", text)
            print(f"[stateful] 改單/取消 → 清除狀態，進待處理 user={user_id[:10]}...: {text!r}")
            return "稍等一下喔"

    # ── 等待數量：客戶確認要購買幾個 ──────────────
    if action == "awaiting_quantity":
        # 秒殺擋下 — 內部群已標記「沒貨 XXXX」→ 直接清 state 回擋
        from storage import sold_out as _so_aq
        _aq_pc = state.get("prod_cd", "")
        if _aq_pc and _so_aq.is_sold_out(_aq_pc):
            state_manager.clear(user_id)
            print(f"[sold-out] 擋下 awaiting_quantity {_aq_pc} user={user_id[:10]}...", flush=True)
            return tone.sold_out_secret_kill()
        # 客戶要看照片 → 發產品照片
        _photo_kw = ["照片", "圖片", "看圖", "有圖", "看一下", "長什麼樣", "長怎樣", "看看"]
        if any(kw in text for kw in _photo_kw):
            prod_cd = state.get("prod_cd", "")
            prod_name = state.get("prod_name", "這款")
            if prod_cd:
                _ph_imgs = _get_product_image_urls([prod_cd], max_images=1)
                if _ph_imgs:
                    return (f"這就是「{prod_name}」的照片唷～\n請問{tone.boss()}要幾個呢？", _ph_imgs)
            return f"這款目前沒有照片{tone.suffix_light()} 請問{tone.boss()}需要幾個「{prod_name}」呢？"
        # 問裝箱數：「一箱幾個」「一箱多少」「幾個一箱」「幾入」→ 從 Ecount 規格回
        _pack_kw = ["一箱幾", "一件幾", "一箱多少", "一件多少", "幾個一箱", "幾入一箱",
                    "幾個裝", "幾入裝", "裝箱數", "一箱是幾", "一箱有幾", "一箱有多少",
                    "一箱幾入", "一箱幾個", "箱裝幾"]
        if any(kw in text for kw in _pack_kw):
            prod_cd = state.get("prod_cd", "")
            prod_name = state.get("prod_name", "這款")
            if prod_cd:
                from handlers.ordering import detect_per_box as _dpb
                from services.ecount import ecount_client as _ec_pk
                _per_box = _dpb(prod_cd)
                _item_pk = _ec_pk.get_product_cache_item(prod_cd)
                _size_des = (_item_pk.get("size_des") if _item_pk else "") or ""
                if _per_box > 0:
                    _ub = "入" if ("入" in _size_des) else "個"
                    import re as _re_pk
                    _m_unit = _re_pk.search(r'\d+\s*(盒|個|入|包|罐|條|瓶)', _size_des)
                    if _m_unit:
                        _ub = _m_unit.group(1)
                    return f"「{prod_name}」1箱 = {_per_box}{_ub} 唷{tone.suffix_light()}"
                if _size_des:
                    return f"「{prod_name}」規格：{_size_des}{tone.suffix_light()}"
            return f"裝箱數我確認一下再回覆{tone.boss()}哦"
        # 問庫存數量攔截：「有多少個」「大概幾個」「還剩多少」→ 不透露數量
        _ask_qty_kw = ["多少個", "幾個", "多少", "剩多少", "剩幾", "有幾個", "有多少"]
        if any(kw in text for kw in _ask_qty_kw):
            prod_name = state.get("prod_name", "")
            issue_store.add(user_id, "ask_stock_qty", f"客戶問數量：{prod_name}")
            return tone.ask_qty_deflect()
        # 問句排除：「是不是三個顏色」「有幾個顏色」不是回答數量
        _question_kw = ["請問", "是不是", "是否", "有幾", "幾個", "多少", "嗎", "呢", "？", "?", "顏色", "款式", "種"]
        _is_question = any(kw in text for kw in _question_kw)
        qty = extract_quantity(text)
        if qty and not _is_question:
            # 統一直接加購物車（不再多一步確認，避免 state 被覆蓋）
            state_manager.clear(user_id)
            return handle_order_quantity(user_id, text, state, line_api)
        elif qty and _is_question:
            # 有數字但是問句 → 清 state，走正常 dispatch（交給 Claude 回答）
            state_manager.clear(user_id)
            return None
        elif any(kw in text for kw in ["不要", "算了", "取消", "不訂", "不用",
                                        "好吧", "謝謝", "感謝", "沒關係", "不用了",
                                        "好的謝謝", "好吧謝謝", "下次再說", "先不要"]):
            state_manager.clear(user_id)
            import random as _random
            b = tone.boss()
            return _random.choice([
                f"好的{tone.suffix_light()} {b}有需要再找我哦",
                f"沒問題～{b}有需要隨時說哦",
                f"好的，{b}有需要再告訴我{tone.suffix_light()}",
            ])
        else:
            # 明顯是新查詢（有沒有、還有、推薦等）→ 清 state，走正常流程
            _new_query_kw = ["有沒有", "有什麼", "還有", "推薦", "其他", "別的",
                            "元左右", "塊左右", "元以下", "塊以下", "多元", "多塊",
                            "貨運", "寄貨", "寄送", "要寄", "送貨", "宅配", "配送"]
            if any(kw in text for kw in _new_query_kw):
                state_manager.clear(user_id)
                print(f"[stateful] awaiting_quantity 但像新查詢，清除狀態: {text[:20]!r}", flush=True)
                return None
            # 訊息明顯不是回答數量（問句、其他意圖）→ 清 state，走正常流程
            _other_intent = detect_intent(text)
            if _other_intent and _other_intent not in (Intent.UNKNOWN, Intent.CONFIRMATION):
                state_manager.clear(user_id)
                print(f"[stateful] awaiting_quantity 但意圖={_other_intent.value}，清除狀態", flush=True)
                return None  # 回 None 讓訊息走正常 dispatch
            # 非數字回覆 → 走 dispatch 正常回覆，但盡量保留 state
            # （客戶想下單時會 tag 訊息+數量，sent_image_map 能追蹤到貨號）
            _saved_state = dict(state)
            print(f"[stateful] awaiting_quantity 但非數字回覆，走 dispatch: {text[:20]!r}", flush=True)
            state_manager.clear(user_id)
            _fallback_intent = detect_intent(text)
            _fallback_reply = _dispatch(user_id, text, _fallback_intent, line_api)
            # 只有 dispatch 沒設新 state 時才還原（避免蓋掉新流程）
            if not state_manager.get(user_id):
                state_manager.set(user_id, _saved_state)
                print(f"[stateful] 還原 awaiting_quantity state", flush=True)
            return _fallback_reply if _fallback_reply else ""

    # ── 多產品清單：客戶問照片 → 發照片（翻頁），其他 → 清 state 走正常流程 ──
    elif action == "recent_products":
        _photo_kw = ["照片", "圖片", "看圖", "有圖", "看一下", "看看", "給我看", "其他"]
        if any(kw in text for kw in _photo_kw):
            # 優先：客戶指定了貨號 → 發指定的照片 + 設 awaiting_quantity
            _specified = _PROD_CODE_RE.findall(text)
            if _specified:
                _sc = _specified[0].upper()
                _sc_imgs = _get_product_image_urls([_sc], max_images=1)
                from services.ecount import ecount_client as _ec_rp
                _sc_item = _ec_rp.get_product_cache_item(_sc)
                _sc_name = (_sc_item.get("name") if _sc_item else None) or _sc
                state_manager.set(user_id, {
                    "action":    "awaiting_quantity",
                    "prod_cd":   _sc,
                    "prod_name": _sc_name,
                })
                if _sc_imgs:
                    return (f"這是「{_sc_name}」的照片唷～\n請問要幾個呢？", _sc_imgs)
                return f"「{_sc_name}」目前沒有照片{tone.suffix_light()} 請問要幾個呢？"
            _rp_codes = state.get("prod_codes", [])
            _rp_offset = state.get("photo_offset", 0)
            if _rp_codes:
                _page_codes = _rp_codes[_rp_offset:_rp_offset + 4]
                if not _page_codes:
                    # 已經看完所有照片，從頭來
                    _page_codes = _rp_codes[:4]
                    _rp_offset = 0
                _rp_imgs = _get_product_image_urls(_page_codes, max_images=4)
                _next_offset = _rp_offset + 4
                _has_more = _next_offset < len(_rp_codes)
                if _has_more:
                    state_manager.set(user_id, {
                        "action":       "recent_products",
                        "prod_codes":   _rp_codes,
                        "photo_offset": _next_offset,
                    })
                    _hint = f"\n還有其他照片，說「還有嗎」繼續看唷～"
                else:
                    state_manager.clear(user_id)
                    _hint = "\n有喜歡的跟我說要哪個、幾個！"
                if _rp_imgs:
                    return (f"這些是產品照片唷～{_hint}", _rp_imgs)
            state_manager.clear(user_id)
            return "目前這些產品沒有照片，有喜歡的跟我說要哪個、幾個！"
        # 非照片請求 → 清 state，走正常流程
        state_manager.clear(user_id)
        return None

    # ── 模糊描述產品確認：附照片問「是這款嗎？」──────
    elif action == "awaiting_product_confirm":
        prod_cd   = state.get("prod_cd", "")
        prod_name = state.get("prod_name", "此商品")
        qty       = state.get("qty", 1)
        note      = state.get("note", "")
        if any(kw in text for kw in _YES_KW):
            state_manager.clear(user_id)
            from storage import cart as _cart_confirm
            _cart_confirm.add_item(user_id, prod_cd, prod_name, qty, note=note)
            print(f"[stateful] awaiting_product_confirm 確認 → 加購物車: {prod_cd} x{qty}", flush=True)
            return tone.cart_item_added(_cart_confirm.get_cart(user_id))
        elif any(kw in text for kw in _NO_KW):
            state_manager.clear(user_id)
            print(f"[stateful] awaiting_product_confirm 取消: {prod_cd}", flush=True)
            return f"好的{tone.suffix_light()} 那請問{tone.boss()}要找的是哪款呢？可以提供更多資訊嗎？"
        else:
            return f"請問是「{prod_name}」這款嗎？跟我說「是」或「不是」唷"

    # ── 多圖+各X 確認：等待確認送出或其他回覆靜默 ──────
    elif action == "awaiting_multi_img_confirm":
        if any(kw == text.strip() for kw in _YES_KW):
            state_manager.clear(user_id)
            from handlers.ordering import handle_checkout
            return handle_checkout(user_id, line_api)
        else:
            # 非確認 → 靜默，加入待處理
            state_manager.clear(user_id)
            issue_store.add(user_id, "multi_img_order", f"客戶回覆：{text}")
            return None

    # ── 箱/盒確認 ──────────────────────────────────
    elif action == "awaiting_box_confirm":
        prod_cd = state.get("prod_cd", "")
        prod_name = state.get("prod_name", "")

        # 嘗試提取數量
        _box_qty = extract_quantity(text)
        if _box_qty:
            state_manager.clear(user_id)
            from storage import cart as _cart_box
            _cart_box.add_item(user_id, prod_cd, prod_name, _box_qty)
            return tone.cart_item_added(_cart_box.get_cart(user_id))
        elif any(kw in text for kw in ["不要", "算了", "取消", "不訂", "不用"]):
            state_manager.clear(user_id)
            return f"好的{tone.suffix_light()} 已取消"
        else:
            return f"請問「{prod_name}」要幾箱呢？"

    # ── 圖片下單確認框：等待「確認」或「取消」──────
    elif action == "awaiting_image_order_confirm":
        prod_cd   = state.get("prod_cd", "")
        prod_name = state.get("prod_name", "此商品")
        qty       = state.get("qty", 1)
        if any(kw in text for kw in _YES_KW):
            state_manager.clear(user_id)
            # 加入購物車後立即結帳
            from handlers.ordering import handle_checkout
            from storage import cart as cart_store
            cart_store.add_item(user_id, prod_cd, prod_name, qty)
            return handle_checkout(user_id, line_api)
        elif any(kw in text for kw in _NO_KW):
            state_manager.clear(user_id)
            return f"好的{tone.suffix_light()} 已取消，{tone.boss()}有需要再找我哦"
        else:
            # 非確認/取消 → 重新顯示確認框
            return (
                f"確認是這款對吧～\n"
                f"{prod_name}（{prod_cd}）× {qty} 個"
            )

    # ── 等待數量（缺貨）：舊 state 相容 → 轉為購物車流程 ─────
    elif action == "awaiting_restock_qty":
        qty = extract_quantity(text)
        if qty:
            prod_name = state.get("prod_name", "此商品")
            prod_cd = state.get("prod_cd", "")
            state_manager.clear(user_id)
            # 加入購物車（統一走訂單流程）
            from storage import cart as _cart_restock
            _cart_restock.add_item(user_id, prod_cd, prod_name, qty)
            return tone.cart_item_added(_cart_restock.get_cart(user_id))
        elif any(kw in text for kw in ["不要", "算了", "取消", "不訂", "不用",
                                       "收到", "好", "好的", "知道了", "了解",
                                       "知道", "ok", "OK", "謝謝", "感謝"]):
            state_manager.clear(user_id)
            return f"好的{tone.suffix_light()} 有需要再找我哦"
        else:
            # 訊息明顯不是回答數量（問句、其他意圖）→ 清 state，走正常流程
            _other_intent = detect_intent(text)
            if _other_intent and _other_intent not in (Intent.UNKNOWN, Intent.CONFIRMATION):
                state_manager.clear(user_id)
                print(f"[stateful] awaiting_restock_qty 但意圖={_other_intent.value}，清除狀態", flush=True)
                return None
            prod_name = state.get("prod_name", "這款")
            return tone.ask_quantity(prod_name)

    # ── 等待客戶確認是否願意等叫貨 ────────────────────────
    elif action == "awaiting_wait_confirm":
        prod_name = state.get("prod_name", "此商品")
        prod_cd = state.get("prod_cd", "")
        qty = state.get("qty", 1)
        wait_time = state.get("wait_time", "")
        restock_id = state.get("restock_id")

        if any(kw in text for kw in _YES_KW):
            from storage.restock import restock_store
            from storage.customers import customer_store as _cs
            from services.ecount import ecount_client

            # 多地址 → 先問客戶選哪個地址（群組有預設則問確認）
            codes = _cs.get_ecount_codes_by_line_id(user_id)
            if len(codes) > 1:
                preferred = state_manager.get_group_cust_cd(user_id)
                if preferred:
                    pref_label = next(
                        (c.get("address_label") or c.get("cust_name") or preferred
                         for c in codes if c["ecount_cust_cd"] == preferred),
                        preferred
                    )
                    state_manager.set(user_id, {
                        "action":     "awaiting_address_selection",
                        "prod_cd":    prod_cd,
                        "prod_name":  prod_name,
                        "qty":        qty,
                        "restock_id": restock_id,
                        "preferred_cust_cd": preferred,
                    })
                    return tone.ask_group_address_confirm(pref_label)
                state_manager.set(user_id, {
                    "action":     "awaiting_address_selection",
                    "prod_cd":    prod_cd,
                    "prod_name":  prod_name,
                    "qty":        qty,
                    "restock_id": restock_id,
                })
                return tone.ask_address_selection(codes)

            state_manager.clear(user_id)
            cust_code = (
                codes[0]["ecount_cust_cd"] if codes
                else _cs.get_ecount_cust_code(
                    user_id, default=settings.ECOUNT_DEFAULT_CUST_CD
                )
            )
            slip_no = ecount_client.save_order(
                cust_code=cust_code,
                items=[{"prod_cd": prod_cd, "qty": qty}],
                phone=_user_phone(user_id),
            )
            if restock_id:
                restock_store.update_status(restock_id, "confirmed")
            if slip_no:
                return tone.restock_wait_confirmed(prod_name, qty, slip_no)
            else:
                print(f"[ordering] 訂單建立失敗（wait_confirm）: {cust_code} | {prod_name} x{qty}")
                issue_store.add(user_id, "order_failed", f"{prod_name} × {qty} 個（等待確認後）")
                return None
        elif any(kw in text for kw in _NO_KW):
            state_manager.clear(user_id)
            from storage.restock import restock_store
            if restock_id:
                restock_store.update_status(restock_id, "cancelled")
            # 登記到貨通知：有現貨時自動 push 給客戶
            if prod_cd:
                notify_store.add(user_id, prod_cd, prod_name, qty)
                print(f"[notify] 登記到貨通知：{prod_name}（{prod_cd}）x{qty} for {user_id[:10]}...")
            return tone.restock_wait_declined(prod_name)
        else:
            return tone.restock_wait_ask(prod_name, qty, wait_time)

    # ── 等待地址選擇（多地址客戶下單）──────────────
    elif action == "awaiting_address_selection":
        prod_cd          = state.get("prod_cd", "")
        prod_name        = state.get("prod_name", "")
        qty              = state.get("qty", 1)
        restock_id       = state.get("restock_id")
        preferred_cust_cd = state.get("preferred_cust_cd")  # 群組預設地址

        codes = customer_store.get_ecount_codes_by_line_id(user_id)

        # 群組預設地址確認流程：接受「是/否」
        if preferred_cust_cd:
            if any(kw in text for kw in _YES_KW):
                cust_code = preferred_cust_cd
                addr_label = next(
                    (c.get("address_label") or "" for c in codes if c["ecount_cust_cd"] == preferred_cust_cd), ""
                )
                is_undecided = "未決定" in addr_label
                state_manager.clear(user_id)
                state_manager.clear_group_cust_cd(user_id)
                from services.ecount import ecount_client
                slip_no = ecount_client.save_order(
                    cust_code=cust_code,
                    items=[{"prod_cd": prod_cd, "qty": qty}],
                    phone=_user_phone(user_id),
                )
                if restock_id:
                    restock_store.update_status(restock_id, "confirmed")
                if slip_no:
                    print(f"[ordering] 群組預設地址訂單: {slip_no} | {cust_code} | {prod_name} x{qty}")
                    if is_undecided:
                        issue_store.add(user_id, "address_pending",
                            f"{prod_name} × {qty} 個（單號 {slip_no}，待確認配送地址）")
                    return tone.order_confirmed(prod_name, qty, slip_no)
                else:
                    issue_store.add(user_id, "order_failed", f"{prod_name} × {qty} 個（群組地址後）")
                    return None
            elif any(kw in text for kw in _NO_KW):
                # 移除 preferred，改為一般地址選擇
                state_manager.set(user_id, {
                    "action":     "awaiting_address_selection",
                    "prod_cd":    prod_cd,
                    "prod_name":  prod_name,
                    "qty":        qty,
                    "restock_id": restock_id,
                })
                return tone.ask_address_selection(codes)
            else:
                # 重新確認
                pref_label = next(
                    (c.get("address_label") or c.get("cust_name") or preferred_cust_cd
                     for c in codes if c["ecount_cust_cd"] == preferred_cust_cd),
                    preferred_cust_cd
                )
                return tone.ask_group_address_confirm(pref_label)

        m = re.search(r"^\s*(\d+)\s*$", text)
        if m:
            idx = int(m.group(1)) - 1
            if 0 <= idx < len(codes):
                selected     = codes[idx]
                cust_code    = selected["ecount_cust_cd"]
                addr_label   = (selected.get("address_label") or "").strip()
                is_undecided = "未決定" in addr_label
                state_manager.clear(user_id)
                from services.ecount import ecount_client
                slip_no = ecount_client.save_order(
                    cust_code=cust_code,
                    items=[{"prod_cd": prod_cd, "qty": qty}],
                    phone=_user_phone(user_id),
                )
                if restock_id:
                    restock_store.update_status(restock_id, "confirmed")
                if slip_no:
                    print(f"[ordering] 訂單建立成功: {slip_no} | {cust_code} | {prod_name} x{qty}")
                    if is_undecided:
                        issue_store.add(user_id, "address_pending",
                            f"{prod_name} × {qty} 個（單號 {slip_no}，待確認配送地址）")
                        print(f"[ordering] 地址未決定，已記錄待處理: {slip_no}")
                    return tone.order_confirmed(prod_name, qty, slip_no)
                else:
                    print(f"[ordering] 訂單建立失敗（address_selection）: {cust_code} | {prod_name} x{qty}")
                    issue_store.add(user_id, "order_failed", f"{prod_name} × {qty} 個（地址選擇後）")
                    return None

        # 無效輸入 → 重新顯示選項
        return tone.ask_address_selection(codes)

    # ── 等待地址選擇（購物車結帳版）──────────────────
    elif action == "awaiting_address_selection_checkout":
        from storage import cart as cart_store
        cart = cart_store.get_cart(user_id)
        codes = customer_store.get_ecount_codes_by_line_id(user_id)
        # 結帳/確認詞（客戶重複 tag 清單說「好了/確認/送出」）→ 靜默不再重問
        _CONF_NOISE = set(CHECKOUT_KEYWORDS) | set(AFFIRMATIVE_KEYWORDS)
        _txt_strip_addr = text.strip()
        if _txt_strip_addr in _CONF_NOISE or any(k in _txt_strip_addr for k in
                                                 ["確認訂單", "送出訂單", "幫我送出", "幫忙送出"]):
            print(f"[addr_select] 已在選地址狀態，忽略結帳詞: {_txt_strip_addr!r}", flush=True)
            return None
        # 取消 → 清 state
        if any(k in _txt_strip_addr for k in ["取消", "不要了", "算了"]):
            state_manager.clear(user_id)
            return f"好的{tone.suffix_light()} 已取消"
        m = re.search(r"^\s*(\d+)\s*$", text)
        if m:
            idx = int(m.group(1)) - 1
            if 0 <= idx < len(codes):
                selected     = codes[idx]
                cust_code    = selected["ecount_cust_cd"]
                addr_label   = (selected.get("address_label") or "").strip()
                is_undecided = "未決定" in addr_label
                state_manager.clear(user_id)
                from services.ecount import ecount_client as _ec
                items   = [{"prod_cd": i["prod_cd"], "qty": i["qty"]} for i in cart]
                slip_no = _ec.save_order(cust_code=cust_code, items=items, phone=_user_phone(user_id))
                if slip_no:
                    cart_store.clear_cart(user_id)
                    if is_undecided:
                        desc = "、".join(f"{i['prod_name']}×{i['qty']}" for i in cart)
                        issue_store.add(user_id, "address_pending",
                            f"{desc}（單號 {slip_no}，待確認配送地址）")
                        print(f"[ordering] 地址未決定，已記錄待處理: {slip_no}")
                    return tone.checkout_confirmed(cart)
                else:
                    desc = "、".join(f"{i['prod_name']}×{i['qty']}" for i in cart)
                    issue_store.add(user_id, "order_failed", desc)
                    return None
        return tone.ask_address_selection(codes)

    # ── 等待客戶提供聯絡資料（姓名 + 手機） ──────────
    elif action == "awaiting_contact_info":
        prod_cd   = state.get("prod_cd", "")
        prod_name = state.get("prod_name", "")
        qty       = state.get("qty", 1)

        phone_match = _PHONE_RE.search(text)
        if phone_match:
            phone_str = phone_match.group()
            try:
                customer_store.update_phone(user_id, phone_str)
            except Exception:
                pass
            name_text = _PHONE_RE.sub("", text).strip()
            if name_text:
                try:
                    customer_store.update_real_name(user_id, name_text)
                except Exception:
                    pass

            state_manager.clear(user_id)
            from handlers.ordering import _resolve_cust_code, _create_ecount_customer
            from services.ecount import ecount_client as _ec

            # 第二次 JSON 比對（不重複同步，已在 handle_checkout 階段同步過）
            cust_code = _resolve_cust_code(user_id, do_refresh=False)
            if not cust_code:
                # JSON 仍找不到 → Ecount API 建立新客戶
                cust_code = _create_ecount_customer(user_id)

            if not cust_code:
                # 全部失敗 → 寫入待處理，客戶端不回應
                print(f"[ordering] 客戶代碼解析全失敗: {user_id} | {prod_name} x{qty}")
                issue_store.add(user_id, "order_failed", f"{prod_name} × {qty} 個")
                return None

            slip_no = _ec.save_order(
                cust_code=cust_code,
                items=[{"prod_cd": prod_cd, "qty": qty}],
                phone=phone_str,
            )
            if slip_no:
                print(f"[ordering] 訂單建立成功: {slip_no} | {cust_code} | {prod_name} x{qty}")
                return tone.order_confirmed(prod_name, qty, slip_no)
            else:
                print(f"[ordering] 訂單建立失敗: {cust_code} | {prod_name} x{qty}")
                issue_store.add(user_id, "order_failed", f"{prod_name} × {qty} 個")
                return None
        else:
            return tone.ask_contact_info()

    # ── 等待聯絡資料（購物車結帳版）──────────────────
    elif action == "awaiting_contact_info_checkout":
        phone_match = _PHONE_RE.search(text)
        if phone_match:
            phone_str = phone_match.group()
            try:
                customer_store.update_phone(user_id, phone_str)
            except Exception:
                pass
            name_text = _PHONE_RE.sub("", text).strip()
            if name_text:
                try:
                    customer_store.update_real_name(user_id, name_text)
                except Exception:
                    pass
            state_manager.clear(user_id)
            # 交回 handle_checkout 統一處理：建 Ecount 客戶 + 登記到貨通知 + 送出訂單
            # （之前直接 _ec.save_order 漏掉 notify_store.add → 到貨通知沒登記）
            from handlers.ordering import handle_checkout
            return handle_checkout(user_id, line_api)
        else:
            return tone.ask_contact_info()

    # ── 等待產品名稱 ───────────────────────────────
    elif action == "awaiting_product":
        # 長訊息且不含貨號 → 使用者已換話題，清除狀態並轉真人
        _has_prod_code = bool(re.search(r'[A-Za-z]{1,3}-?\d{3,6}', text))
        if len(text) > 20 and not _has_prod_code:
            state_manager.clear(user_id)
            from handlers.escalate import handle_unknown
            return handle_unknown(user_id, text, line_api)
        state_manager.clear(user_id)
        from handlers.inventory import query_product
        return query_product(user_id, text, line_api)

    # ── 等待客戶說要登記通知的產品名稱 ────────────
    elif action == "awaiting_notify_product":
        # 長訊息且不含貨號 → 客戶已經換話題，清狀態並轉真人（避免把整句話當產品名查）
        _has_prod_code = bool(re.search(r'[A-Za-z]{1,3}-?\d{3,6}', text))
        if len(text) > 20 and not _has_prod_code:
            state_manager.clear(user_id)
            from handlers.escalate import handle_unknown
            return handle_unknown(user_id, text, line_api)
        state_manager.clear(user_id)
        from services.ecount import ecount_client as _ec
        from storage.notify import notify_store
        result_item = _ec.lookup(text)
        if not result_item:
            # 查不到 → pending_store 等真人確認
            from storage.pending import pending_store
            pending_store.add(user_id, text)
            return tone.product_not_found(text)
        prod_code = result_item["code"]
        prod_name = result_item["name"] or prod_code
        qty = result_item.get("qty", 0)
        if qty and qty > 0:
            # 有貨 → 進下單流程
            state_manager.set(user_id, {
                "action":    "awaiting_quantity",
                "prod_cd":   prod_code,
                "prod_name": prod_name,
            })
            return tone.notify_request_in_stock(prod_name)
        # 無貨 → 登記到貨通知
        notify_store.add(user_id, prod_code, prod_name, 1)
        print(f"[notify] 登記到貨通知：{prod_name}（{prod_code}）for {user_id[:10]}...")
        return tone.notify_request_ack(prod_name)

    # ── 等待訂單編號 ───────────────────────────────
    elif action == "awaiting_order_id":
        state_manager.clear(user_id)
        from handlers.orders import handle_order_tracking
        return handle_order_tracking(user_id, text)

    # ── 群組預設地址確認（是/否）──────────────────
    elif action == "awaiting_group_address_confirm":
        from storage import cart as cart_store
        from services.ecount import ecount_client

        preferred_cust_cd = state_manager.get_group_cust_cd(user_id)
        if not preferred_cust_cd:
            # 無預設代碼（不應發生），降級為一般地址選擇
            codes = customer_store.get_ecount_codes_by_line_id(user_id)
            state_manager.set(user_id, {"action": "awaiting_address_selection_checkout"})
            return tone.ask_address_selection(codes)

        if any(kw in text for kw in _YES_KW):
            state_manager.clear(user_id)
            state_manager.clear_group_cust_cd(user_id)
            cart = cart_store.get_cart(user_id)
            items = [{"prod_cd": i["prod_cd"], "qty": i["qty"]} for i in cart]
            slip_no = ecount_client.save_order(cust_code=preferred_cust_cd, items=items, phone=_user_phone(user_id))
            if slip_no:
                print(f"[ordering] 群組地址訂單建立成功: {slip_no} | {preferred_cust_cd}")
                cart_store.clear_cart(user_id)
                return tone.checkout_confirmed(cart)
            else:
                print(f"[ordering] 群組地址訂單建立失敗: {preferred_cust_cd}")
                desc = "、".join(f"{i['prod_name']}×{i['qty']}" for i in cart)
                issue_store.add(user_id, "order_failed", desc)
                return None
        elif any(kw in text for kw in _NO_KW):
            # 改選其他地址
            codes = customer_store.get_ecount_codes_by_line_id(user_id)
            state_manager.set(user_id, {"action": "awaiting_address_selection_checkout"})
            return tone.ask_address_selection(codes)
        else:
            # 重新詢問確認
            codes = customer_store.get_ecount_codes_by_line_id(user_id)
            pref_label = next(
                (c.get("address_label") or c.get("cust_name") or preferred_cust_cd
                 for c in codes if c["ecount_cust_cd"] == preferred_cust_cd),
                preferred_cust_cd
            )
            return tone.ask_group_address_confirm(pref_label)

    else:
        state_manager.clear(user_id)
        from handlers.escalate import handle_unknown
        return handle_unknown(user_id, text, line_api)


_PRODUCT_PHOTO_DIR = Path(r"H:\其他電腦\我的電腦\小蠻牛\產品照片")


def _push_product_images(user_id: str, prod_codes: list[str], line_api) -> None:
    """找到產品照片後 push 給客戶（最多每個產品 1 張，總共最多 5 張）"""
    from linebot.v3.messaging import PushMessageRequest, ImageMessage
    from config import settings as _cfg

    if not _PRODUCT_PHOTO_DIR.exists() or _push_quota_exhausted:
        print(f"[recommend] 跳過 push: dir={_PRODUCT_PHOTO_DIR.exists()} quota={_push_quota_exhausted}", flush=True)
        return

    images = []
    for code in prod_codes:
        if len(images) >= 5:
            break
        # 找第一張 jpg（A 版優先）
        for suffix in ["A.jpg", "B.jpg", "C.jpg", ".jpg", "A.png", ".png"]:
            p = _PRODUCT_PHOTO_DIR / f"{code}{suffix}"
            if p.exists():
                # 需要一個公開 URL → 用 server 的靜態路由提供
                url = f"https://xmnline.duckdns.org/product-photo/{code}{suffix}"
                images.append(ImageMessage(
                    original_content_url=url,
                    preview_image_url=url,
                ))
                break

    if not images:
        print(f"[recommend] 找不到圖片: {prod_codes}", flush=True)
        return
    print(f"[recommend] 準備 push {len(images)} 張圖: {prod_codes}", flush=True)
    try:
        resp = line_api.push_message(PushMessageRequest(
            to=user_id, messages=images,
        ))
        # 記錄圖片 msg_id → product_code（純圖無前置文字，所以從 index 0 開始對應）
        if hasattr(resp, 'sent_messages') and resp.sent_messages:
            with _sent_image_map_lock:
                _ts = __import__('time').time()
                for i, code in enumerate(prod_codes[:len(resp.sent_messages)]):
                    _sent_image_map[resp.sent_messages[i].id] = {"code": code, "ts": _ts}
                _save_sent_image_map()
            print(f"[recommend] 記錄 {len(images)} 張圖 msg_id → 貨號", flush=True)
        print(f"[recommend] push {len(images)} 張產品圖給 {user_id[:10]}...", flush=True)
    except Exception as e:
        if _is_quota_429(e):
            _mark_push_exhausted()
        else:
            print(f"[recommend] push 圖片失敗: {e}", flush=True)


def _get_product_image_urls(prod_codes: list[str], max_images: int = 5) -> list[str]:
    """找產品照片 URL（最多 max_images 張）"""
    urls = []
    if not _PRODUCT_PHOTO_DIR.exists():
        return urls
    for code in prod_codes:
        if len(urls) >= max_images:
            break
        for suffix in ["A.jpg", "B.jpg", "C.jpg", ".jpg", "A.png", ".png"]:
            p = _PRODUCT_PHOTO_DIR / f"{code}{suffix}"
            if p.exists():
                urls.append(f"https://xmnline.duckdns.org/product-photo/{code}{suffix}")
                break
    return urls


_RECOMMEND_EXCLUDE_KW = {"泡澡球", "洗衣球", "衛生紙", "抽取式", "面紙"}


def _is_recommend_excluded(code: str, name: str) -> bool:
    """排除不推薦的品項"""
    if code.upper().startswith("HH"):
        return True
    for kw in _RECOMMEND_EXCLUDE_KW:
        if kw in name:
            return True
    return False


def _get_recommend_hint() -> str:
    """取得推薦品項提示：最新5個有現貨 + 庫存最多3個 + 預購品"""
    import json
    from services.ecount import ecount_client
    from handlers.inventory import _load_preorder_cache

    avail_path = Path(__file__).parent / "data" / "available.json"
    if not avail_path.exists():
        return ""
    # 超過 30 分鐘未更新 → 先同步
    import time as _t
    if _t.time() - avail_path.stat().st_mtime > 30 * 60:
        try:
            from services.ecount import ecount_client
            ecount_client._ensure_available()
        except Exception:
            pass
    try:
        avail = json.loads(avail_path.read_text(encoding="utf-8"))
    except Exception:
        return ""

    # 建立品名查詢
    def _get_name(code):
        info = ecount_client.get_product_cache_item(code)
        return (info.get("name") if info else None) or code

    # ── 最新 10 個有現貨的（從 PO文尾端找，給 Claude 選）──
    po_path = Path(r"H:\其他電腦\我的電腦\小蠻牛\產品PO文.txt")
    newest = []
    if po_path.exists():
        try:
            text = po_path.read_text(encoding="utf-8")
            blocks = re.split(r"\n{2,}", text)
            for block in reversed(blocks):
                if len(newest) >= 10:
                    break
                m = re.search(r"([A-Za-z]{1,3}-?\d{3,6})", block)
                if not m:
                    continue
                code = m.group(1).upper()
                if code in [c for c, _ in newest]:
                    continue
                qty = 0
                if code in avail:
                    d = avail[code]
                    qty = d.get("available", 0) if isinstance(d, dict) else d
                if qty <= 0:
                    continue
                name = _get_name(code)
                if _is_recommend_excluded(code, name):
                    continue
                newest.append((code, name))
        except Exception:
            pass

    # ── 庫存最多 3 個 ──
    stock_items = []
    for code, data in avail.items():
        qty = data.get("available", 0) if isinstance(data, dict) else data
        if qty <= 0:
            continue
        name = _get_name(code)
        if _is_recommend_excluded(code, name):
            continue
        stock_items.append((code, name, qty))
    stock_items.sort(key=lambda x: x[2], reverse=True)
    top_stock = [(c, n) for c, n, _ in stock_items[:10]]

    # ── 預購品 ──
    po_cache = _load_preorder_cache()
    preorder = []
    for code, info in po_cache.items():
        name = info.get("name", code) if isinstance(info, dict) else code
        eta = info.get("eta", "") if isinstance(info, dict) else ""
        if not _is_recommend_excluded(code, name):
            preorder.append((code, name, eta))

    # 組合
    parts = []
    if newest:
        lines = ["【最新上架有現貨（從中挑 5 個推薦）】"]
        for code, name in newest:
            lines.append(f"{code}: {name}（有現貨）")
        parts.append("\n".join(lines))
    if top_stock:
        lines = ["【庫存充足的熱門品項（從中挑 3 個推薦）】"]
        for code, name in top_stock:
            lines.append(f"{code}: {name}（庫存充足）")
        parts.append("\n".join(lines))
    if preorder:
        lines = ["【預購品（全部推薦，註明預購和到貨時間）】"]
        for code, name, eta in preorder:
            eta_str = f"，{eta}" if eta else ""
            lines.append(f"{code}: {name}（預購{eta_str}）")
        parts.append("\n".join(lines))

    return "\n\n".join(parts)


def _handle_recommendation(user_id: str, text: str, line_api) -> str | tuple:
    """推薦/新貨查詢 → 交給 Claude 回答，附帶產品圖"""
    from services.claude_ai import ask_claude_text

    # 額外附上推薦品項提示，讓 Claude 推薦
    hint = _get_recommend_hint()
    if hint:
        enhanced_text = f"{text}\n\n⚠️ 嚴格規定：只能從以下清單中推薦，不要推薦清單以外的商品！\n\n{hint}"
    else:
        enhanced_text = text

    reply = ask_claude_text(enhanced_text, user_id=user_id)
    if not reply:
        return "讓我看看有什麼好東西，稍等一下唷～"

    reply = _strip_oos_lines_from_recommendation(reply)

    # chat_history 由 _send_reply 在送出成功後自動記（caller 會做）

    # 從 Claude 回覆中提取貨號 → 附帶產品圖
    codes_raw = _PROD_CODE_RE.findall(reply)
    codes = list(dict.fromkeys(c.upper() for c in codes_raw))
    image_urls = _get_product_image_urls(codes) if codes else []

    if image_urls:
        return (reply, image_urls)
    return reply


def _strip_oos_lines_from_recommendation(reply: str) -> str:
    """硬阻擋：把 Claude 回覆裡缺貨的推薦行刪掉（預購品保留）。
    Why: Claude 偶爾無視 system prompt「只推薦有現貨」的規則，
    需要在輸出端兜底，避免推到客戶手上的清單含缺貨品。"""
    import json as _json
    from handlers.inventory import _load_preorder_cache
    from handlers.internal import _PROD_CODE_RE as _PCR

    avail_path = Path(__file__).parent / "data" / "available.json"
    if not avail_path.exists():
        return reply
    try:
        avail = _json.loads(avail_path.read_text(encoding="utf-8"))
    except Exception:
        return reply
    po_cache = _load_preorder_cache()

    def _qty(code: str) -> int:
        d = avail.get(code)
        if d is None:
            return 0
        return d.get("available", 0) if isinstance(d, dict) else d

    kept_lines = []
    dropped = []
    for line in reply.split("\n"):
        codes = [c.upper() for c in _PCR.findall(line)]
        if not codes:
            kept_lines.append(line)
            continue
        if "預購" in line:
            kept_lines.append(line)
            continue
        in_stock_or_preorder = any(_qty(c) > 0 or c in po_cache for c in codes)
        if in_stock_or_preorder:
            kept_lines.append(line)
        else:
            dropped.append(line.strip())

    if dropped:
        print(f"[recommend] 過濾掉缺貨推薦 {len(dropped)} 行：{dropped}", flush=True)

    cleaned = "\n".join(kept_lines)
    # 若刪到只剩抬頭/結語，補一句友善提示
    if dropped and not _PCR.findall(cleaned):
        cleaned += "\n\n（剛才挑的幾款庫存不太夠，幫您再確認看看～）"
    return cleaned


def _execute_claude_command(user_id: str, cmd: dict, line_api, original_text: str = "") -> str | None:
    """執行 Claude 指令引擎回傳的結構化指令"""
    action = cmd.get("action")

    if action == "add_cart":
        code = cmd.get("code", "").upper()
        qty = int(cmd.get("qty", 1))
        note = cmd.get("note", "")
        if not code:
            return None
        from services.ecount import ecount_client as _ec_cmd
        item = _ec_cmd.get_product_cache_item(code)
        name = (item.get("name") if item else "") or code
        # 模糊描述（客戶沒報貨號）→ 先確認再加購物車
        _has_code = bool(_PROD_CODE_RE.search(original_text)) if original_text else False
        if not _has_code:
            from storage.state import state_manager as _sm_ac
            _sm_ac.set(user_id, {
                "action":    "awaiting_product_confirm",
                "prod_cd":   code,
                "prod_name": name,
                "qty":       qty,
                "note":      note,
            })
            _price = f"　${int(item['price'])}" if item and item.get("price") and item["price"] > 0 else ""
            _confirm_text = f"請問是這款嗎？\n\n「{name}」（{code}）{_price}\n× {qty} 個\n\n確認的話跟我說「是」，不是的話跟我說「不是」唷"
            _confirm_imgs = _get_product_image_urls([code], max_images=1)
            print(f"[claude-cmd] add_cart 模糊 → 先確認: {code} x{qty}", flush=True)
            if _confirm_imgs:
                return (_confirm_text, _confirm_imgs)
            return _confirm_text
        from storage import cart as _cart_cmd
        _cart_cmd.add_item(user_id, code, name, qty, note=note)
        print(f"[claude-cmd] add_cart: {code} x{qty} note={note!r}", flush=True)
        return tone.cart_item_added(_cart_cmd.get_cart(user_id))

    elif action == "ask_quantity":
        code = cmd.get("code", "").upper()
        if not code:
            return None
        from services.ecount import ecount_client as _ec_cmd2
        item = _ec_cmd2.get_product_cache_item(code)
        name = (item.get("name") if item else "") or code
        from storage.state import state_manager as _sm_cmd
        _sm_cmd.set(user_id, {
            "action": "awaiting_quantity",
            "prod_cd": code,
            "prod_name": name,
        })
        print(f"[claude-cmd] ask_quantity: {code}", flush=True)
        _aq_text = tone.ask_quantity(name)
        # 模糊描述（客戶沒報貨號）→ 附產品照片讓客戶確認是否是這款
        _has_code = bool(_PROD_CODE_RE.search(original_text)) if original_text else False
        if not _has_code:
            _aq_imgs = _get_product_image_urls([code], max_images=1)
            if _aq_imgs:
                return (_aq_text, _aq_imgs)
        return _aq_text

    elif action == "checkout":
        from storage import cart as _cart_co
        if _cart_co.is_empty(user_id):
            return None
        print(f"[claude-cmd] checkout", flush=True)
        return handle_checkout(user_id, line_api)

    elif action == "reply":
        reply_text = cmd.get("text", "")
        if reply_text:
            # 從回覆中提取貨號，若客戶問「現貨」且 Claude 把缺貨品也列進來 → 從回覆中移除並重新組回覆
            _raw_codes = _PROD_CODE_RE.findall(reply_text)
            _raw_codes = list(dict.fromkeys(c.upper() for c in _raw_codes))
            _xianhuo_kw = any(k in (original_text or "") for k in ["現貨", "有貨", "有什麼", "什麼款", "哪些", "哪款"])
            if _xianhuo_kw and _raw_codes:
                from services.ecount import ecount_client as _ec_chk
                _oos = []
                for _c in _raw_codes:
                    _it = _ec_chk.lookup(_c)
                    _q = (_it or {}).get("qty")
                    if _q is None or _q <= 0:
                        _oos.append(_c)
                if _oos:
                    # 逐行過濾：若該行含缺貨貨號就整行刪掉
                    _kept_lines = []
                    for _ln in reply_text.split("\n"):
                        if any(_c in _ln.upper() for _c in _oos):
                            continue
                        _kept_lines.append(_ln)
                    reply_text = "\n".join(_kept_lines)
                    print(f"[claude-cmd] reply 過濾缺貨 {_oos}，保留 {[c for c in _raw_codes if c not in _oos]}", flush=True)
            print(f"[claude-cmd] reply: {reply_text[:30]!r}", flush=True)
            # 從（過濾後的）回覆中提取貨號 → 附產品圖 + 設 awaiting_quantity
            _reply_codes = _PROD_CODE_RE.findall(reply_text)
            _reply_codes = list(dict.fromkeys(c.upper() for c in _reply_codes))
            if len(_reply_codes) == 1:
                _rc = _reply_codes[0]
                from services.ecount import ecount_client as _ec_reply
                _r_item = _ec_reply.get_product_cache_item(_rc)
                _r_name = (_r_item.get("name") if _r_item else None) or _rc
                from storage.state import state_manager as _sm_reply
                _sm_reply.set(user_id, {
                    "action":    "awaiting_quantity",
                    "prod_cd":   _rc,
                    "prod_name": _r_name,
                })
                print(f"[claude-cmd] reply 設 awaiting_quantity: {_rc}", flush=True)
            elif len(_reply_codes) > 1:
                # 多產品清單 → 記住貨號，客戶問照片時可以發（offset=4 因為 reply 已附前 4 張）
                from storage.state import state_manager as _sm_multi
                _sm_multi.set(user_id, {
                    "action":       "recent_products",
                    "prod_codes":   _reply_codes[:16],
                    "photo_offset": 4,
                })
                print(f"[claude-cmd] reply 多產品，記 recent_products: {_reply_codes[:16]}", flush=True)
            if _reply_codes:
                _reply_imgs = _get_product_image_urls(_reply_codes, max_images=4 if len(_reply_codes) > 1 else 1)
                if _reply_imgs:
                    return (reply_text, _reply_imgs)
            return reply_text
        return None

    elif action == "escalate":
        reason = cmd.get("reason", "")
        issue_store.add(user_id, "claude_escalate", f"Claude 轉真人：{reason}")
        print(f"[claude-cmd] escalate: {reason}", flush=True)
        from handlers.hours import _is_open_now, next_open_reply
        from datetime import datetime as _dt_esc
        import pytz as _pz_esc
        _now_esc = _dt_esc.now(_pz_esc.timezone(settings.BUSINESS_TZ))
        if _is_open_now(_now_esc):
            return "您的問題我已經登記起來了喔～我幫您確認一下，稍後回覆您"
        return "您的問題我已經登記起來了喔～" + next_open_reply()

    return None


def _dispatch(
    user_id: str, text: str, intent: Intent, line_api: MessagingApi
) -> str:
    """根據意圖分發到對應 handler"""
    # ── 秒殺擋下：訊息含已標記沒貨的貨號 → 直接擋（一律優先）──
    from storage import sold_out as _so_d
    _d_codes = _PROD_CODE_RE.findall(text)
    _d_blocked = [c.upper() for c in _d_codes if _so_d.is_sold_out(c.upper())]
    if _d_blocked:
        print(f"[sold-out] dispatch 擋下 {_d_blocked} user={user_id[:10]}...", flush=True)
        return tone.sold_out_secret_kill()

    # ── 全局攔截：貨號 + 照片關鍵字 → 直接發照片 ──
    _photo_dispatch_kw = ["照片", "圖片", "看圖", "有圖", "看一下"]
    if any(kw in text for kw in _photo_dispatch_kw):
        _pd_codes = _PROD_CODE_RE.findall(text)
        if _pd_codes:
            _pd_code = _pd_codes[0].upper()
            _pd_imgs = _get_product_image_urls([_pd_code], max_images=1)
            from services.ecount import ecount_client as _ec_pd
            _pd_item = _ec_pd.get_product_cache_item(_pd_code)
            _pd_name = (_pd_item.get("name") if _pd_item else None) or _pd_code
            from storage.state import state_manager as _sm_pd
            _sm_pd.set(user_id, {
                "action":    "awaiting_quantity",
                "prod_cd":   _pd_code,
                "prod_name": _pd_name,
            })
            print(f"[dispatch] 貨號+照片 → 發照片: {_pd_code}", flush=True)
            if _pd_imgs:
                return (f"這是「{_pd_name}」的照片唷～\n請問要幾個呢？", _pd_imgs)
            return f"「{_pd_name}」目前沒有照片{tone.suffix_light()} 請問要幾個呢？"

    # ── 改數量：「貨號改N」→ 直接更新購物車該品項 ──
    try:
        from handlers.internal import _PROD_CODE_PAT as _CHG_CODE_PAT
        _CHG_RE = re.compile(rf'({_CHG_CODE_PAT})\s*改\s*(\d+)')
        _chg_matches = list(_CHG_RE.finditer(text))
        if _chg_matches:
            from storage import cart as _cart_chg2
            from services.ecount import ecount_client as _ec_chg2
            _chg_cart = _cart_chg2.get_cart(user_id)
            if _chg_cart:
                _chg_any = False
                for _cm in _chg_matches:
                    _ccode = _cm.group(1).strip().upper()
                    _cqty  = int(_cm.group(2))
                    _exist = next((it for it in _chg_cart if (it.get("prod_cd","").upper() == _ccode)), None)
                    if _exist:
                        _cart_chg2.set_item(user_id, _ccode, _exist["prod_name"], _cqty,
                                            note=_exist.get("note","") or "")
                        _chg_any = True
                    else:
                        _citem = _ec_chg2.get_product_cache_item(_ccode)
                        _cname = (_citem.get("name") if _citem else "") or _ccode
                        _cart_chg2.add_item(user_id, _ccode, _cname, _cqty)
                        _chg_any = True
                if _chg_any:
                    from storage.state import state_manager as _sm_chg2
                    _sm_chg2.clear(user_id)
                    print(f"[dispatch] 改數量: {[(m.group(1), m.group(2)) for m in _chg_matches]}", flush=True)
                    return tone.cart_item_added(_cart_chg2.get_cart(user_id))
    except Exception as _chg_e:
        print(f"[dispatch] 改數量解析失敗: {_chg_e}", flush=True)

    # ── 批次下單解析：多行 / 單行含多個「貨號*數量」→ 直接加購物車 ──
    # 支援：
    #   Z3590*2
    #   S0632*2
    #   Z3592*6 (3色各2)   ← 括號或尾段當備註
    try:
        from handlers.internal import _STAFF_ORDER_ITEM_RE as _BATCH_RE, _parse_qty as _batch_parse_qty
        from handlers.ordering import resolve_unit as _batch_resolve_unit
        _batch_items = []
        for _bl in text.splitlines():
            _bm = _BATCH_RE.search(_bl)
            if not _bm:
                continue
            _bcode = _bm.group(1).strip().upper()
            _bqty  = _batch_parse_qty(_bm.group(2))
            _bunit = _bm.group(3) if _bm.lastindex and _bm.lastindex >= 3 else None
            _brest = _bl[_bm.end():].strip()
            # 剝括號（全形半形都處理）
            _brest = re.sub(r'^[\(（]\s*|\s*[\)）]\s*$', '', _brest).strip()
            _brest = re.sub(r'^備[註誌记]\s*[:：]?\s*', '', _brest).strip()
            _bactual_cd, _bactual_qty, _bwarn = _batch_resolve_unit(_bcode, _bqty, _bunit)
            _batch_items.append((_bactual_cd, _bactual_qty, _brest, _bwarn))

        if _batch_items:
            from storage import cart as _cart_batch
            from services.ecount import ecount_client as _ec_batch
            _warnings = []
            for _bcd, _bq, _bnote, _bw in _batch_items:
                _bitem = _ec_batch.get_product_cache_item(_bcd)
                _bname = (_bitem.get("name") if _bitem else "") or _bcd
                _cart_batch.add_item(user_id, _bcd, _bname, _bq, note=_bnote)
                if _bw:
                    _warnings.append(_bw)
            print(f"[dispatch] 批次下單: {[(c,q,n) for c,q,n,_ in _batch_items]}", flush=True)
            from storage.state import state_manager as _sm_batch
            _sm_batch.clear(user_id)
            _reply_batch = tone.cart_item_added(_cart_batch.get_cart(user_id))
            if _warnings:
                _reply_batch += "\n" + "\n".join(_warnings)
            return _reply_batch
    except Exception as _batch_e:
        print(f"[dispatch] 批次下單解析失敗: {_batch_e}", flush=True)

    if intent == Intent.RECOMMENDATION:
        return _handle_recommendation(user_id, text, line_api)
    elif intent == Intent.INVENTORY:
        return handle_inventory(user_id, text, line_api)
    elif intent == Intent.PRICE:
        return handle_price(user_id, text)
    elif intent == Intent.ORDER_TRACKING:
        # 先看購物車有沒有未送出的
        from storage import cart as _cart_ot
        _ot_cart = _cart_ot.get_cart(user_id)
        if _ot_cart:
            lines = ["目前還沒送出的訂單："]
            for item in _ot_cart:
                lines.append(f"  • {item['prod_name']} × {item['qty']}")
            lines.append("\n還有其他要訂的嗎？如果好了就跟我說幫你送出喔")
            return "\n".join(lines)
        # 查已建立的訂單（未備貨 + 已備貨未取）
        _cust = customer_store.get_by_line_id(user_id)
        _cust_name = (_cust.get("real_name") or _cust.get("display_name") or "").strip() if _cust else ""
        if _cust_name:
            _ot_reply = handle_internal_customer_orders(
                f"{_cust_name}訂單", as_customer_reply=True,
            )
            if _ot_reply:
                return _ot_reply
        return handle_order_tracking(user_id, text)
    elif intent == Intent.DELIVERY:
        return handle_delivery(user_id, text)
    elif intent == Intent.BUSINESS_HOURS:
        return handle_business_hours(text)
    elif intent == Intent.GREETING:
        return tone.greeting_reply()
    elif intent == Intent.CREDIT_CARD:
        return "抱歉我們沒有刷卡喔"
    elif intent == Intent.BANK_ACCOUNT:
        return f"匯款資訊如下：\n{tone._get_bank_info()}"
    elif intent == Intent.CONFIRMATION:
        # 純短確認詞 + cart 非空 → 視為送出（bot 剛問「好了就送出」客戶回「好」的情境）
        from storage import cart as _cart_conf
        if text.strip() in set(AFFIRMATIVE_KEYWORDS) and not _cart_conf.is_empty(user_id):
            print(f"[confirmation] 純確認詞+cart非空 → 視為送出: {text.strip()!r}", flush=True)
            return handle_checkout(user_id, line_api)
        # 只有真正的感謝語才回覆，純確認詞（好的/了解/收到）靜默不回
        _THANKS_KW = ["謝謝", "感謝", "感恩", "辛苦了", "謝啦", "謝了", "多謝"]
        if any(kw in text for kw in _THANKS_KW):
            return tone.confirmation_ack()
        return None  # 好的/了解/收到/OK → 靜默
    # ── 台型查詢：「巨無霸有什麼」「中巨有哪些」─────────────────────
    _machine_type = detect_machine_query(text)
    if _machine_type:
        return handle_machine_query(user_id, _machine_type, line_api)
    # ── 新場景 ─────────────────────────────────────
    elif intent == Intent.BARGAINING:
        return handle_bargaining(user_id, text)
    elif intent == Intent.SPEC:
        return handle_spec(user_id, text, line_api)
    elif intent == Intent.RETURN:
        return handle_return(user_id, text, line_api)
    elif intent == Intent.MULTI_PRODUCT:
        return handle_multi_product(user_id, text)
    elif intent == Intent.ADDRESS_CHANGE:
        return handle_address_change(user_id, text, line_api)
    elif intent == Intent.ADDRESS_QUERY:
        return tone.address_query()
    elif intent == Intent.COMPLAINT:
        # 購物車有東西 + 「不對」「錯了」→ 清購物車 + 靜默待處理
        from storage import cart as _cart_complaint
        if not _cart_complaint.is_empty(user_id):
            _cart_complaint.clear_cart(user_id)
            issue_store.add(user_id, "complaint", f"（購物車品項有誤）{text}")
            return "問題已記錄，稍等下唷～"
        return handle_complaint(user_id, text, line_api)
    elif intent == Intent.ORDER_CHANGE:
        issue_store.add(user_id, "order_change", text)
        print(f"[dispatch] 改單/取消 → 進待處理 user={user_id[:10]}...: {text!r}")
        return "稍等一下喔"
    elif intent == Intent.URGENT_ORDER:
        return handle_urgent_order(user_id, text, line_api)
    elif intent == Intent.NOTIFY_REQUEST:
        return handle_notify_request(user_id, text, line_api)
    elif intent == Intent.MACHINE_SIZE:
        # 含貨號 → 可能是客戶貼 PO 文問庫存，不當台型查詢
        if _PROD_CODE_RE.search(text):
            return handle_inventory(user_id, text, line_api)
        # 嘗試識別「需要中巨貨」「K霸 300 內」這類推薦請求 → 主動列現貨
        from handlers.service import extract_machine_type_loose, extract_budget, handle_machine_recommend
        _mt = extract_machine_type_loose(text)
        if _mt:
            _budget = extract_budget(text)
            _reply = handle_machine_recommend(user_id, _mt, _budget, line_api)
            if _reply:
                # 同步登 issue 讓真人也能看到（但不阻塞回覆）
                issue_store.add(user_id, "machine_size", f"{text}（已自動列{_mt}{f' {_budget}內' if _budget else ''}現貨）")
                return _reply
        # 看不出台型 → 登 issue + 短 ack（避免客戶以為訊息沒到）
        issue_store.add(user_id, "machine_size", text)
        return "收到～小編幫您看一下適合的，馬上來"
    elif intent == Intent.VISIT_STORE:
        # 優先用 real_name（客戶提供的真實姓名），否則用 LINE 顯示名稱
        _cust = customer_store.get_by_line_id(user_id)
        _display = (_cust.get("real_name") if _cust else None) or ""
        if not _display:
            _prof = _get_profile_cached(line_api, user_id)
            _display = _prof.display_name if _prof else ""
        return handle_visit(user_id, text, _display)
    elif intent == Intent.CHECKOUT:
        from storage import cart as cart_store
        if cart_store.is_empty(user_id):
            # 購物車空的 → 當一般確認語處理
            return tone.confirmation_ack()
        return handle_checkout(user_id, line_api)
    else:
        # 購物車有東西 → 判斷客戶意圖
        from storage import cart as _cart_chk
        if not _cart_chk.is_empty(user_id):
            # 客戶要看照片（指定貨號 or 購物車最後一個）
            _photo_cart_kw = ["照片", "圖片", "看圖", "有圖", "看一下照片", "看一下圖"]
            if any(kw in text for kw in _photo_cart_kw):
                # 優先：客戶指定了貨號 → 發指定的照片
                _specified_codes = _PROD_CODE_RE.findall(text)
                if _specified_codes:
                    _sc = _specified_codes[0].upper()
                    _sc_imgs = _get_product_image_urls([_sc], max_images=1)
                    from services.ecount import ecount_client as _ec_photo
                    _sc_item = _ec_photo.get_product_cache_item(_sc)
                    _sc_name = (_sc_item.get("name") if _sc_item else None) or _sc
                    # 設 awaiting_quantity，客戶回數量就能直接下單
                    from storage.state import state_manager as _sm_photo
                    _sm_photo.set(user_id, {
                        "action":    "awaiting_quantity",
                        "prod_cd":   _sc,
                        "prod_name": _sc_name,
                    })
                    if _sc_imgs:
                        return (f"這是「{_sc_name}」的照片唷～\n請問要幾個呢？", _sc_imgs)
                    return f"「{_sc_name}」目前沒有照片{tone.suffix_light()} 請問要幾個呢？"
                # 沒指定貨號 → 發購物車最後一個
                _cart_items = _cart_chk.get_cart(user_id)
                _cart_codes = [it["prod_cd"] for it in _cart_items if it.get("prod_cd")]
                if _cart_codes:
                    _cart_imgs = _get_product_image_urls(_cart_codes[-1:], max_images=1)
                    if _cart_imgs:
                        _last_name = _cart_items[-1].get("prod_name", "")
                        return (f"這是「{_last_name}」的照片唷～", _cart_imgs)
            _cancel_kw = ["不用", "算了", "取消", "不要"]
            if any(kw in text for kw in _cancel_kw):
                _cart_chk.clear_cart(user_id)
                return "好的，已取消訂單～有需要再找我哦"
            # 純數量（三個、10個）→ 可能要追加，不自動結帳
            from handlers.ordering import extract_quantity as _eq_cart
            if _eq_cart(text):
                return f"請問要訂什麼商品呢？或是跟我說「好了」幫您送出目前的訂單唷"
            # 「謝謝」「麻煩了」等禮貌語 + 購物車有東西 → 結帳確認
            _polite_kw = ["謝謝", "感謝", "麻煩了", "辛苦", "先這樣", "就這些", "就先這樣", "就好"]
            if any(k in text for k in _polite_kw):
                return handle_checkout(user_id, line_api)
            # 備註性質的訊息 → 存到最後一個品項的備註
            _note_kw = ["備註", "幫我", "請幫", "分配", "混裝", "平均",
                        "顏色", "款式", "不要黑", "不要白", "不要紅", "換色"]
            if any(k in text for k in _note_kw) and not any(k in text for k in _cancel_kw):
                _cart_chk.set_note(user_id, text.strip())
                return f"好的，已備註：{text.strip()}"
        # 先問 Claude 指令引擎，再 fallback 到文字回覆
        from services.claude_ai import ask_claude_command, ask_claude_text
        _cmd = ask_claude_command(text, user_id=user_id)
        if _cmd:
            _cmd_result = _execute_claude_command(user_id, _cmd, line_api, original_text=text)
            if _cmd_result:
                return _cmd_result

        # Claude 指令引擎沒回應 → fallback 到文字回覆
        _claude_reply = ask_claude_text(text, user_id=user_id)
        if _claude_reply:
            # Claude 回覆裡的產品代碼
            _claude_codes_raw = _PROD_CODE_RE.findall(_claude_reply)
            _claude_codes = list(dict.fromkeys(c.upper() for c in _claude_codes_raw))
            if len(_claude_codes) == 1:
                _cc = _claude_codes[0]
                from services.ecount import ecount_client as _ec_cl
                _cl_item = _ec_cl.get_product_cache_item(_cc)
                _cl_name = (_cl_item.get("name") if _cl_item else None) or _cc
                from storage.state import state_manager as _sm_cl
                _sm_cl.set(user_id, {
                    "action":    "awaiting_quantity",
                    "prod_cd":   _cc,
                    "prod_name": _cl_name,
                })
                print(f"[claude-ai] 設 awaiting_quantity: {_cc} ({_cl_name})", flush=True)
            elif len(_claude_codes) > 1:
                from storage.state import state_manager as _sm_multi2
                _sm_multi2.set(user_id, {
                    "action":       "recent_products",
                    "prod_codes":   _claude_codes[:16],
                    "photo_offset": 4,
                })
                print(f"[claude-ai] 回覆含多個產品，記 recent_products: {_claude_codes[:16]}", flush=True)
            if _claude_codes:
                _img_urls = _get_product_image_urls(_claude_codes, max_images=4)
                if _img_urls:
                    return (_claude_reply, _img_urls)

            # 客戶語意含取消/延後 → 登記待處理
            _cancel_defer_kw = ["先給別人", "給其他客人", "先讓給", "不用留", "延後", "有空再",
                                "有時間再", "下次再", "先不要", "暫時不", "改天再", "先取消"]
            if any(k in text for k in _cancel_defer_kw):
                issue_store.add(user_id, "cancel_defer", f"客戶取消/延後：{text[:80]}")
                print(f"[claude-ai] 客戶取消/延後，記待處理: {text[:30]!r}", flush=True)

            # Claude 回覆「確認一下」「稍後回覆」→ 代表無法回答，嘗試從 LINE OA 補上下文重試
            _unsure_kw = ["確認一下", "稍後回覆", "幫您確認", "幫你確認", "稍等", "查一下", "幫您查", "幫你查", "沒有資料"]
            if any(k in _claude_reply for k in _unsure_kw):
                print(f"[claude-ai] 回覆含不確定語氣，嘗試從 LINE OA 補上下文: {text[:30]!r}", flush=True)
                # 嘗試從 LINE OA 抓最近對話補上下文
                _oa_retry_ok = False
                try:
                    _cust_oa = customer_store.get_by_line_id(user_id)
                    _cust_name_oa = (_cust_oa.get("chat_label") or _cust_oa.get("real_name") or _cust_oa.get("display_name") or "") if _cust_oa else ""
                    if _cust_name_oa:
                        from services.line_oa_chat import read_chat_sync
                        _oa_msgs = read_chat_sync(_cust_name_oa, max_messages=15)
                        if _oa_msgs:
                            _oa_context = "\n".join(
                                f"{'客服' if m['role'] == 'staff' else '客戶'}：{m['text']}"
                                for m in _oa_msgs[-10:]
                            )
                            _retry_prompt = f"【LINE 官方帳號完整對話紀錄】\n{_oa_context}\n\n---\n客戶最新訊息：{text}"
                            _retry_reply = ask_claude_text(_retry_prompt, user_id=user_id)
                            if _retry_reply and not any(k in _retry_reply for k in _unsure_kw):
                                print(f"[claude-ai] LINE OA 上下文重試成功", flush=True)
                                _claude_reply = _retry_reply
                                _oa_retry_ok = True
                except Exception as _e_oa:
                    print(f"[claude-ai] LINE OA 上下文重試失敗: {_e_oa}", flush=True)

                if not _oa_retry_ok:
                    issue_store.add(user_id, "claude_unsure", f"Claude 無法回答：{text[:50]}")
                    print(f"[claude-ai] 重試仍無法回答，記待處理", flush=True)
                    # 休息時間 → 改回覆下次上班時間
                    from handlers.hours import _is_open_now, next_open_reply
                    from datetime import datetime as _dt_cl
                    import pytz as _pz_cl
                    from config import settings as _cfg_cl
                    _now_cl = _dt_cl.now(_pz_cl.timezone(_cfg_cl.BUSINESS_TZ))
                    if not _is_open_now(_now_cl):
                        _claude_reply = next_open_reply()
            return _claude_reply
        return handle_unknown(user_id, text, line_api)


if __name__ == "__main__":
    import os as _os
    _os.environ["WATCHFILES_FORCE_POLLING"] = "true"
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True,
                reload_delay=1.5,
                reload_excludes=["data", "*.log", "截圖用", "static", ".claude"])
