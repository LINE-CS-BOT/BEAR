"""
自動從 Ecount 訂貨單出貨處理同步未處理訂單 → data/unfulfilled_orders.json

執行方式：
  python scripts/sync_unfulfilled.py

改用 Excel 下載方式，比爬 HTML 表格更準確完整。
"""

import asyncio
import json
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).parent.parent
OUTPUT_PATH = ROOT / "data" / "unfulfilled_orders.json"
UNCLAIMED_PATH = ROOT / "data" / "unclaimed_orders.json"
_TMP_XLSX = ROOT / "data" / "_tmp_unfulfilled.xlsx"
_TMP_UNCLAIMED_XLSX = ROOT / "data" / "_tmp_unclaimed.xlsx"

sys.path.insert(0, str(ROOT))
from scripts._chrome_helper import (
    launch_chrome_if_needed,
    connect_get_page,
    ensure_logged_in,
    ERP_URL,
)

_ORDER_HASH = (
    "#menuType=MENUTAB_000004"
    "&menuSeq=836HEBN5VVQ8A2M"
    "&groupSeq=MENUTREE_000030"
    "&prgId=E040230&depth=3&version=V3"
    "&menuUrl=%2FECERP%2FSVC%2FESD%2FESD030M"
    "&isFavMenu=Y"
)

_PAGE_LOADED_SEL = "#tabIng"


async def _download_excel(page, tmp_path: Path) -> bool:
    """點頁面上的 Excel 按鈕下載檔案"""
    excel_selectors = [
        '[data-cid="Excel"]',
        '[data-cid="excel"]',
        '[data-cid*="Excel"]',
        'button:has-text("Excel")',
        '.btn:has-text("Excel")',
    ]
    excel_btn = None
    for sel in excel_selectors:
        loc = page.locator(sel)
        if await loc.count() > 0:
            excel_btn = loc.first
            break

    if excel_btn is None:
        print("[unfulfilled] ✗ 找不到 Excel 按鈕")
        return False

    try:
        async with page.expect_download(timeout=15000) as dl_info:
            await excel_btn.click(timeout=5000)
        dl = await dl_info.value
        await dl.save_as(str(tmp_path))
        if tmp_path.stat().st_size == 0:
            print("[unfulfilled] ✗ Excel 下載為空檔")
            return False
        print(f"[unfulfilled] ✓ Excel 已下載 → {tmp_path.name}")
        return True
    except Exception as e:
        # 嘗試下拉選單
        await asyncio.sleep(0.8)
        dropdown = page.locator(
            'li:has-text("Excel"), a:has-text("Excel"), '
            '[data-cid*="xls"], [data-cid*="excel"], [data-cid*="Excel"]'
        ).first
        try:
            async with page.expect_download(timeout=15000) as dl_info:
                await dropdown.click(timeout=5000)
            dl = await dl_info.value
            await dl.save_as(str(tmp_path))
            if tmp_path.stat().st_size == 0:
                return False
            print(f"[unfulfilled] ✓ Excel 已下載 → {tmp_path.name}")
            return True
        except Exception as e2:
            print(f"[unfulfilled] ✗ Excel 下載失敗: {e2}")
            return False


def _parse_unfulfilled_excel(xlsx_path: Path) -> list[dict]:
    """解析未備貨 Excel，回傳 [{code, name, customer, qty, date_no, delivery_date, note}, ...]"""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    # 動態尋找表頭列
    hdrs = None
    header_row_idx = 1
    _HDR_KEYWORDS = ["品項", "編碼", "客戶", "餘量", "日期", "交付", "名稱"]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        vals = [str(v or "").strip() for v in row]
        joined = " ".join(vals)
        if sum(1 for kw in _HDR_KEYWORDS if kw in joined) >= 2:
            hdrs = vals
            header_row_idx = i + 1
            break

    if not hdrs:
        print(f"[unfulfilled] ✗ 找不到表頭列")
        print(f"[unfulfilled]   前 5 列: {[list(r) for r in ws.iter_rows(min_row=1, max_row=5, values_only=True)]}")
        return []

    print(f"[unfulfilled] 表頭（第{header_row_idx}列）: {hdrs}")

    # 建立欄位索引
    col_map = {}
    _ALIASES = {
        "code":          ["品項編碼", "品號", "貨號", "編碼", "Prod"],
        "name":          ["品項名稱", "品名", "名稱"],
        "customer":      ["客戶名稱", "客戶", "Cust"],
        "qty":           ["餘量", "數量", "Qty"],
        "date_no":       ["日期-號碼", "日期", "Date"],
        "delivery_date": ["交付日期", "交付", "Delivery"],
        "note":          ["摘要", "備註", "Note", "Remark"],
    }
    for key, aliases in _ALIASES.items():
        for i, h in enumerate(hdrs):
            if any(a in h for a in aliases):
                col_map[key] = i
                break

    print(f"[unfulfilled] 欄位對應: {col_map}")

    if "code" not in col_map:
        print("[unfulfilled] ✗ 找不到品項編碼欄位")
        return []

    # 解析資料
    results = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        vals = list(row)
        code = str(vals[col_map["code"]] or "").strip() if "code" in col_map and col_map["code"] < len(vals) else ""
        if not code:
            continue

        def _get(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(vals):
                return ""
            return str(vals[idx] or "").strip()

        def _get_num(key):
            v = _get(key)
            try:
                return float(v.replace(",", "")) if v else 0
            except ValueError:
                return 0

        results.append({
            "code":          code,
            "name":          _get("name"),
            "customer":      _get("customer"),
            "qty":           _get_num("qty"),
            "date_no":       _get("date_no"),
            "delivery_date": _get("delivery_date"),
            "note":          _get("note"),
        })

    return results


def _parse_unclaimed_excel(xlsx_path: Path) -> list[dict]:
    """解析未取訂單 Excel"""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    hdrs = None
    header_row_idx = 1
    _HDR_KEYWORDS = ["客戶", "品項", "日期", "數量", "銷貨"]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        vals = [str(v or "").strip() for v in row]
        joined = " ".join(vals)
        if sum(1 for kw in _HDR_KEYWORDS if kw in joined) >= 2:
            hdrs = vals
            header_row_idx = i + 1
            break

    if not hdrs:
        return []

    col_map = {}
    _ALIASES = {
        "date_no":  ["日期-號碼", "日期", "Date"],
        "customer": ["客戶名稱", "客戶", "Cust"],
        "product":  ["品項名稱", "品名", "品項"],
        "qty":      ["數量", "餘量", "Qty"],
    }
    for key, aliases in _ALIASES.items():
        for i, h in enumerate(hdrs):
            if any(a in h for a in aliases):
                col_map[key] = i
                break

    results = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        vals = list(row)
        def _get(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(vals):
                return ""
            return str(vals[idx] or "").strip()
        def _get_num(key):
            v = _get(key)
            try:
                return float(v.replace(",", "")) if v else 0
            except ValueError:
                return 0

        customer = _get("customer")
        if not customer:
            continue
        results.append({
            "date_no":  _get("date_no"),
            "customer": customer,
            "product":  _get("product"),
            "qty":      _get_num("qty"),
        })

    return results


async def sync_unfulfilled():
    """同步未處理訂單資料（Excel 下載方式）"""
    from playwright.async_api import async_playwright

    launch_chrome_if_needed()
    async with async_playwright() as p:
        browser, page = await connect_get_page(p)
        if not page:
            print("[unfulfilled] ✗ 無法連接 Chrome")
            return False

        ec_sid = await ensure_logged_in(page)
        if not ec_sid:
            print("[unfulfilled] ✗ 未登入 Ecount")
            return False

        # 1. 導航到訂貨單出貨處理
        url = f"{ERP_URL}?w_flag=1&ec_req_sid={ec_sid}{_ORDER_HASH}"
        print("[unfulfilled] 導航訂貨單出貨處理...")
        try:
            await page.goto("about:blank", timeout=5000)
            await page.goto(url, timeout=20000)
            await page.wait_for_load_state("networkidle", timeout=12000)
            await page.wait_for_timeout(3000)
        except Exception as e:
            print(f"[unfulfilled] ✗ 導航失敗: {e}")
            return False

        if not await page.query_selector(_PAGE_LOADED_SEL):
            print("[unfulfilled] ✗ 頁面未載入")
            return False

        # 2. 點擊「未處理」tab
        try:
            await page.locator("a#tabIng").click(timeout=5000)
            print("[unfulfilled] ✓ 已點擊「未處理」tab")
            await page.wait_for_timeout(5000)
        except Exception as e:
            print(f"[unfulfilled] ✗ 點擊未處理 tab 失敗: {e}")
            return False

        # 3. 下載 Excel
        if not await _download_excel(page, _TMP_XLSX):
            return False

    # 4. 解析 Excel
    data = _parse_unfulfilled_excel(_TMP_XLSX)

    if data:
        OUTPUT_PATH.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"[unfulfilled] ✓ 未處理訂單已存：{len(data)} 筆")
        print(f"[unfulfilled]   → {OUTPUT_PATH}")
        for d in data[:5]:
            print(f"  {d['code']:10s} {d['name'][:20]:20s} {d['customer']:10s} x{d['qty']:g}")
        return True
    else:
        print("[unfulfilled] ✗ Excel 解析無資料")
        return False


async def sync_unclaimed():
    """同步未取訂單（已到貨但銷貨單為「輸入」）— Excel 下載方式"""
    from playwright.async_api import async_playwright

    launch_chrome_if_needed()
    async with async_playwright() as p:
        browser, page = await connect_get_page(p)
        if not page:
            print("[unclaimed] ✗ 無法連接 Chrome")
            return False

        ec_sid = await ensure_logged_in(page)
        if not ec_sid:
            print("[unclaimed] ✗ 未登入 Ecount")
            return False

        # 1. 導航到訂貨單出貨處理
        url = f"{ERP_URL}?w_flag=1&ec_req_sid={ec_sid}{_ORDER_HASH}"
        print("[unclaimed] 導航訂貨單出貨處理...")
        try:
            await page.goto("about:blank", timeout=5000)
            await page.goto(url, timeout=20000)
            await page.wait_for_load_state("networkidle", timeout=12000)
            await page.wait_for_timeout(3000)
        except Exception as e:
            print(f"[unclaimed] ✗ 導航失敗: {e}")
            return False

        # 2. 點擊「已到貨」tab
        try:
            await page.locator("a#tabConfirm").click(timeout=5000)
            print("[unclaimed] ✓ 已點擊「已到貨」tab")
            await page.wait_for_timeout(3000)
        except Exception as e:
            print(f"[unclaimed] ✗ 點擊已到貨 tab 失敗: {e}")
            return False

        # 3. 下載 Excel
        if not await _download_excel(page, _TMP_UNCLAIMED_XLSX):
            # 沒有資料也是正常的
            UNCLAIMED_PATH.write_text("[]", encoding="utf-8")
            print("[unclaimed] 目前沒有未取訂單")
            return True

    # 4. 解析 Excel
    data = _parse_unclaimed_excel(_TMP_UNCLAIMED_XLSX)

    if data:
        UNCLAIMED_PATH.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"[unclaimed] ✓ 未取訂單已存：{len(data)} 筆")
        for d in data[:5]:
            print(f"  {d['customer']:15s} {d['product'][:25]:25s} x{d['qty']}")
        return True
    else:
        UNCLAIMED_PATH.write_text("[]", encoding="utf-8")
        print("[unclaimed] 目前沒有未取訂單")
        return True


def main():
    asyncio.run(sync_unfulfilled())
    asyncio.run(sync_unclaimed())


if __name__ == "__main__":
    main()
