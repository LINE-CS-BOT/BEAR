"""
完全自動化：從 Ecount 庫存情況頁同步可售庫存 → data/available.json

執行方式：
  python scripts/auto_sync_unfulfilled.py

導航策略（依序）：
  1. Chrome 已在庫存情況頁（#searchGroup 存在）→ reload 清空表單
  2. goto ERP 首頁（等同點左上 ECOUNT logo）→ 點頂部導航「庫存情況」
  3. 嘗試候選 prgId 清單 → about:blank + goto 逐一試
  4. 全失敗 → 提示手動導航一次（之後永遠自動）
"""

import asyncio
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT        = Path(__file__).parent.parent
OUTPUT_PATH = ROOT / "data" / "available.json"

# 庫存情況的候選 prgId（Ecount IB 區）
_INVENTORY_PRGIDS = [
    "E040701",          # 已確認（小蠻牛帳號）
    "BA017B", "BA017",  "BI000001", "INV001",
    "STK001", "INVBAL", "BA00001",  "BI001",
]

# 選單中「庫存情況」的文字清單（多語系備選）
_INVENTORY_MENU_TEXTS = ["庫存情況", "재고현황", "Inventory Balance"]

sys.path.insert(0, str(ROOT))
from scripts._chrome_helper import (
    launch_chrome_if_needed,
    connect_get_page,
    ensure_logged_in,
    load_web_config,
    save_web_config,
    ERP_URL,
)


# ---------------------------------------------------------------------------
# 庫存情況頁面導航
# ---------------------------------------------------------------------------

def _capture_inventory_hash(page) -> str | None:
    """從目前頁面 URL 取出 hash 部分（#...），儲存到 config"""
    url = page.url
    if "#" not in url:
        return None
    h = "#" + url.split("#", 1)[1]
    save_web_config({"inventory_hash": h})
    print(f"[sync] 已儲存庫存情況 URL hash: {h[:60]}")
    return h


async def _click_inventory_menu(page) -> bool:
    """
    在當前頁面嘗試點頂部導航列「庫存情況」，確認 #searchGroup 出現後回傳 True。
    適用於任何已登入的 ERP 頁面（包含供應表單等）。
    注意：SPA hash 導航不觸發 networkidle，改用 wait_for_selector 等目標元素。
    """
    for menu_text in _INVENTORY_MENU_TEXTS:
        selectors = [
            f'text="{menu_text}"',       # 精確比對
            f':text("{menu_text}")',      # contains 比對
            f'a:has-text("{menu_text}")',
        ]
        for sel in selectors:
            try:
                locs = await page.locator(sel).all()
                for loc in locs:
                    try:
                        if not await loc.is_visible():
                            continue
                        await loc.click(timeout=5000)
                        print(f"[sync] 點擊「{menu_text}」，等待 #searchGroup 出現...")
                        # SPA hash 導航不觸發 networkidle，直接等目標元素
                        await page.wait_for_selector("#searchGroup", timeout=15000)
                        print("[sync] ✓ #searchGroup 已出現")
                        return True
                    except Exception:
                        continue
            except Exception:
                continue
    return False


async def _navigate_to_inventory(page, ec_sid: str) -> bool:
    """
    導向庫存情況頁面，確認 #searchGroup 出現。
    策略 1 → 2 → 3 → 4 依序嘗試。
    """
    # ── 策略 1：使用已儲存的 inventory_hash（最快最可靠）────────────────
    cfg = load_web_config()
    saved_hash = cfg.get("inventory_hash", "")
    if saved_hash:
        url = f"{ERP_URL}?w_flag=1&ec_req_sid={ec_sid}{saved_hash}"
        print(f"[sync] 使用已儲存 hash 導航庫存情況...")
        try:
            await page.goto("about:blank", timeout=5000)
            await page.goto(url, timeout=20000)
            await page.wait_for_load_state("networkidle", timeout=12000)
            await page.wait_for_timeout(2000)
            if await page.query_selector("#searchGroup"):
                print("[sync] ✓ 已儲存 hash 導航成功")
                return True
            print("[sync] 已儲存 hash 失效，改用其他策略...")
        except Exception as e:
            print(f"[sync] 已儲存 hash 失敗: {e}")

    # ── 策略 2：回 ERP 首頁（等同點左上 ECOUNT logo）再點「庫存情況」 ─
    print("[sync] 回 ERP 首頁（ECOUNT logo）再點「庫存情況」...")
    try:
        url_home = f"{ERP_URL}?w_flag=1&ec_req_sid={ec_sid}"
        await page.goto(url_home, timeout=25000)
        await page.wait_for_load_state("networkidle", timeout=15000)
        await page.wait_for_timeout(2000)
    except Exception as e:
        print(f"[sync] ERP 首頁導向失敗: {e}")

    if await _click_inventory_menu(page):
        _capture_inventory_hash(page)
        return True

    # ── 策略 3：逐一嘗試候選 prgId（about:blank + goto 完整重載）─────
    print("[sync] 嘗試候選 prgId 清單...")
    for prgid in _INVENTORY_PRGIDS:
        url = (
            f"{ERP_URL}?w_flag=1&ec_req_sid={ec_sid}"
            f"#menuType=MENUTREE_000004&prgId={prgid}&depth=3"
        )
        print(f"[sync]   嘗試 prgId={prgid}...")
        try:
            await page.goto("about:blank", timeout=5000)
            await page.goto(url, timeout=20000)
            await page.wait_for_load_state("networkidle", timeout=12000)
            await page.wait_for_timeout(3000)
            if await page.query_selector("#searchGroup"):
                h = f"#menuType=MENUTREE_000004&prgId={prgid}&depth=3"
                save_web_config({"inventory_hash": h})
                print(f"[sync] ✓ prgId={prgid} 成功，已儲存")
                return True
        except Exception:
            pass

    # ── 策略 4：全失敗 ─────────────────────────────────────────────────
    print("[sync] ✗ 無法自動導航到庫存情況頁面")
    print("[sync]")
    print("[sync] 請做一次手動設定：")
    print("[sync]   1. 在 Chrome 中點開「庫存情況」頁面")
    print("[sync]   2. 重新執行此腳本（會自動捕捉頁面 URL 並永久儲存）")
    return False


# ---------------------------------------------------------------------------
# 填表 → 查詢 → 提取
# ---------------------------------------------------------------------------

async def _enable_include_zero(page) -> bool:
    """
    啟用「包含0」篩選條件（三次 retry，支援 native click + hover 兩種觸發方式）。
    步驟：
      1. 點 balQty_BETWEEN input → FN 按鈕從 hidden 變可見
      2. Hover FN 按鈕 → 子選單出現
      3. 點「包含0」
    """
    _INPUT = "input[data-cid='balQty_BETWEEN']"
    _FN    = "button[data-function='fn'][data-cid='balQty_BETWEEN']"

    for attempt in range(3):
        try:
            inp = page.locator(_INPUT).first

            # Step 1：native click 觸發 FN 按鈕（比 JS click 更能模擬真實互動）
            try:
                await inp.click(timeout=3000)
            except Exception:
                # fallback：JS click + focus
                await page.evaluate(
                    "() => { const el = document.querySelector(\"input[data-cid='balQty_BETWEEN']\");"
                    " if (el) { el.focus(); el.click(); } }"
                )

            await page.wait_for_timeout(800)  # 等待 FN 按鈕出現

            # Step 2：等 FN 按鈕變可見
            fn_btn = page.locator(_FN).first
            try:
                await fn_btn.wait_for(state="visible", timeout=4000)
            except Exception:
                # 嘗試 hover input 再等一次
                try:
                    await inp.hover(timeout=2000)
                    await page.wait_for_timeout(600)
                    await fn_btn.wait_for(state="visible", timeout=3000)
                except Exception:
                    if attempt < 2:
                        print(f"[sync] 包含0：第{attempt+1}次 FN 未出現，重試...")
                        await page.wait_for_timeout(1000)
                        continue
                    print("[sync] 包含0：FN 按鈕未出現，略過")
                    return False

            # Step 3：Hover FN 按鈕 → 子選單
            await fn_btn.hover()
            await page.wait_for_timeout(500)

            # Step 4：點「包括0」（Ecount 用字為「包括」而非「包含」）
            zero_loc = page.locator('.dropdown-menu-fn li:has-text("包括0")').first
            try:
                await zero_loc.wait_for(state="visible", timeout=3000)
                await zero_loc.click()
                print("[sync] ✓ 已選取「包括0」")
                await page.wait_for_timeout(300)
                return True
            except Exception:
                print("[sync] 包括0：選單項未出現，略過")
                return False

        except Exception as e:
            print(f"[sync] 包含0 第{attempt+1}次失敗: {e}")
            await page.wait_for_timeout(800)

    return False


def _parse_inventory_excel(xlsx_path: Path) -> dict[str, dict]:
    """解析庫存情況 Excel，回傳 {code: {incoming, unfilled, balance, available, preorder, unit_price}, ...}"""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    # 動態尋找表頭列
    hdrs = None
    header_row_idx = 1
    _HDR_KEYWORDS = ["品號", "貨號", "入庫", "結存", "可售", "未出"]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        vals = [str(v or "").strip() for v in row]
        joined = " ".join(vals)
        if sum(1 for kw in _HDR_KEYWORDS if kw in joined) >= 2:
            hdrs = vals
            header_row_idx = i + 1
            break

    if not hdrs:
        print(f"[sync] ✗ Excel 找不到表頭列")
        return {}

    print(f"[sync] Excel 表頭（第{header_row_idx}列）: {hdrs}")

    # 建立欄位索引
    col_map = {}
    _ALIASES = {
        "code":       ["品號", "貨號", "品項編碼", "Prod"],
        "incoming":   ["入庫量", "入庫", "未進貨", "Incoming"],
        "unfilled":   ["未出量", "未出庫", "未出貨", "未出", "Unfilled"],
        "balance":    ["結存量", "結存", "庫存數量", "Balance"],
        "available":  ["可售量", "可售", "可售庫存", "Available"],
        "preorder":   ["預購量", "預購", "可預購數量", "可預購", "Pre-order"],
        "unit_price": ["出庫單價", "單價", "Unit Price"],
    }
    for key, aliases in _ALIASES.items():
        for i, h in enumerate(hdrs):
            if any(a in h for a in aliases):
                col_map[key] = i
                break

    print(f"[sync] Excel 欄位對應: {col_map}")

    if "code" not in col_map:
        print("[sync] ✗ Excel 找不到品號欄位")
        return {}

    # 解析資料
    result = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        vals = list(row)
        code_idx = col_map["code"]
        if code_idx >= len(vals):
            continue
        code = str(vals[code_idx] or "").strip().upper()
        if not code or not re.match(r'^[A-Za-z0-9\-]+$', code):
            continue

        def _num(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(vals):
                return 0
            v = str(vals[idx] or "").strip().replace(",", "")
            try:
                return round(float(v)) if v else 0
            except ValueError:
                return 0

        def _float(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(vals):
                return 0.0
            v = str(vals[idx] or "").strip().replace(",", "")
            try:
                return float(v) if v else 0.0
            except ValueError:
                return 0.0

        entry = {
            "incoming":  _num("incoming"),
            "unfilled":  _num("unfilled"),
            "balance":   _num("balance"),
            "available": _num("available"),
            "preorder":  _num("preorder"),
        }
        if "unit_price" in col_map:
            entry["unit_price"] = _float("unit_price")
        result[code] = entry

    print(f"[sync] ✓ Excel 解析完成：{len(result)} 筆")
    return result


async def _fill_and_extract(page) -> dict[str, dict]:
    """填倉庫=101 → 啟用包含0 → 點查詢 → 下載 Excel 提取庫存數量"""

    # 填倉庫代碼（noneEvent class，用 JS 設值）
    try:
        await page.evaluate("""
            () => {
                const el = document.querySelector("input[placeholder='倉庫']");
                if (!el) return;
                el.value = '';
                el.dispatchEvent(new Event('input',  {bubbles: true}));
                el.value = '101';
                el.dispatchEvent(new Event('input',  {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                el.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true}));
                el.dispatchEvent(new Event('blur',   {bubbles: true}));
            }
        """)
        print("[sync] 倉庫 = 101")
        await page.wait_for_timeout(500)
    except Exception as e:
        print(f"[sync] 倉庫設值略過: {e}")

    # 啟用「包含0」
    await _enable_include_zero(page)

    # 點查詢按鈕（JS click 繞過可見度限制）
    try:
        clicked = await page.evaluate("""
            () => {
                const btns = document.querySelectorAll('#searchGroup');
                if (!btns.length) return false;
                btns[btns.length - 1].click();
                return true;
            }
        """)
        if clicked:
            print("[sync] 點擊查詢(F8)")
        else:
            print("[sync] ✗ 找不到 #searchGroup")
            return {}
    except Exception as e:
        print(f"[sync] ✗ 查詢按鈕失敗: {e}")
        return {}

    # 等待結果
    print("[sync] 等待結果...")
    try:
        await page.wait_for_selector("#grid-main tr:nth-child(2)", timeout=30000)
    except Exception as e:
        print(f"[sync] ✗ 等待結果逾時: {e}")
        return {}

    # ── 下載 Excel（比爬 HTML 更準確完整，最多重試 3 次）──────────────
    _TMP_XLSX = ROOT / "data" / "_tmp_inventory.xlsx"
    for _attempt in range(3):
        excel_selectors = [
            '[data-cid="Excel"]',
            '[data-cid="excel"]',
            '[data-cid*="Excel"]',
            'button:has-text("Excel")',
            '.btn:has-text("Excel")',
        ]
        excel_btn = None
        for sel in excel_selectors:
            loc = page.locator(sel)
            if await loc.count() > 0:
                excel_btn = loc.first
                break

        if not excel_btn:
            print(f"[sync] ⚠️ 第{_attempt+1}次：找不到 Excel 按鈕")
            if _attempt < 2:
                await page.wait_for_timeout(3000)
                continue
            break

        try:
            async with page.expect_download(timeout=30000) as dl_info:
                await excel_btn.click(timeout=5000)
            dl = await dl_info.value
            await dl.save_as(str(_TMP_XLSX))
            if _TMP_XLSX.stat().st_size > 0:
                print(f"[sync] ✓ Excel 已下載 → {_TMP_XLSX.name}（第{_attempt+1}次）")
                return _parse_inventory_excel(_TMP_XLSX)
            else:
                print(f"[sync] ⚠️ 第{_attempt+1}次：Excel 下載為空檔")
        except Exception as e:
            print(f"[sync] ⚠️ 第{_attempt+1}次 Excel 下載失敗: {e}")

        if _attempt < 2:
            await page.wait_for_timeout(3000)

    print("[sync] ✗ Excel 3 次下載都失敗，不更新庫存資料")
    return {}


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

async def run() -> dict[str, int]:
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser, page = await connect_get_page(p)

        # 確認登入
        ec_sid = await ensure_logged_in(page)
        if not ec_sid:
            await browser.close()
            return {}

        # 導航到庫存情況頁
        ok = await _navigate_to_inventory(page, ec_sid)
        if not ok:
            await browser.close()
            return {}

        # 填表 + 查詢 + 提取
        result = await _fill_and_extract(page)
        await browser.close()
        return result


def main():
    # 1. 自動啟動 Chrome
    if not launch_chrome_if_needed():
        print("[sync] ✗ Chrome 無法啟動")
        sys.exit(1)

    # 2. 執行同步
    try:
        result = asyncio.run(run())
    except Exception as e:
        print(f"[sync] ✗ 同步失敗: {e}")
        sys.exit(1)

    if not result:
        print("[sync] 未取得資料，保留現有 available.json")
        sys.exit(1)

    # 3. 驗證資料品質
    MIN_ENTRIES = 200  # 正常情況應有 900+ 筆，<200 視為異常
    count = len(result)
    if count < MIN_ENTRIES:
        print(f"[sync] ✗ 資料筆數異常（{count} < {MIN_ENTRIES}），保留現有 available.json")
        sys.exit(1)

    # 確認不是全部都是 0（代表 Chrome 可能在錯誤的頁面）
    nonzero = sum(
        1 for d in result.values()
        if any(d.get(k, 0) != 0 for k in ("available", "balance", "incoming", "unfilled", "preorder"))
    )
    if nonzero == 0:
        print(f"[sync] ✗ 所有 {count} 筆資料欄位均為 0，保留現有 available.json")
        sys.exit(1)
    print(f"[sync] ✓ 品質驗證通過：{count} 筆，其中 {nonzero} 筆有非零數值")

    # 4. 儲存
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[sync] {now}  更新 {len(result)} 筆 → {OUTPUT_PATH.name}")
    top = sorted(result.items(), key=lambda x: x[1].get("available", 0), reverse=True)[:5]
    has_price = any("unit_price" in d for d in result.values())
    for code, d in top:
        price_str = f"  出庫單價:{d.get('unit_price',0):>8.2f}" if has_price else ""
        print(f"  {code:<15} 可售:{d.get('available',0):>5}  可預購:{d.get('preorder',0):>5}{price_str}")

    # 庫存同步後更新預購清單
    try:
        from handlers.inventory import refresh_preorder_list
        refresh_preorder_list()
    except Exception as e:
        print(f"[sync] 更新預購清單失敗: {e}")


if __name__ == "__main__":
    main()
