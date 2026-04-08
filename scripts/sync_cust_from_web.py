"""
完全自動化：從 Ecount 客戶清單頁同步客戶資料 → customers.db

執行方式：
  python scripts/sync_cust_from_web.py            # 自動爬取 + 同步
  python scripts/sync_cust_from_web.py --dry-run  # 只顯示結果，不寫入
  python scripts/sync_cust_from_web.py --save-only # 只存 JSON，不同步 DB
  python scripts/sync_cust_from_web.py --from-file # 從已存 JSON 重新比對

導航策略（依序）：
  1. about:blank → ERP 主頁 SITE MAP → 點擊「客戶/供應商列表」連結（真實點擊）
  2. 全失敗 → 提示手動導航一次（之後永遠自動）

比對邏輯：
  - LINE 客戶有地址 → 先用地址模糊比對（路名 + 門牌號碼）
  - LINE 客戶無地址 → 用手機號碼比對
  - 同手機多個 Ecount 代碼（多地址）→ 全部存入 customer_ecount_codes
    （下單時 bot 會自動詢問客戶選擇地址）
"""

import re
import sys
import json
import argparse
import asyncio
from pathlib import Path
from datetime import datetime

# Windows 終端機強制 UTF-8
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from scripts._chrome_helper import (
    launch_chrome_if_needed,
    connect_get_page,
    ensure_logged_in,
    ERP_URL,
)

OUTPUT_JSON     = ROOT / "data" / "ecount_customers.json"
ECOUNT_CFG_JSON = ROOT / "data" / "ecount_web_config.json"  # 儲存已知的客戶清單 prgId

# Ecount 客戶清單的 prgId（多個備選，按優先順序嘗試）
_CUST_PRGIDS = [
    "A000035",   # 已知可用（首次有頭登入時有效）
    "ESA001M",   # 從頁面 data-ecpath 發現
    "G000001",   # 거래처 목록（通用）
    "GS00001",   # 거래처 검색
    "AR000001",  # 應收客戶
    "A000001",
    "AC00001",
    "A100001",
]


# ---------------------------------------------------------------------------
# 頁面導航 + 資料提取
# ---------------------------------------------------------------------------

_CUST_HEADER_KEYWORDS = {"編碼", "編號", "名稱", "客戶", "CUST", "BUSINESS"}
_TXN_HEADER_KEYWORDS  = {"日期", "金額", "交易", "單據", "會計", "帳"}


def _is_customer_list_headers(headers: list[str]) -> bool:
    """判斷表頭是否為客戶清單（含編碼/名稱）而非交易/訂單清單"""
    joined = " ".join(headers).upper()
    has_cust = any(k.upper() in joined for k in _CUST_HEADER_KEYWORDS)
    has_txn  = any(k in joined for k in _TXN_HEADER_KEYWORDS)
    return has_cust and not has_txn


async def _get_headers(page) -> list[str]:
    """讀取 #grid-main 表頭"""
    try:
        return await page.evaluate("""
            () => {
                const t = document.getElementById('grid-main');
                if (!t || !t.rows[0]) return [];
                return Array.from(t.rows[0].cells).map(c => (c.textContent || '').trim());
            }
        """)
    except Exception:
        return []


def _load_saved_prgid() -> str | None:
    """從 ecount_web_config.json 載入已確認的客戶清單 prgId"""
    try:
        if ECOUNT_CFG_JSON.exists():
            cfg = json.loads(ECOUNT_CFG_JSON.read_text(encoding="utf-8"))
            return cfg.get("cust_list_prgid")
    except Exception:
        pass
    return None


def _save_prgid(prgid: str):
    """儲存已確認的客戶清單 prgId 到 ecount_web_config.json"""
    try:
        cfg = {}
        if ECOUNT_CFG_JSON.exists():
            cfg = json.loads(ECOUNT_CFG_JSON.read_text(encoding="utf-8"))
        cfg["cust_list_prgid"] = prgid
        ECOUNT_CFG_JSON.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  [config] 儲存 prgId={prgid} → {ECOUNT_CFG_JSON.name}")
    except Exception as e:
        print(f"  [config] 儲存失敗: {e}")


async def _navigate_to_cust_list(page, ec_sid: str) -> bool:
    """
    導向客戶清單，策略依序：
    1. ERP 主頁（about:blank 完全重載）→ 點選 SITE MAP「客戶/供應商列表」連結（真實點擊）
    2. 全失敗 → 提示手動導航
    """
    # ── 策略 0：用已存的 prgId 直接導航 ──────────────────────────
    saved_id = _load_saved_prgid()
    if saved_id:
        try:
            direct_url = f"{ERP_URL}?w_flag=1&ec_req_sid={ec_sid}#prgId={saved_id}"
            print(f"  嘗試用已存 prgId={saved_id} 直接導航...")
            await page.goto("about:blank", timeout=5000)
            await page.goto(direct_url, timeout=25000)
            await page.wait_for_load_state("networkidle", timeout=15000)
            await page.wait_for_timeout(3000)
            if await page.query_selector("#grid-main"):
                headers = await _get_headers(page)
                if _is_customer_list_headers(headers):
                    print(f"  ✅ 用 prgId={saved_id} 直接導航成功")
                    return True
            print(f"  [warn] prgId={saved_id} 導航後找不到客戶表格，嘗試其他方式")
        except Exception as e:
            print(f"  [warn] prgId={saved_id} 導航失敗: {e}")

    # ── 策略 1：先導到 ERP 主頁（about:blank 確保 SPA 完全重載）──────
    try:
        url_base = f"{ERP_URL}?w_flag=1&ec_req_sid={ec_sid}"
        print("  [info] 載入 ERP 主頁...")
        await page.goto("about:blank", timeout=5000)
        await page.goto(url_base, timeout=25000)
        await page.wait_for_load_state("networkidle", timeout=15000)
        await page.wait_for_timeout(2000)
    except Exception as e:
        print(f"  [warn] ERP 主頁導向: {e}")

    # ── 策略 2：點選首頁 SITE MAP 中的客戶清單連結（真實點擊）──────
    _CUST_MENU_TEXTS = ["客戶/供應商列表", "거래처 목록", "Customer List", "客戶清單"]
    print("  嘗試點擊首頁選單「客戶/供應商列表」...")
    for menu_text in _CUST_MENU_TEXTS:
        try:
            locs = await page.locator(f'text="{menu_text}"').all()
            for loc in locs:
                if await loc.is_visible():
                    await loc.click(timeout=5000)
                    print(f"  點擊「{menu_text}」，等待載入...")
                    await page.wait_for_load_state("networkidle", timeout=20000)
                    await page.wait_for_timeout(2000)
                    if await page.query_selector("#grid-main"):
                        headers = await _get_headers(page)
                        if _is_customer_list_headers(headers):
                            m = re.search(r"prgId=([^&#]+)", page.url)
                            if m:
                                _save_prgid(m.group(1))
                            print(f"  ✅ 「{menu_text}」→ 客戶清單載入成功")
                            return True
                    break
        except Exception as e:
            print(f"  「{menu_text}」點擊失敗: {e}")

    # ── 策略 3：全失敗 ────────────────────────────────────────────────
    print("  ✗ 無法自動導航到客戶清單頁面")
    print("  請做一次手動設定：")
    print("    1. 在 Chrome 中點開「客戶/供應商列表」頁面")
    print("    2. 重新執行此腳本（會自動捕捉 prgId 並永久儲存）")
    return False


async def _click_query(page):
    """
    載入全部客戶資料：
    1. 優先：JS 強制點「查詢5,000筆以上」(#moreData) → 一次載入所有資料
    2. 備用：點一般「查詢」按鈕
    """
    # ── 優先：強制點 #moreData 載入全部（繞過 visibility 限制）──────────
    try:
        rows_before = await page.evaluate(
            "() => document.getElementById('grid-main')?.rows?.length || 0"
        )
        result = await page.evaluate("""
            () => {
                const btn = document.getElementById('moreData');
                if (!btn) return 'not_found';
                btn.style.display = 'inline-block';
                btn.style.visibility = 'visible';
                btn.click();
                return 'clicked';
            }
        """)
        if result == 'clicked':
            print("  [info] 強制點「查詢5,000筆以上」，等待全部資料載入...")
            await page.wait_for_load_state("networkidle", timeout=30000)
            try:
                await page.wait_for_selector("#grid-main tr:nth-child(2)", timeout=20000)
            except Exception:
                pass
            rows_after = await page.evaluate(
                "() => document.getElementById('grid-main')?.rows?.length || 0"
            )
            print(f"  [info] 載入前 {int(rows_before)-1} 筆 → 載入後 {int(rows_after)-1} 筆")
            return
        else:
            print("  [info] #moreData 不存在，改用一般查詢")
    except Exception as e:
        print(f"  [info] moreData 點擊失敗: {e}")

    # ── 備用：點一般查詢按鈕 ────────────────────────────────────────────
    try:
        btn = page.locator('button:has-text("查詢"):not(#moreData)')
        if await btn.count() > 0:
            await btn.first.click(timeout=5000)
            await page.wait_for_load_state("networkidle", timeout=15000)
            await page.wait_for_selector("#grid-main tr:nth-child(2)", timeout=15000)
    except Exception as e:
        print(f"  [info] 點查詢略過: {e}")


_PAGE_EXTRACT_JS = """
    () => {
        const table = document.getElementById('grid-main');
        if (!table) return { error: 'no #grid-main' };

        // 先讀表頭，找出各欄位置
        const headers = [];
        const headerRow = table.rows[0];
        if (headerRow) {
            for (let i = 0; i < headerRow.cells.length; i++) {
                headers.push((headerRow.cells[i].textContent || '').trim());
            }
        }

        // 尋找關鍵欄的 index
        const idxCode  = headers.findIndex(h => h.includes('編碼') || h.includes('編號') || h === 'CUST_CD' || h === 'BUSINESS_NO');
        const idxName  = headers.findIndex(h => h.includes('名稱') || h.includes('客戶名') || h === 'CUST_NAME');
        const idxPhone = headers.findIndex(h => h.includes('手機') || h.includes('HP'));
        const idxTel   = headers.findIndex(h => h.includes('電話') && !h.includes('手機'));
        const idxAddr  = headers.findIndex(h => h.includes('地址') || h.includes('住址') || h.toUpperCase().includes('ADDRESS'));

        if (idxCode < 0) {
            // fallback：假設固定欄位 col 1=編碼, col 2=名稱, col 4=手機, col 5=電話, col 6=地址
            const custs = [];
            for (let i = 1; i < table.rows.length; i++) {
                const row = table.rows[i];
                const code  = (row.cells[1]?.textContent || '').trim();
                const name  = (row.cells[2]?.textContent || '').trim();
                const phone = (row.cells[4]?.textContent || '').trim();
                const tel   = (row.cells[5]?.textContent || '').trim();
                const addr  = (row.cells[6]?.textContent || '').trim();
                if (code) custs.push({ code, name, phone, tel, addr });
            }
            return { customers: custs, fallback: true, headers };
        }

        const custs = [];
        for (let i = 1; i < table.rows.length; i++) {
            const row = table.rows[i];
            const code  = (row.cells[idxCode]?.textContent || '').trim();
            const name  = idxName  >= 0 ? (row.cells[idxName]?.textContent  || '').trim() : '';
            const phone = idxPhone >= 0 ? (row.cells[idxPhone]?.textContent || '').trim() : '';
            const tel   = idxTel   >= 0 ? (row.cells[idxTel]?.textContent   || '').trim() : '';
            const addr  = idxAddr  >= 0 ? (row.cells[idxAddr]?.textContent  || '').trim() : '';
            if (code) custs.push({ code, name, phone, tel, addr });
        }
        return { customers: custs, headers };
    }
"""


async def _extract_page(page) -> list[dict]:
    """提取目前頁面的客戶資料"""
    result = await page.evaluate(_PAGE_EXTRACT_JS)
    if "error" in result:
        print(f"  ❌ 表格提取失敗: {result['error']}")
        return []
    if result.get("fallback"):
        print(f"  ⚠️  用 fallback 欄位（表頭: {result.get('headers', [])}）")
    return result.get("customers", [])


async def _scan_pagination(page) -> list[dict]:
    """掃描頁面上所有可能的翻頁元素（用於診斷）"""
    return await page.evaluate("""
        () => {
            const els = document.querySelectorAll(
                '[data-cid], [class*="pag"], [id*="page"], [class*="next"], [id*="next"], ' +
                '[class*="prev"], [id*="prev"], button, a[onclick]'
            );
            const results = [];
            for (const el of els) {
                const txt = (el.textContent || '').trim().replace(/\\s+/g, ' ').slice(0, 20);
                const cid = el.getAttribute('data-cid') || '';
                const id  = el.id || '';
                const cls = el.className || '';
                // 只記錄看起來像翻頁的元素
                if (
                    cid.toLowerCase().includes('page') || cid.toLowerCase().includes('next') ||
                    cid.toLowerCase().includes('prev') || id.toLowerCase().includes('page') ||
                    id.toLowerCase().includes('next') || txt === '>' || txt === '>>' ||
                    txt === '다음' || txt === 'Next' || txt === '下一頁' || txt === '›' ||
                    cls.includes('next') || cls.includes('pag')
                ) {
                    results.push({
                        tag: el.tagName, id, cid, cls: cls.slice(0,40), txt,
                        visible: el.offsetParent !== null,
                        disabled: el.disabled || el.getAttribute('disabled') !== null
                    });
                }
            }
            return results;
        }
    """)


async def _click_next_page(page) -> bool:
    """
    點「下一頁」按鈕；成功翻頁回傳 True，無更多頁面回傳 False。
    先掃描頁面上的翻頁元素，再嘗試多種方式點擊。
    """
    # 取得目前第一列（翻頁後用來確認內容確實換了）
    first_row_before = await page.evaluate(
        "() => { const t = document.getElementById('grid-main'); "
        "return t && t.rows[1] ? t.rows[1].cells[1]?.textContent?.trim() : ''; }"
    )

    async def _verify_page_changed() -> bool:
        await page.wait_for_load_state("networkidle", timeout=15000)
        await asyncio.sleep(0.5)
        try:
            await page.wait_for_selector("#grid-main tr:nth-child(2)", timeout=10000)
        except Exception:
            pass
        first_row_after = await page.evaluate(
            "() => { const t = document.getElementById('grid-main'); "
            "return t && t.rows[1] ? t.rows[1].cells[1]?.textContent?.trim() : ''; }"
        )
        return first_row_after != first_row_before

    # ── 策略 0：Ecount 頁碼輸入框 [data-cid="paging"] ───────────────────
    try:
        paging_inp = page.locator('[data-cid="paging"]').first
        if await paging_inp.count() > 0 and await paging_inp.is_visible():
            cur_val = await paging_inp.input_value()
            cur_page = int(cur_val) if str(cur_val).strip().isdigit() else 1

            # 從頁面「N / 24」格式取總頁數
            total_str = await page.evaluate("""
                () => {
                    for (const el of document.querySelectorAll('*')) {
                        if (el.children.length > 0) continue;
                        const m = (el.textContent || '').match(/\\/\\s*(\\d+)/);
                        if (m && parseInt(m[1]) > 1) return m[1];
                    }
                    return '1';
                }
            """)
            total_pages = int(total_str) if str(total_str).strip().isdigit() else 9999
            print(f"  [info] 頁碼輸入：目前第 {cur_page} 頁 / 共 {total_pages} 頁")

            if cur_page >= total_pages:
                print("  [info] 已是最後一頁")
                return False

            next_page = cur_page + 1
            await paging_inp.triple_click()
            await paging_inp.fill(str(next_page))
            await paging_inp.press("Enter")
            if await _verify_page_changed():
                return True
            print("  [info] 頁碼輸入後內容未更新，改用其他方式...")
    except Exception as e:
        print(f"  [info] paging input 略過: {e}")

    # ── 策略 1：標準 CSS selector ───────────────────────────────────────
    selectors = [
        '[data-cid="pageNext"]', '[data-cid="nextPage"]', '[data-cid="next"]',
        '#pageNext', '.paging-next', '.btn-next',
        'button[title="다음"]', 'button[title="Next"]', 'button[title="下一頁"]',
        'a[title="下一頁"]',   'button:has-text("다음")', 'button:has-text("Next")',
    ]
    for sel in selectors:
        try:
            btn = page.locator(sel)
            if await btn.count() == 0:
                continue
            if await btn.first.is_disabled():
                return False
            await btn.first.click(timeout=5000)
            await page.wait_for_load_state("networkidle", timeout=15000)
            await page.wait_for_selector("#grid-main tr:nth-child(2)", timeout=10000)
            first_row_after = await page.evaluate(
                "() => { const t = document.getElementById('grid-main'); "
                "return t && t.rows[1] ? t.rows[1].cells[1]?.textContent?.trim() : ''; }"
            )
            if first_row_after != first_row_before:
                return True
        except Exception:
            continue

    # ── 策略 2：掃描頁面元素 + JS click ────────────────────────────────
    items = await _scan_pagination(page)
    if items:
        print(f"  [診斷] 找到 {len(items)} 個翻頁相關元素:")
        for it in items:
            print(f"    tag={it['tag']} id={it['id']} cid={it['cid']} txt={it['txt']!r} visible={it['visible']} disabled={it['disabled']}")

        # 找「下一頁」且未 disabled 的
        nxt = next(
            (x for x in items
             if not x['disabled'] and x['visible']
             and (x['cid'].lower() in ('pagenext','nextpage','next')
                  or x['id'].lower() in ('pagenext','next')
                  or x['txt'] in ('>', '›', '다음', 'Next', '下一頁', '>>'))
             ),
            None
        )
        if nxt:
            try:
                result = await page.evaluate(f"""
                    () => {{
                        let el = null;
                        if ({json.dumps(nxt['id'])})
                            el = document.getElementById({json.dumps(nxt['id'])});
                        if (!el && {json.dumps(nxt['cid'])})
                            el = document.querySelector('[data-cid={json.dumps(nxt['cid'])}]');
                        if (!el) return 'not_found';
                        el.click();
                        return 'clicked';
                    }}
                """)
                if result == 'clicked':
                    await page.wait_for_load_state("networkidle", timeout=15000)
                    first_row_after = await page.evaluate(
                        "() => { const t = document.getElementById('grid-main'); "
                        "return t && t.rows[1] ? t.rows[1].cells[1]?.textContent?.trim() : ''; }"
                    )
                    if first_row_after != first_row_before:
                        return True
            except Exception as e:
                print(f"  [info] JS 翻頁失敗: {e}")
    else:
        print("  [診斷] 找不到翻頁按鈕（可能 #moreData 已載入全部資料）")

    return False


async def _download_and_parse_excel(page) -> list[dict] | None:
    """
    點右下角 Excel 按鈕下載客戶/供應商列表（所有頁），解析 xlsx 回傳客戶清單。
    需要 openpyxl（pip install openpyxl）。
    """
    from openpyxl import load_workbook

    tmp_path = ROOT / "data" / "_tmp_ecount_cust.xlsx"

    # ── 掃描頁面所有按鈕（診斷用）────────────────────────────────────────
    btn_info = await page.evaluate("""
        () => {
            const btns = document.querySelectorAll('button, a.btn, [data-cid]');
            return Array.from(btns).map(b => ({
                tag: b.tagName,
                cid: b.getAttribute('data-cid') || '',
                txt: (b.textContent || '').trim().slice(0, 30),
                cls: (b.className || '').slice(0, 50),
                visible: b.offsetParent !== null,
            })).filter(b => b.txt || b.cid);
        }
    """)
    print(f"  [診斷] 頁面按鈕清單（共 {len(btn_info)} 個）：")
    for b in btn_info:
        print(f"    [{b['cid'] or '—'}] {b['txt']!r}  visible={b['visible']}  cls={b['cls'][:30]}")

    # 找 Excel 按鈕：多種 selector 嘗試
    excel_selectors = [
        '[data-cid="Excel"]',
        '[data-cid="excel"]',
        '[data-cid*="Excel"]',
        'button:has-text("Excel")',
        '.btn:has-text("Excel")',
        'button[class*="btn"]:has-text("Excel")',
    ]
    excel_btn = None
    for sel in excel_selectors:
        loc = page.locator(sel)
        if await loc.count() > 0:
            excel_btn = loc.first
            print(f"  [info] 找到 Excel 按鈕（selector: {sel}）")
            break

    if excel_btn is None:
        print("  [info] 找不到 Excel 按鈕，改用翻頁方式")
        return None

    print("  [info] 點 Excel 按鈕，下載全部資料（含所有頁）...")

    # 方案 A：直接點擊 → 捕捉下載
    try:
        async with page.expect_download(timeout=10000) as dl_info:
            await excel_btn.click(timeout=5000)
        dl = await dl_info.value
        await dl.save_as(str(tmp_path))
        if tmp_path.stat().st_size == 0:
            print("  [warn] Excel 下載為空檔（0 bytes），視為失敗")
            return None
        print(f"  [info] 已下載 → {tmp_path.name}")

    except Exception:
        # 方案 B：可能打開了 dropdown，找下拉選項
        print("  [info] 嘗試從 Excel 下拉選單選取...")
        await asyncio.sleep(0.8)
        dropdown = page.locator(
            'li:has-text("Excel"), a:has-text("Excel"), '
            '[data-cid*="xls"], [data-cid*="excel"], [data-cid*="Excel"]'
        ).first
        if await dropdown.count() > 0:
            try:
                async with page.expect_download(timeout=15000) as dl_info:
                    await dropdown.click(timeout=5000)
                dl = await dl_info.value
                await dl.save_as(str(tmp_path))
                if tmp_path.stat().st_size == 0:
                    print("  [warn] Excel 下載為空檔（0 bytes），視為失敗")
                    return None
                print(f"  [info] 已下載 → {tmp_path.name}")
            except Exception as e:
                print(f"  [info] Excel 下載失敗: {e}")
                return None
        else:
            print("  [info] 找不到 Excel 下拉選項，改用翻頁方式")
            return None

    # 解析 Excel
    try:
        wb = load_workbook(str(tmp_path), data_only=True)
        ws = wb.active

        # ── 動態尋找表頭列（Ecount Excel 第1行通常是標題如「公司名稱: XXX」）──
        hdrs = None
        header_row_idx = 1
        _HDR_KEYWORDS = ["編碼", "編號", "名稱", "CUST", "BUSINESS", "客戶", "手機", "地址", "電話"]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
            vals = [str(v or "").strip() for v in row]
            joined = " ".join(vals)
            if sum(1 for kw in _HDR_KEYWORDS if kw in joined) >= 2:
                hdrs = vals
                header_row_idx = i + 1  # 1-based
                break

        if not hdrs:
            # 最終 fallback：用第1行
            hdrs = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            header_row_idx = 1

        print(f"  [info] Excel 表頭 (第{header_row_idx}行): {hdrs}")

        def col(keywords):
            for kw in keywords:
                for i, h in enumerate(hdrs):
                    if kw in h:
                        return i
            return -1

        i_code  = col(["編碼", "編號", "CUST_CD", "BUSINESS_NO", "客戶/供應商編碼"])
        i_name  = col(["名稱", "CUST_NAME", "客戶/供應商名稱"])
        i_addr  = col(["地址", "住址", "ADDR"])
        i_phone = col(["手機", "HP", "HP_NO"])
        i_tel   = col(["電話", "TEL"])

        if i_code < 0:
            print(f"  [info] Excel 找不到「編碼」欄，無法解析（表頭: {hdrs}）")
            wb.close()
            return None

        def cell(row, i):
            return str(row[i] or "").strip() if 0 <= i < len(row) else ""

        custs = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            code = cell(row, i_code)
            if not code:
                continue
            custs.append({
                "code":  code,
                "name":  cell(row, i_name),
                "phone": cell(row, i_phone),
                "tel":   cell(row, i_tel),
                "addr":  cell(row, i_addr),
            })

        wb.close()
        print(f"  [info] Excel 解析完成：{len(custs)} 筆")
        return custs if custs else None   # 0筆→None，讓 fallback 翻頁

    except Exception as e:
        print(f"  [info] Excel 解析失敗: {e}")
        return None


async def _extract_customers(page) -> list[dict]:
    """
    從客戶清單頁抓取所有客戶。
    優先：Excel 下載（一次取得全部頁資料）
    備用：翻頁逐頁抓取
    """
    # ── Excel 下載（唯一方式，失敗重試 2 次）─────────────────────────────
    for attempt in range(3):
        custs = await _download_and_parse_excel(page)
        if custs:
            with_addr = sum(1 for c in custs if c.get("addr"))
            print(f"  共取得 {len(custs)} 筆客戶（其中 {with_addr} 筆有地址）")
            return custs
        if attempt < 2:
            print(f"  ⚠️ Excel 下載失敗，第 {attempt+2} 次重試...")
            await asyncio.sleep(3)

    print("  ✗ Excel 下載 3 次都失敗，無法同步客戶清單")
    return []


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

async def run() -> list[dict]:
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser, page = await connect_get_page(p)

        # 確認登入
        ec_sid = await ensure_logged_in(page)
        if not ec_sid:
            await browser.close()
            return []

        # 導航到客戶清單頁
        print("[web] 自動導向客戶清單頁面 ...")
        nav_ok = await _navigate_to_cust_list(page, ec_sid)
        if not nav_ok:
            # session 可能過期，重新登入再試一次
            print("[web] 導航失敗，嘗試重新登入...")
            await page.goto(ERP_URL, timeout=30000)
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            ec_sid2 = await ensure_logged_in(page)
            if ec_sid2 and ec_sid2 != ec_sid:
                print(f"[web] 取得新 session，重試導航...")
                nav_ok = await _navigate_to_cust_list(page, ec_sid2)
            if not nav_ok:
                print("[web] ✗ 無法自動導向到客戶清單頁面")
                print("[web]")
                print("[web] 請做一次手動設定：")
                print("[web]   1. 在 Chrome 中點開「客戶清單」頁面")
                print("[web]   2. 重新執行此腳本（會自動捕捉 prgId 並永久儲存）")
                await browser.close()
                return []

        # 提取客戶資料
        print("[web] ✓ 導向成功，開始提取客戶資料...")
        customers = await _extract_customers(page)
        await browser.close()
        return customers


# ---------------------------------------------------------------------------
# 同步邏輯
# ---------------------------------------------------------------------------

def _normalize(phone: str) -> str:
    return phone.replace(" ", "").replace("-", "").strip() if phone else ""


def _addr_key(addr: str) -> str:
    """提取地址關鍵部分（路/街名 + 門牌號碼），用於模糊比對"""
    addr = addr.replace(" ", "").replace("\u3000", "")
    m = re.search(
        r'([\u4e00-\u9fff]{1,8}[\u8def\u8857\u9053\u5df7\u5f04])'
        r'(?:[\u4e00-\u9fff\d\u6bb5]*?)(\d+)[\u865f\u53f7]?',
        addr
    )
    if m:
        return m.group(1) + m.group(2)
    return addr[:12] if len(addr) >= 4 else ""


def _addr_match(line_addr: str, ecount_addr: str) -> bool:
    """
    地址模糊比對：
    1. 去空白後互相包含 → True
    2. 提取路名+號碼比對 → True
    """
    if not line_addr or not ecount_addr:
        return False
    la = line_addr.replace(" ", "").replace("\u3000", "")
    ea = ecount_addr.replace(" ", "").replace("\u3000", "")
    if la in ea or ea in la:
        return True
    la_key = _addr_key(la)
    ea_key = _addr_key(ea)
    if la_key and ea_key and len(la_key) >= 4:
        return la_key in ea_key or ea_key in la_key
    return False


def _do_sync(ecount_custs: list[dict], dry_run: bool):
    from storage.customers import customer_store

    # ── 建立索引 ───────────────────────────────────────────
    # phone → list[{code, addr}]（同手機可能多筆，如多地址）
    phone_to_entries: dict[str, list[dict]] = {}
    # name → list[{code, addr}]（姓名 fallback）
    name_to_entries: dict[str, list[dict]] = {}
    # 所有有地址的條目（用於地址比對）
    addr_entries: list[dict] = []  # {code, addr}

    for c in ecount_custs:
        code = c["code"]
        addr = (c.get("addr") or "").strip()
        entry = {"code": code, "addr": addr}

        for ph in (_normalize(c.get("phone", "")), _normalize(c.get("tel", ""))):
            if ph:
                lst = phone_to_entries.setdefault(ph, [])
                if not any(e["code"] == code for e in lst):
                    lst.append(entry)

        if c.get("name"):
            lst = name_to_entries.setdefault(c["name"], [])
            if not any(e["code"] == code for e in lst):
                lst.append(entry)

        if addr:
            addr_entries.append(entry)

    line_custs = customer_store.all(limit=9999)
    print(f"\n比對 {len(line_custs)} 筆 LINE 客戶 vs {len(ecount_custs)} 筆 Ecount 客戶")
    print(f"  電話索引: {len(phone_to_entries)} 支 | 地址索引: {len(addr_entries)} 筆有地址")

    updated = skipped = unmatched = multi_addr = 0

    for c in line_custs:
        uid       = c.get("line_user_id") or ""
        db_id     = c.get("id")
        name      = (c.get("display_name") or "").strip()
        phone     = _normalize(c.get("phone") or "")
        line_addr = (c.get("address") or "").strip()
        old_cd    = c.get("ecount_cust_cd") or ""

        matched: list[dict] = []

        # ── 優先：LINE 有地址 → 用地址比對 ──────────────
        if line_addr:
            for ae in addr_entries:
                if _addr_match(line_addr, ae["addr"]):
                    if not any(e["code"] == ae["code"] for e in matched):
                        matched.append(ae)

        # ── 無地址或地址未命中 → 用手機比對 ─────────────
        if not matched and phone:
            matched = list(phone_to_entries.get(phone, []))

        # ── fallback：用姓名比對 ──────────────────────────
        if not matched and name:
            matched = list(name_to_entries.get(name, []))

        if not matched:
            unmatched += 1
            # 舊代碼若存在（如 L-XXXX）→ 清空，讓下單時走即時解析流程
            if old_cd:
                if dry_run:
                    print(f"  [清除] {name or uid}  {old_cd} → (空白)")
                else:
                    if uid:
                        customer_store.update_ecount_cust_cd(uid, "")
                    elif db_id:
                        customer_store.update_ecount_cust_cd_by_db_id(db_id, "")
                    print(f"  [清除] {name or uid}  {old_cd} → (空白)")
            continue

        default_code = matched[0]["code"]
        is_multi     = len(matched) > 1
        if is_multi:
            multi_addr += 1

        # 判斷是否需要更新
        needs_update = old_cd != default_code
        action = "覆蓋" if (needs_update and old_cd) else "新增"

        if needs_update:
            multi_tag = f" (+{len(matched) - 1}個地址)" if is_multi else ""
            print(f"  [{action}] {name or uid}  {old_cd or '(空)'} → {default_code}{multi_tag}")

        if not dry_run:
            # 更新 customers.ecount_cust_cd（預設代碼）
            if needs_update:
                if uid:
                    customer_store.update_ecount_cust_cd(uid, default_code)
                elif db_id:
                    customer_store.update_ecount_cust_cd_by_db_id(db_id, default_code)

            # 取得 DB id（upsert_ecount_code 需要）
            target_id = db_id
            if not target_id and uid:
                row = customer_store.get_by_line_id(uid)
                if row:
                    target_id = row["id"]

            # 儲存所有代碼 + 地址到 customer_ecount_codes
            if target_id:
                for entry in matched:
                    customer_store.upsert_ecount_code(
                        target_id, entry["code"], entry.get("addr", "")
                    )

        if needs_update:
            updated += 1
        else:
            skipped += 1

    print(f"\n比對結果：更新 {updated} | 已是最新 {skipped} | 無對應 {unmatched}")
    if multi_addr:
        print(f"  其中 {multi_addr} 筆客戶有多個地址/代碼（下單時將詢問客戶選擇）")


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="從 Ecount 網頁同步客戶清單")
    parser.add_argument("--dry-run",   action="store_true", help="不實際寫入 DB")
    parser.add_argument("--save-only", action="store_true", help="只存 JSON，不同步 DB")
    parser.add_argument("--from-file", action="store_true",
                        help=f"從 {OUTPUT_JSON} 讀取（跳過爬蟲）")
    parser.add_argument("--auto", action="store_true",
                        help="排程用（非互動模式，失敗時直接退出）")
    args = parser.parse_args()

    print("=" * 60)
    print("Ecount 客戶清單同步")
    print("=" * 60)

    if args.from_file:
        if not OUTPUT_JSON.exists():
            print(f"❌ 找不到 {OUTPUT_JSON}，請先執行一次（不加 --from-file）")
            sys.exit(1)
        ecount_custs = json.loads(OUTPUT_JSON.read_text(encoding="utf-8"))
        print(f"從 {OUTPUT_JSON} 讀取 {len(ecount_custs)} 筆")
    else:
        # 1. 自動啟動 Chrome
        if not launch_chrome_if_needed():
            print("[web] ✗ Chrome 無法啟動")
            sys.exit(1)

        # 2. 執行爬蟲
        try:
            ecount_custs = asyncio.run(run())
        except Exception as e:
            print(f"[web] ✗ 執行失敗: {e}")
            sys.exit(1)

        if not ecount_custs:
            print("❌ 未取得任何客戶資料，結束")
            sys.exit(1)

        # 存 JSON 以便 --from-file 重跑（防止覆蓋：新資料比舊的少 50% 以上就不存）
        OUTPUT_JSON.parent.mkdir(exist_ok=True)
        old_count = 0
        if OUTPUT_JSON.exists():
            try:
                old_count = len(json.loads(OUTPUT_JSON.read_text(encoding="utf-8")))
            except Exception:
                pass
        if old_count > 0 and len(ecount_custs) < old_count * 0.5:
            print(f"⚠️ 新資料 {len(ecount_custs)} 筆遠少於現有 {old_count} 筆，跳過覆蓋（可能抓取不完整）")
        else:
            OUTPUT_JSON.write_text(
                json.dumps(ecount_custs, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"[{ts}] 已存 {len(ecount_custs)} 筆 → {OUTPUT_JSON}")

    if args.save_only:
        print("--save-only 模式，跳過 DB 同步")
        return

    _do_sync(ecount_custs, dry_run=args.dry_run)

    if args.dry_run:
        print("\n【DRY RUN 模式，未實際寫入】")

    print("=" * 60)


if __name__ == "__main__":
    main()
